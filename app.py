import html as html_lib
import json
import os
import sqlite3
import traceback
import uuid
import io
import csv
from email import policy
from email.parser import BytesParser
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import parse_qs, quote
from wsgiref.simple_server import make_server
from readingtown.routes.auth import handle_auth_routes
from readingtown.routes.notifications import handle_notifications_routes
from readingtown.routes.logs import handle_logs_routes
from readingtown.routes.api import handle_api_routes
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "lms.db")
ROLE_OWNER = "owner"
ROLE_MANAGER = "manager"
ROLE_TEACHER = "teacher"
ROLE_PARENT = "parent"
ROLE_STUDENT = "student"
ROLE_LABELS = {
    ROLE_OWNER: "원장",
    ROLE_MANAGER: "매니저",
    ROLE_TEACHER: "강사",
    ROLE_PARENT: "학부모",
    ROLE_STUDENT: "학생",
}
CURRENT_LANG = "en"
LANG_LABELS = {"ko": "한국어", "en": "English", "zh": "中文"}
NAV_LABELS = {
    "dashboard": "menu.dashboard",
    "users": "menu.users",
    "students": "menu.students",
    "academics": "menu.academics",
    "masterdata": "menu.masterdata",
    "schedule": "menu.schedule",
    "attendance": "menu.attendance",
    "homework": "menu.homework",
    "exams": "menu.exams",
    "counseling": "menu.counseling",
    "payments": "menu.payments",
    "announcements": "menu.announcements",
    "library": "menu.library",
    "logs": "menu.logs",
    "logout": "common.logout",
    "login_as": "common.login_as",
    "lang": "common.language",
}
I18N_TEXTS = {
    "ko": {
        "menu.dashboard": "대시보드", "menu.users": "사용자", "menu.students": "학생관리", "menu.academics": "학사구조", "menu.masterdata": "마스터데이터", "menu.schedule": "스케줄",
        "menu.attendance": "출결", "menu.homework": "숙제", "menu.exams": "시험/성적", "menu.counseling": "상담/특이사항",
        "menu.payments": "수납", "menu.announcements": "공지/알림", "menu.library": "도서대출", "menu.logs": "시스템 로그",
        "common.login_as": "로그인", "common.language": "언어", "common.login": "로그인", "common.logout": "로그아웃",
        "common.save": "저장", "common.edit": "수정", "common.delete": "삭제", "common.search": "검색",
        "common.no_data": "데이터 없음", "common.selected": "선택됨", "common.forbidden": "접근 권한이 없습니다",
        "common.yes": "예", "common.no": "아니오", "common.prev": "이전", "common.next": "다음",
        "login.title": "LMS 로그인", "login.username": "아이디", "login.password": "비밀번호", "login.failed": "로그인 실패",
        "dash.title": "영어학원 LMS 대시보드", "dash.stats": "운영 현황", "dash.quick": "빠른 이동", "dash.users": "사용자", "dash.students": "학생", "dash.classes": "반", "dash.attendance": "출결 기록",
        "users.page_title": "학생/학부모/강사 관리(사용자 기반)", "users.add": "사용자 추가", "users.list": "사용자 목록", "users.saved": "사용자가 저장되었습니다.",
        "students.search": "검색", "students.list": "학생 목록", "common.reset": "초기화",
        "students.detail.title": "학생 상세", "students.detail.back": "학생 목록", "students.detail.edit": "수정",
        "students.detail.section.basic": "기본정보", "students.detail.section.attendance": "최근 출결", "students.detail.section.homework": "최근 숙제 제출/피드백",
        "students.detail.section.exams": "최근 시험/성적", "students.detail.section.counseling": "최근 상담 기록", "students.detail.section.payments": "최근 수납 기록", "students.detail.section.loans": "최근 도서 대출 기록",
        "students.field.student_no": "학생번호", "students.field.name_ko": "중국이름", "students.field.name_en": "영문이름", "students.field.phone": "연락처",
        "students.field.guardian_name": "보호자명", "students.field.guardian_phone": "보호자 연락처", "students.field.class": "현재 반", "students.field.credits": "남은 크레딧",
        "students.field.status": "상태", "students.field.enrolled_at": "입학일", "students.field.leave_period": "휴학 기간", "students.field.memo": "메모",
        "students.field.lesson_date": "날짜", "students.field.note": "메모", "students.field.homework": "숙제", "students.field.submitted": "제출여부", "students.field.submitted_at": "제출일",
        "students.field.feedback": "피드백", "students.field.exam_name": "시험명", "students.field.score": "점수", "students.field.exam_date": "시험일", "students.field.recorded_at": "기록일",
        "students.field.special_note": "특이사항", "students.field.paid_date": "결제일", "students.field.amount": "금액", "students.field.package_hours": "패키지시간",
        "students.field.remaining_classes": "잔여수업수", "students.field.code": "코드", "students.field.title": "제목", "students.field.loaned_at": "대출일", "students.field.returned_at": "반납일",
        "students.msg.saved": "저장되었습니다", "students.msg.pw_saved": "비밀번호가 변경되었습니다", "students.msg.no_user": "연결된 계정이 없습니다", "students.msg.empty_pw": "비밀번호를 입력하세요", "students.msg.validation_error": "입력값 검증에 실패했습니다. 필수값/날짜/참조값을 확인하세요.",
        "status.active": "정상", "status.leave": "휴학", "status.ended": "종료",
        "counseling.title": "상담 기록/학생 특이사항", "counseling.student_id": "학생ID", "counseling.parent_id": "학부모ID", "counseling.memo": "메모", "counseling.special": "특이사항", "counseling.list": "상담 기록",
        "picker.student": "학생 검색 선택", "picker.class": "반 검색 선택", "picker.teacher": "강사 검색 선택",
        "attendance.title": "출결 관리", "attendance.input": "출결 입력", "attendance.list": "출결 기록",
        "homework.title": "숙제 관리", "homework.add": "숙제 등록", "homework.input": "제출/피드백 입력", "homework.list": "숙제 목록", "homework.submissions": "제출/피드백 목록",
        "exams.title": "시험/성적 관리", "exams.add": "시험 등록", "exams.score_input": "점수 입력", "exams.list": "시험 목록", "exams.scores": "점수 목록",
        "payments.title": "수납 기록", "payments.input": "결제 입력", "payments.list": "최근 수납 기록",
        "ann.title": "공지/알림 구조", "ann.write": "공지 작성", "ann.list": "공지 목록", "ann.noti": "알림 데이터",
        "library.title": "도서 대출 관리", "library.book_add": "도서 등록", "library.loan": "대출 처리", "library.return": "반납 처리", "library.books": "도서 목록", "library.history": "대출/반납 이력",
        "field.class_id": "반ID", "field.student_id": "학생ID", "field.teacher_id": "강사ID", "field.date": "날짜", "field.status": "상태", "field.note": "메모", "field.title": "제목", "field.content": "내용", "field.type": "유형", "field.target": "대상", "field.data": "데이터", "field.writer": "작성자", "field.created_at": "생성일", "field.score": "점수",
        "status.present": "출석", "status.late": "지각", "status.absent": "결석", "status.makeup": "보강", "status.cancelled": "휴강",
    },
    "en": {
        "menu.dashboard": "Dashboard", "menu.users": "Users", "menu.students": "Students", "menu.academics": "Academics", "menu.masterdata": "Master Data", "menu.schedule": "Schedule",
        "menu.attendance": "Attendance", "menu.homework": "Homework", "menu.exams": "Exams/Scores", "menu.counseling": "Counseling",
        "menu.payments": "Payments", "menu.announcements": "Announcements", "menu.library": "Library", "menu.logs": "System Logs",
        "common.login_as": "Signed in", "common.language": "Language", "common.login": "Login", "common.logout": "Logout",
        "common.save": "Save", "common.edit": "Edit", "common.delete": "Delete", "common.search": "Search",
        "common.no_data": "No Data", "common.selected": "Selected", "common.forbidden": "Forbidden",
        "common.yes": "Yes", "common.no": "No", "common.prev": "Prev", "common.next": "Next",
        "login.title": "LMS Login", "login.username": "Username", "login.password": "Password", "login.failed": "Login failed",
        "dash.title": "ReadingTown LMS Dashboard", "dash.stats": "Operations Overview", "dash.quick": "Quick Links", "dash.users": "Users", "dash.students": "Students", "dash.classes": "Classes", "dash.attendance": "Attendance Records",
        "users.page_title": "User Management", "users.add": "Add User", "users.list": "User List", "users.saved": "User saved.",
        "students.search": "Search", "students.list": "Student List", "common.reset": "Reset",
        "students.detail.title": "Student Detail", "students.detail.back": "Student List", "students.detail.edit": "Edit",
        "students.detail.section.basic": "Basic Info", "students.detail.section.attendance": "Recent Attendance", "students.detail.section.homework": "Recent Homework Submissions/Feedback",
        "students.detail.section.exams": "Recent Exams/Scores", "students.detail.section.counseling": "Recent Counseling Records", "students.detail.section.payments": "Recent Payments", "students.detail.section.loans": "Recent Book Loans",
        "students.field.student_no": "Student No", "students.field.name_ko": "Chinese Name", "students.field.name_en": "English Name", "students.field.phone": "Phone",
        "students.field.guardian_name": "Guardian", "students.field.guardian_phone": "Guardian Phone", "students.field.class": "Class", "students.field.homeroom_teacher": "Homeroom Teacher", "students.field.credits": "Remaining Credits",
        "students.field.status": "Status", "students.field.enrolled_at": "Enrollment Date", "students.field.leave_period": "Leave Period", "students.field.memo": "Memo",
        "students.field.lesson_date": "Date", "students.field.note": "Note", "students.field.homework": "Homework", "students.field.submitted": "Submitted", "students.field.submitted_at": "Submitted At",
        "students.field.feedback": "Feedback", "students.field.exam_name": "Exam", "students.field.score": "Score", "students.field.exam_date": "Exam Date", "students.field.recorded_at": "Recorded At",
        "students.field.special_note": "Special Note", "students.field.paid_date": "Paid Date", "students.field.amount": "Amount", "students.field.package_hours": "Package Hours",
        "students.field.remaining_classes": "Remaining Classes", "students.field.code": "Code", "students.field.title": "Title", "students.field.loaned_at": "Loaned At", "students.field.returned_at": "Returned At",
        "students.msg.saved": "Saved successfully", "students.msg.pw_saved": "Password updated", "students.msg.no_user": "No linked user account", "students.msg.empty_pw": "Please enter a password", "students.msg.validation_error": "Validation failed. Check required fields, date formats, and references.",
        "status.active": "Active", "status.leave": "Leave", "status.ended": "Ended",
        "counseling.title": "Counseling / Student Notes", "counseling.student_id": "Student ID", "counseling.parent_id": "Parent ID", "counseling.memo": "Memo", "counseling.special": "Special Note", "counseling.list": "Counseling Records",
        "picker.student": "Student Picker", "picker.class": "Class Picker", "picker.teacher": "Teacher Picker",
        "attendance.title": "Attendance Management", "attendance.input": "Attendance Entry", "attendance.list": "Attendance Records",
        "homework.title": "Homework Management", "homework.add": "Add Homework", "homework.input": "Submission/Feedback Input", "homework.list": "Homework List", "homework.submissions": "Submission/Feedback List",
        "exams.title": "Exam/Score Management", "exams.add": "Add Exam", "exams.score_input": "Score Input", "exams.list": "Exam List", "exams.scores": "Score List",
        "payments.title": "Payment Records", "payments.input": "Payment Entry", "payments.list": "Recent Payments",
        "ann.title": "Announcements/Notifications", "ann.write": "Write Announcement", "ann.list": "Announcement List", "ann.noti": "Notification Data",
        "library.title": "Library Loan Management", "library.book_add": "Add Book", "library.loan": "Loan Process", "library.return": "Return Process", "library.books": "Book List", "library.history": "Loan/Return History",
        "field.class_id": "Class ID", "field.student_id": "Student ID", "field.teacher_id": "Teacher ID", "field.date": "Date", "field.status": "Status", "field.note": "Note", "field.title": "Title", "field.content": "Content", "field.type": "Type", "field.target": "Target", "field.data": "Data", "field.writer": "Writer", "field.created_at": "Created At", "field.score": "Score",
        "status.present": "Present", "status.late": "Late", "status.absent": "Absent", "status.makeup": "Makeup", "status.cancelled": "Cancelled",
    },
    "zh": {
        "menu.dashboard": "仪表盘", "menu.users": "用户", "menu.students": "学生管理", "menu.academics": "学术结构", "menu.masterdata": "主数据", "menu.schedule": "排课",
        "menu.attendance": "考勤", "menu.homework": "作业", "menu.exams": "考试/成绩", "menu.counseling": "咨询/备注",
        "menu.payments": "收费", "menu.announcements": "公告/通知", "menu.library": "图书借阅", "menu.logs": "系统日志",
        "common.login_as": "当前登录", "common.language": "语言", "common.login": "登录", "common.logout": "退出登录",
        "common.save": "保存", "common.edit": "编辑", "common.delete": "删除", "common.search": "搜索",
        "common.no_data": "无数据", "common.selected": "已选择", "common.forbidden": "无权限",
        "common.yes": "是", "common.no": "否", "common.prev": "上一页", "common.next": "下一页",
        "login.title": "LMS 登录", "login.username": "账号", "login.password": "密码", "login.failed": "登录失败",
        "dash.title": "ReadingTown LMS 仪表盘", "dash.stats": "运营概况", "dash.quick": "快捷入口", "dash.users": "用户", "dash.students": "学生", "dash.classes": "班级", "dash.attendance": "考勤记录",
        "users.page_title": "用户管理", "users.add": "新增用户", "users.list": "用户列表", "users.saved": "用户已保存。",
        "students.search": "搜索", "students.list": "学生列表", "common.reset": "重置",
        "students.detail.title": "学生详情", "students.detail.back": "学生列表", "students.detail.edit": "编辑",
        "students.detail.section.basic": "基本信息", "students.detail.section.attendance": "最近考勤", "students.detail.section.homework": "最近作业提交/反馈",
        "students.detail.section.exams": "最近考试/成绩", "students.detail.section.counseling": "最近咨询记录", "students.detail.section.payments": "最近缴费记录", "students.detail.section.loans": "最近图书借阅记录",
        "students.field.student_no": "学号", "students.field.name_ko": "中文名", "students.field.name_en": "英文名", "students.field.phone": "联系电话",
        "students.field.guardian_name": "监护人", "students.field.guardian_phone": "监护人电话", "students.field.class": "当前班级", "students.field.credits": "剩余学分",
        "students.field.status": "状态", "students.field.enrolled_at": "入学日", "students.field.leave_period": "休学期间", "students.field.memo": "备注",
        "students.field.lesson_date": "日期", "students.field.note": "备注", "students.field.homework": "作业", "students.field.submitted": "是否提交", "students.field.submitted_at": "提交日期",
        "students.field.feedback": "反馈", "students.field.exam_name": "考试名", "students.field.score": "分数", "students.field.exam_date": "考试日期", "students.field.recorded_at": "记录日",
        "students.field.special_note": "特殊事项", "students.field.paid_date": "支付日", "students.field.amount": "金额", "students.field.package_hours": "套餐课时",
        "students.field.remaining_classes": "剩余课次", "students.field.code": "编码", "students.field.title": "标题", "students.field.loaned_at": "借出日", "students.field.returned_at": "归还日",
        "students.msg.saved": "已保存", "students.msg.pw_saved": "密码已更新", "students.msg.no_user": "未关联用户账号", "students.msg.empty_pw": "请输入密码", "students.msg.validation_error": "校验失败，请检查必填项、日期格式与关联引用。",
        "status.active": "正常", "status.leave": "休学", "status.ended": "结束",
        "counseling.title": "咨询记录/学生备注", "counseling.student_id": "学生ID", "counseling.parent_id": "家长ID", "counseling.memo": "备注", "counseling.special": "特殊事项", "counseling.list": "咨询记录",
        "picker.student": "学生搜索选择", "picker.class": "班级搜索选择", "picker.teacher": "教师搜索选择",
        "attendance.title": "考勤管理", "attendance.input": "考勤录入", "attendance.list": "考勤记录",
        "homework.title": "作业管理", "homework.add": "作业登记", "homework.input": "提交/反馈录入", "homework.list": "作业列表", "homework.submissions": "提交/反馈列表",
        "exams.title": "考试/成绩管理", "exams.add": "考试登记", "exams.score_input": "成绩录入", "exams.list": "考试列表", "exams.scores": "成绩列表",
        "payments.title": "缴费记录", "payments.input": "缴费录入", "payments.list": "最近缴费记录",
        "ann.title": "公告/通知结构", "ann.write": "发布公告", "ann.list": "公告列表", "ann.noti": "通知数据",
        "library.title": "图书借阅管理", "library.book_add": "登记图书", "library.loan": "借出处理", "library.return": "归还处理", "library.books": "图书列表", "library.history": "借还记录",
        "field.class_id": "班级ID", "field.student_id": "学生ID", "field.teacher_id": "教师ID", "field.date": "日期", "field.status": "状态", "field.note": "备注", "field.title": "标题", "field.content": "内容", "field.type": "类型", "field.target": "对象", "field.data": "数据", "field.writer": "创建者", "field.created_at": "创建时间", "field.score": "分数",
        "status.present": "出勤", "status.late": "迟到", "status.absent": "缺勤", "status.makeup": "补课", "status.cancelled": "停课",
    },
}

I18N_TEXTS["ko"].update({
    "common.add": "추가", "role.owner": "원장", "role.manager": "매니저", "role.teacher": "강사", "role.parent": "학부모", "role.student": "학생",
    "login.default_accounts": "기본 계정: owner/1234, manager/1234, teacher/1234, parent/1234, student/1234",
    "field.id": "ID", "field.name": "이름", "field.role": "역할", "field.student": "학생", "field.submitted": "제출", "field.homework_title": "숙제명", "field.due_date": "마감일",
    "field.submission_count": "제출수", "field.total_targets": "총대상(등록수)", "field.avg_score": "평균점수", "field.score_entries": "입력건수",
    "students.field.leave_start_date": "휴학시작", "students.field.leave_end_date": "휴학종료", "students.password_reset": "학생 계정 비밀번호 변경", "students.new_password": "새 비밀번호",
    "academics.title": "코스/레벨/반/시간표 관리", "academics.register": "등록", "academics.course": "코스", "academics.level": "레벨", "academics.course_name": "코스명", "academics.level_name": "레벨명",
    "academics.class_name": "반명", "academics.class_list": "반 목록", "academics.course_id": "코스ID", "academics.level_id": "레벨ID", "academics.schedule_class_id": "시간표 반ID",
    "academics.day_of_week": "요일", "academics.start_time": "시작", "academics.end_time": "종료", "academics.teacher": "담당 강사", "academics.student_count": "학생 수", "academics.schedule": "시간표",
    "academics.class_detail.title": "반 상세", "academics.back_to_list": "학사구조 목록", "academics.basic_info": "기본정보", "academics.course_level": "코스/레벨", "academics.students": "소속 학생",
    "academics.sort": "정렬", "academics.sort_name": "이름 정렬", "academics.sort_student_no": "학생번호 정렬", "academics.sort_status": "상태 정렬", "academics.export_students_csv": "학생 CSV 내보내기", "academics.export_attendance_csv": "출결 CSV 내보내기",
    "academics.recent_attendance": "최근 출결", "academics.recent_homework": "최근 숙제", "academics.recent_exams": "최근 시험/성적",
    "forbidden.teacher_class_only": "담당 반만 처리할 수 있습니다", "forbidden.teacher_exam_only": "담당 시험만 처리할 수 있습니다",
    "library.not_found": "도서를 찾을 수 없습니다", "library.loan_done": "대여 완료", "server.start": "LMS 서버 실행",
    "academics.timetable_title": "시간표 관리", "academics.timetable_desc": "주간 수업 배치와 수업별 업무 이동을 관리합니다.", "academics.filter": "필터", "academics.week_prev": "이전 주", "academics.week_current": "현재 주", "academics.week_next": "다음 주",
    "academics.day_filter": "요일", "academics.teacher_filter": "강사", "academics.classroom_filter": "교실", "academics.course_level_filter": "코스/레벨", "academics.class_filter": "반 검색", "academics.search": "검색", "academics.add_lesson": "새 수업 추가",
    "academics.timetable": "주간 타임테이블", "academics.time_slot": "시간대", "academics.teacher_room": "강사/교실", "academics.students_summary": "학생", "academics.more_students": "+더보기", "academics.view_class": "반 상세", "academics.go_attendance": "출결", "academics.go_homework": "숙제", "academics.go_exams": "성적",
    "academics.lesson_detail": "수업 상세", "academics.edit_schedule": "수업 수정", "academics.schedule_form": "시간표 등록/수정", "academics.schedule_pick_class": "반 선택", "academics.schedule_pick_teacher": "강사 선택", "academics.schedule_teacher_auto": "반 담당 강사가 기본 선택됩니다", "academics.schedule_autofill": "선택한 반 정보를 자동으로 불러옵니다", "academics.schedule_pick_day": "요일 선택", "academics.schedule_pick_time": "시간 선택", "academics.schedule_pick_room": "교실 선택", "academics.schedule_pick_status": "상태 선택", "academics.schedule_pick_student": "학생", "academics.go_structure": "학사구조 관리로 이동", "academics.validation_class_required": "반을 먼저 선택하세요", "academics.validation_end_before_start": "종료 시간이 시작 시간보다 이를 수 없습니다", "academics.validation_conflict_class": "같은 반의 시간표가 겹칩니다", "academics.validation_conflict_teacher": "같은 강사의 시간표가 겹칩니다", "academics.validation_conflict_room": "같은 교실의 시간표가 겹칩니다", "academics.saved": "시간표가 저장되었습니다", "academics.updated": "시간표가 수정되었습니다", "academics.day_all": "전체", "academics.status": "상태", "academics.classroom": "교실", "academics.action.attendance_eval": "출결 및 평가", "lesson.record.page_title": "수업기록 입력", "lesson.record.header": "학생별 기록 입력", "lesson.record.desc": "시간표에서 선택한 수업의 학생 전체를 한 번에 기록합니다.", "lesson.record.class_info": "수업 정보", "lesson.record.back_schedule": "시간표로 돌아가기", "lesson.record.input_title": "출결 및 평가 입력", "lesson.record.empty_students": "이 반에는 등록된 학생이 없어 입력할 수 없습니다. 학생을 반에 배정한 후 다시 시도하세요.", "lesson.record.invalid_class": "class_id가 올바르지 않습니다.", "lesson.record.not_found_class": "존재하지 않는 반입니다.", "lesson.record.saved": "수업기록(출결/평가)이 저장되었습니다", "lesson.score.participation": "수업참여", "lesson.score.fluency": "유창성", "lesson.score.vocabulary": "어휘", "lesson.score.reading": "읽기", "lesson.score.homework": "숙제", "lesson.score.attitude": "수업태도", "lesson.score.teacher_memo": "교사 메모", "students.detail.section.evaluations": "최근 수업 평가기록", "students.eval.avg.title": "최근 평가 평균(최근 10회)", "students.eval.avg.participation": "평균 수업참여", "students.eval.avg.fluency": "평균 유창성", "students.eval.avg.vocabulary": "평균 어휘", "students.eval.avg.reading": "평균 읽기", "students.eval.avg.homework": "평균 숙제", "students.eval.avg.attitude": "평균 수업태도", "academics.copy_week.button": "이번 주를 다음 주로 복사", "academics.copy_week.desc": "현재 선택한 주간 시간표를 다음 주로 안전 복사합니다(중복은 건너뜀).", "academics.copy_week.source": "원본 주간", "academics.copy_week.target": "대상 주간", "academics.copy_week.result": "복사 결과", "academics.copy_week.none": "복사할 시간표가 없습니다", "academics.copy_week.done": "복사가 완료되었습니다", "academics.copy_week.copied": "복사", "academics.copy_week.skipped": "중복 건너뜀", "academics.week_label": "Week", "academics.week_range": "Week Range", "academics.ref_date": "Reference Date", "academics.selected_day": "선택 요일", "academics.selected_date": "선택 날짜", "attendance.manual_input_title": "수동 출결 입력(비상/관리자용 · 메인 운영동선 아님)", "lesson.record.weekday_mismatch": "선택한 날짜의 요일이 시간표 요일과 일치하지 않습니다", "lesson.record.no_students": "반에 소속된 학생이 없습니다", "lesson.record.score_range": "는 1~5 사이 정수여야 합니다", "validation.date": "YYYY-MM-DD 형식이어야 합니다", "validation.status_invalid": "status 값이 올바르지 않습니다", "users.teacher_type": "Teacher Type", "users.teacher_type.foreign": "외국교사", "users.teacher_type.chinese": "중국교사"
})
I18N_TEXTS["en"].update({
    "common.add": "Add", "role.owner": "Owner", "role.manager": "Manager", "role.teacher": "Teacher", "role.parent": "Parent", "role.student": "Student",
    "login.default_accounts": "Default accounts: owner/1234, manager/1234, teacher/1234, parent/1234, student/1234",
    "field.id": "ID", "field.name": "Name", "field.role": "Role", "field.student": "Student", "field.submitted": "Submitted", "field.homework_title": "Homework", "field.due_date": "Due Date",
    "field.submission_count": "Submissions", "field.total_targets": "Total Targets", "field.avg_score": "Average Score", "field.score_entries": "Score Entries",
    "students.field.leave_start_date": "Leave Start", "students.field.leave_end_date": "Leave End", "students.password_reset": "Reset Student Password", "students.new_password": "New Password",
    "academics.title": "Course/Level/Class/Schedule", "academics.register": "Register", "academics.course": "Course", "academics.level": "Level", "academics.course_name": "Course Name", "academics.level_name": "Level Name",
    "academics.class_name": "Class Name", "academics.class_list": "Class List", "academics.course_id": "Course ID", "academics.level_id": "Level ID", "academics.schedule_class_id": "Schedule Class ID",
    "academics.day_of_week": "Day", "academics.start_time": "Start", "academics.end_time": "End", "academics.teacher": "Teacher", "academics.student_count": "Students", "academics.schedule": "Schedule",
    "academics.class_detail.title": "Class Detail", "academics.back_to_list": "Academics List", "academics.basic_info": "Basic Info", "academics.course_level": "Course/Level", "academics.students": "Class Students",
    "academics.sort": "Sort", "academics.sort_name": "Sort by Name", "academics.sort_student_no": "Sort by Student No", "academics.sort_status": "Sort by Status", "academics.export_students_csv": "Export Students CSV", "academics.export_attendance_csv": "Export Attendance CSV",
    "academics.recent_attendance": "Recent Attendance", "academics.recent_homework": "Recent Homework", "academics.recent_exams": "Recent Exams/Scores",
    "forbidden.teacher_class_only": "Teacher can only access assigned classes", "forbidden.teacher_exam_only": "Teacher can only access assigned exams",
    "library.not_found": "Book not found", "library.loan_done": "Loan completed", "server.start": "LMS server running",
    "academics.timetable_title": "Timetable Management", "academics.timetable_desc": "Manage weekly lesson placement and quick jumps to class operations.", "academics.filter": "Filters", "academics.week_prev": "Previous Week", "academics.week_current": "Current Week", "academics.week_next": "Next Week",
    "academics.day_filter": "Day", "academics.teacher_filter": "Teacher", "academics.classroom_filter": "Classroom", "academics.course_level_filter": "Course/Level", "academics.class_filter": "Class Search", "academics.search": "Search", "academics.add_lesson": "Add Lesson",
    "academics.timetable": "Weekly Timetable", "academics.time_slot": "Time Slots", "academics.teacher_room": "Teacher/Room", "academics.students_summary": "Students", "academics.more_students": "+more", "academics.view_class": "Class Detail", "academics.go_attendance": "Attendance", "academics.go_homework": "Homework", "academics.go_exams": "Scores",
    "academics.lesson_detail": "Lesson Detail", "academics.edit_schedule": "Edit Lesson", "academics.schedule_form": "Schedule Create/Edit", "academics.schedule_pick_class": "Select Class", "academics.schedule_pick_teacher": "Select Teacher", "academics.schedule_teacher_auto": "Class teacher is selected by default", "academics.schedule_autofill": "Class info is auto-filled from selected class", "academics.schedule_pick_day": "Select Day", "academics.schedule_pick_time": "Select Time", "academics.schedule_pick_room": "Select Classroom", "academics.schedule_pick_status": "Select Status", "academics.schedule_pick_student": "Students", "academics.go_structure": "Go to academic structure management", "academics.validation_class_required": "Please select a class first", "academics.validation_end_before_start": "End time must be after start time", "academics.validation_conflict_class": "Class schedule conflicts with existing slot", "academics.validation_conflict_teacher": "Teacher schedule conflicts with existing slot", "academics.validation_conflict_room": "Classroom schedule conflicts with existing slot", "academics.saved": "Schedule saved", "academics.updated": "Schedule updated", "academics.day_all": "All", "academics.status": "Status", "academics.classroom": "Classroom", "academics.action.attendance_eval": "Attendance & Evaluation", "lesson.record.page_title": "Lesson Record Entry", "lesson.record.header": "Student Record Entry", "lesson.record.desc": "Record the full class attendance and evaluation in one pass.", "lesson.record.class_info": "Lesson Info", "lesson.record.back_schedule": "Back to Schedule", "lesson.record.input_title": "Attendance & Evaluation Input", "lesson.record.empty_students": "No students are assigned to this class yet. Assign students to the class and try again.", "lesson.record.invalid_class": "Invalid class_id.", "lesson.record.not_found_class": "Class not found.", "lesson.record.saved": "Lesson record (attendance/evaluation) saved.", "lesson.score.participation": "Participation", "lesson.score.fluency": "Fluency", "lesson.score.vocabulary": "Vocabulary", "lesson.score.reading": "Reading", "lesson.score.homework": "Homework", "lesson.score.attitude": "Attitude", "lesson.score.teacher_memo": "Teacher Memo", "students.detail.section.evaluations": "Recent Lesson Evaluations", "students.eval.avg.title": "Evaluation Averages (Recent 10)", "students.eval.avg.participation": "Avg Participation", "students.eval.avg.fluency": "Avg Fluency", "students.eval.avg.vocabulary": "Avg Vocabulary", "students.eval.avg.reading": "Avg Reading", "students.eval.avg.homework": "Avg Homework", "students.eval.avg.attitude": "Avg Attitude", "academics.copy_week.button": "Copy This Week to Next Week", "academics.copy_week.desc": "Safely copy the currently selected week timetable to next week (duplicates are skipped).", "academics.copy_week.source": "Source Week", "academics.copy_week.target": "Target Week", "academics.copy_week.result": "Copy Result", "academics.copy_week.none": "No schedules found to copy", "academics.copy_week.done": "Weekly copy completed", "academics.copy_week.copied": "Copied", "academics.copy_week.skipped": "Skipped (duplicate)", "academics.week_label": "Week", "academics.week_range": "Week Range", "academics.ref_date": "Reference Date", "academics.selected_day": "Selected Day", "academics.selected_date": "Date", "attendance.manual_input_title": "Manual Attendance Input (Emergency/Admin; not the main workflow)", "lesson.record.weekday_mismatch": "Lesson date weekday must match schedule weekday", "lesson.record.no_students": "No students are assigned to this class", "lesson.record.score_range": "must be an integer between 1 and 5", "validation.date": "Must be in YYYY-MM-DD format", "validation.status_invalid": "invalid status value", "users.teacher_type": "Teacher Type", "users.teacher_type.foreign": "Foreign Teacher", "users.teacher_type.chinese": "Chinese Teacher"
})
I18N_TEXTS["zh"].update({
    "common.add": "添加", "role.owner": "院长", "role.manager": "经理", "role.teacher": "教师", "role.parent": "家长", "role.student": "学生",
    "login.default_accounts": "默认账号: owner/1234, manager/1234, teacher/1234, parent/1234, student/1234",
    "field.id": "ID", "field.name": "姓名", "field.role": "角色", "field.student": "学生", "field.submitted": "提交", "field.homework_title": "作业名", "field.due_date": "截止日",
    "field.submission_count": "提交数", "field.total_targets": "总对象", "field.avg_score": "平均分", "field.score_entries": "录入数",
    "students.field.leave_start_date": "休学开始", "students.field.leave_end_date": "休学结束", "students.password_reset": "重置学生密码", "students.new_password": "新密码",
    "academics.title": "课程/级别/班级/课表管理", "academics.register": "登记", "academics.course": "课程", "academics.level": "级别", "academics.course_name": "课程名", "academics.level_name": "级别名",
    "academics.class_name": "班级名", "academics.class_list": "班级列表", "academics.course_id": "课程ID", "academics.level_id": "级别ID", "academics.schedule_class_id": "课表班级ID",
    "academics.day_of_week": "星期", "academics.start_time": "开始", "academics.end_time": "结束", "academics.teacher": "负责教师", "academics.student_count": "学生数", "academics.schedule": "课表",
    "academics.class_detail.title": "班级详情", "academics.back_to_list": "学术结构列表", "academics.basic_info": "基本信息", "academics.course_level": "课程/级别", "academics.students": "所属学生",
    "academics.sort": "排序", "academics.sort_name": "按姓名排序", "academics.sort_student_no": "按学号排序", "academics.sort_status": "按状态排序", "academics.export_students_csv": "导出学生CSV", "academics.export_attendance_csv": "导出考勤CSV",
    "academics.recent_attendance": "最近考勤", "academics.recent_homework": "最近作业", "academics.recent_exams": "最近考试/成绩",
    "forbidden.teacher_class_only": "仅可处理负责班级", "forbidden.teacher_exam_only": "仅可处理负责考试",
    "library.not_found": "未找到图书", "library.loan_done": "借阅完成", "server.start": "LMS 服务已启动",
    "academics.timetable_title": "课表管理", "academics.timetable_desc": "管理每周课程排布并快速跳转到班级业务。", "academics.filter": "筛选", "academics.week_prev": "上一周", "academics.week_current": "本周", "academics.week_next": "下一周",
    "academics.day_filter": "星期", "academics.teacher_filter": "教师", "academics.classroom_filter": "教室", "academics.course_level_filter": "课程/级别", "academics.class_filter": "班级搜索", "academics.search": "搜索", "academics.add_lesson": "新增课程",
    "academics.timetable": "周课表", "academics.time_slot": "时间段", "academics.teacher_room": "教师/教室", "academics.students_summary": "学生", "academics.more_students": "+更多", "academics.view_class": "班级详情", "academics.go_attendance": "考勤", "academics.go_homework": "作业", "academics.go_exams": "成绩",
    "academics.lesson_detail": "课程详情", "academics.edit_schedule": "编辑课程", "academics.schedule_form": "课表新增/编辑", "academics.schedule_pick_class": "选择班级", "academics.schedule_pick_teacher": "选择教师", "academics.schedule_teacher_auto": "默认选择班级负责教师", "academics.schedule_autofill": "所选班级信息将自动填充", "academics.schedule_pick_day": "选择星期", "academics.schedule_pick_time": "选择时间", "academics.schedule_pick_room": "选择教室", "academics.schedule_pick_status": "选择状态", "academics.schedule_pick_student": "学生", "academics.go_structure": "前往学术结构管理", "academics.validation_class_required": "请先选择班级", "academics.validation_end_before_start": "结束时间必须晚于开始时间", "academics.validation_conflict_class": "同一班级时间冲突", "academics.validation_conflict_teacher": "同一教师时间冲突", "academics.validation_conflict_room": "同一教室时间冲突", "academics.saved": "课表已保存", "academics.updated": "课表已更新", "academics.day_all": "全部", "academics.status": "状态", "academics.classroom": "教室", "academics.action.attendance_eval": "考勤与课堂评价", "lesson.record.page_title": "课程记录录入", "lesson.record.header": "学生记录录入", "lesson.record.desc": "一次性录入本节课所有学生的考勤与课堂评价。", "lesson.record.class_info": "课程信息", "lesson.record.back_schedule": "返回课表", "lesson.record.input_title": "考勤与课堂评价录入", "lesson.record.empty_students": "该班级暂无学生，无法录入。请先分配学生后重试。", "lesson.record.invalid_class": "class_id 参数无效。", "lesson.record.not_found_class": "未找到该班级。", "lesson.record.saved": "课程记录（考勤/评价）已保存", "lesson.score.participation": "课堂参与", "lesson.score.fluency": "流利度", "lesson.score.vocabulary": "词汇", "lesson.score.reading": "阅读", "lesson.score.homework": "作业", "lesson.score.attitude": "课堂态度", "lesson.score.teacher_memo": "教师备注", "students.detail.section.evaluations": "最近课堂评价记录", "students.eval.avg.title": "最近10次评价平均分", "students.eval.avg.participation": "平均课堂参与", "students.eval.avg.fluency": "平均流利度", "students.eval.avg.vocabulary": "平均词汇", "students.eval.avg.reading": "平均阅读", "students.eval.avg.homework": "平均作业", "students.eval.avg.attitude": "平均课堂态度", "academics.copy_week.button": "本周复制到下周", "academics.copy_week.desc": "将当前选择周的课表安全复制到下一周（重复项自动跳过）。", "academics.copy_week.source": "源周", "academics.copy_week.target": "目标周", "academics.copy_week.result": "复制结果", "academics.copy_week.none": "没有可复制的课表", "academics.copy_week.done": "周复制完成", "academics.copy_week.copied": "已复制", "academics.copy_week.skipped": "已跳过（重复）", "academics.week_label": "周", "academics.week_range": "周范围", "academics.ref_date": "基准日期", "academics.selected_day": "所选星期", "academics.selected_date": "日期", "attendance.manual_input_title": "手动考勤录入（应急/管理员用，非主流程）", "lesson.record.weekday_mismatch": "课程日期的星期必须与课表星期一致", "lesson.record.no_students": "该班级没有学生", "lesson.record.score_range": "必须是 1 到 5 的整数", "validation.date": "必须是 YYYY-MM-DD 格式", "validation.status_invalid": "状态值无效", "users.teacher_type": "教师类型", "users.teacher_type.foreign": "外教", "users.teacher_type.chinese": "中文教师"
})

I18N_TEXTS["en"].update({
    "homework.class_context": "Class Context",
    "homework.class_required": "Please select a class",
    "homework.teacher": "Teacher",
    "homework.description": "Description",
    "homework.status": "Status",
    "homework.progress": "Submission Progress",
    "homework.select_homework": "Select Homework",
    "homework.targets": "Targets",
    "homework.submission_panel": "Submission & Feedback",
    "homework.save_students": "Save Student Rows",
    "homework.student_not_in_class": "Student is not in this homework class",
    "homework.saved": "Homework saved",
    "homework.rows_saved": "Submission rows saved",
    "homework.validation.title_required": "Title is required",
    "homework.validation.due_date": "Due date must be YYYY-MM-DD",
    "homework.validation.homework_required": "Homework is required",
    "homework.validation.no_targets": "No students found in class",
    "homework.status.active": "Active",
    "homework.status.closed": "Closed"
})
I18N_TEXTS["ko"].update({
    "homework.class_context": "반 컨텍스트",
    "homework.class_required": "반을 선택하세요",
    "homework.teacher": "강사",
    "homework.description": "설명",
    "homework.status": "상태",
    "homework.progress": "제출 진행률",
    "homework.select_homework": "숙제 선택",
    "homework.targets": "대상 학생",
    "homework.submission_panel": "제출/피드백",
    "homework.save_students": "학생별 저장",
    "homework.student_not_in_class": "학생이 숙제 반 소속이 아닙니다",
    "homework.saved": "숙제가 저장되었습니다",
    "homework.rows_saved": "학생 제출/피드백이 저장되었습니다",
    "homework.validation.title_required": "제목은 필수입니다",
    "homework.validation.due_date": "마감일은 YYYY-MM-DD 형식이어야 합니다",
    "homework.validation.homework_required": "숙제를 선택하세요",
    "homework.validation.no_targets": "반에 학생이 없습니다",
    "homework.status.active": "진행중",
    "homework.status.closed": "종료"
})
I18N_TEXTS["zh"].update({
    "homework.class_context": "班级上下文",
    "homework.class_required": "请选择班级",
    "homework.teacher": "教师",
    "homework.description": "说明",
    "homework.status": "状态",
    "homework.progress": "提交进度",
    "homework.select_homework": "选择作业",
    "homework.targets": "目标学生",
    "homework.submission_panel": "提交/反馈",
    "homework.save_students": "保存学生行",
    "homework.student_not_in_class": "学生不属于该作业班级",
    "homework.saved": "作业已保存",
    "homework.rows_saved": "学生提交/反馈已保存",
    "homework.validation.title_required": "标题必填",
    "homework.validation.due_date": "截止日期必须是 YYYY-MM-DD",
    "homework.validation.homework_required": "请选择作业",
    "homework.validation.no_targets": "班级下没有学生",
    "homework.status.active": "进行中",
    "homework.status.closed": "已结束"
})
I18N_TEXTS["en"].update({
    "academics.foreign_teacher": "Foreign Teacher",
    "academics.chinese_teacher": "Chinese Teacher"
})
I18N_TEXTS["ko"].update({
    "academics.foreign_teacher": "외국인 강사",
    "academics.chinese_teacher": "중국인 강사"
})
I18N_TEXTS["zh"].update({
    "academics.foreign_teacher": "外教",
    "academics.chinese_teacher": "中文教师"
})

I18N_TEXTS["en"].update({
    "attendance.absence_charge_type": "Absence Charge Type",
    "attendance.charge.deduct": "Deduct Credit",
    "attendance.charge.no_deduct": "No Credit Deduction",
    "attendance.requires_makeup": "Requires Makeup",
    "attendance.makeup_completed": "Makeup Completed",
    "attendance.students_needing_makeup": "Students Needing Makeup",
    "attendance.credit_impact": "Credit Impact",
    "attendance.correction": "Attendance Correction",
    "attendance.filter.date_from": "Date From",
    "attendance.filter.date_to": "Date To",
    "attendance.filter.deductible": "Deductible Absence",
    "attendance.export_csv": "Export CSV",
    "attendance.updated": "Attendance updated",
    "attendance.makeup_marked": "Makeup marked as completed",
    "attendance.validation.charge_required": "Please choose absence charge type for absent students",
    "attendance.validation.charge_invalid": "Invalid absence charge type",
    "attendance.validation.bool_invalid": "Invalid boolean option",
    "attendance.list_desc": "Use filters below to manage attendance and makeup follow-ups."
})
I18N_TEXTS["ko"].update({
    "attendance.absence_charge_type": "결석 차감 유형",
    "attendance.charge.deduct": "수업 차감",
    "attendance.charge.no_deduct": "차감 없음",
    "attendance.requires_makeup": "보강 필요",
    "attendance.makeup_completed": "보강 완료",
    "attendance.students_needing_makeup": "보강 필요 학생",
    "attendance.credit_impact": "크레딧 반영",
    "attendance.correction": "출결 정정",
    "attendance.filter.date_from": "시작일",
    "attendance.filter.date_to": "종료일",
    "attendance.filter.deductible": "차감 결석",
    "attendance.export_csv": "CSV 내보내기",
    "attendance.updated": "출결이 수정되었습니다",
    "attendance.makeup_marked": "보강 완료로 처리되었습니다",
    "attendance.validation.charge_required": "결석 학생은 차감 유형을 선택해야 합니다",
    "attendance.validation.charge_invalid": "잘못된 결석 차감 유형입니다",
    "attendance.validation.bool_invalid": "잘못된 예/아니오 값입니다",
    "attendance.list_desc": "아래 필터로 출결 및 보강 대상 상태를 관리하세요."
})
I18N_TEXTS["zh"].update({
    "attendance.absence_charge_type": "缺勤扣课类型",
    "attendance.charge.deduct": "扣课",
    "attendance.charge.no_deduct": "不扣课",
    "attendance.requires_makeup": "需要补课",
    "attendance.makeup_completed": "补课完成",
    "attendance.students_needing_makeup": "需补课学生",
    "attendance.credit_impact": "学分影响",
    "attendance.correction": "考勤更正",
    "attendance.filter.date_from": "开始日期",
    "attendance.filter.date_to": "结束日期",
    "attendance.filter.deductible": "可扣课缺勤",
    "attendance.export_csv": "导出CSV",
    "attendance.updated": "考勤已更新",
    "attendance.makeup_marked": "已标记为补课完成",
    "attendance.validation.charge_required": "缺勤学生必须选择扣课类型",
    "attendance.validation.charge_invalid": "缺勤扣课类型无效",
    "attendance.validation.bool_invalid": "布尔值无效",
    "attendance.list_desc": "使用以下筛选管理考勤与补课跟进。"
})


I18N_TEXTS["en"].update({
    "common.query_to_load": "Press Query to load data.",
    "students.export_excel": "Download Excel",
    "students.upload_students": "Upload Students",
    "students.upload_template": "Download Upload Template",
    "students.upload.file_required": "Upload file is required",
    "students.upload.empty_file": "Empty file",
    "students.upload.invalid_headers": "Invalid headers. Missing: {missing}",
    "students.upload.failed": "Upload failed: {error}",
    "students.export.query_first": "Please run query first",
})
I18N_TEXTS["ko"].update({
    "common.query_to_load": "Press Query to load data.",
    "students.export_excel": "Download Excel",
    "students.upload_students": "Upload Students",
    "students.upload_template": "Download Upload Template",
    "students.upload.file_required": "Upload file is required",
    "students.upload.empty_file": "Empty file",
    "students.upload.invalid_headers": "Invalid headers. Missing: {missing}",
    "students.upload.failed": "Upload failed: {error}",
    "students.export.query_first": "Please run query first",
})
I18N_TEXTS["zh"].update({
    "common.query_to_load": "Press Query to load data.",
    "students.export_excel": "Download Excel",
    "students.upload_students": "Upload Students",
    "students.upload_template": "Download Upload Template",
    "students.upload.file_required": "Upload file is required",
    "students.upload.empty_file": "Empty file",
    "students.upload.invalid_headers": "Invalid headers. Missing: {missing}",
    "students.upload.failed": "Upload failed: {error}",
    "students.export.query_first": "Please run query first",
})
def load_locale_files():
    locales_dir = os.path.join(BASE_DIR, "locales")
    if not os.path.isdir(locales_dir):
        return
    for fp in sorted(Path(locales_dir).glob("*.json")):
        code = fp.stem.lower()
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        labels = data.pop("__labels__", {}) if isinstance(data.get("__labels__", {}), dict) else {}
        base = I18N_TEXTS.get(code, {}).copy()
        base.update(data)
        I18N_TEXTS[code] = base
        if isinstance(labels.get("lang"), str) and labels["lang"].strip():
            LANG_LABELS[code] = labels["lang"].strip()
        elif code not in LANG_LABELS:
            LANG_LABELS[code] = code


load_locale_files()
SUPPORTED_LANGS = set(I18N_TEXTS.keys())

NAV_PATHS = {
    "dashboard": "/dashboard",
    "users": "/users",
    "students": "/students",
    "academics": "/academics",
    "masterdata": "/masterdata",
    "schedule": "/schedule",
    "attendance": "/attendance",
    "homework": "/homework",
    "exams": "/exams",
    "counseling": "/counseling",
    "payments": "/payments",
    "announcements": "/announcements",
    "library": "/library",
    "logs": "/logs",
}
ROLE_MENU_KEYS = {
    ROLE_OWNER: ["dashboard", "users", "students", "masterdata", "schedule", "attendance", "homework", "exams", "counseling", "payments", "announcements", "library", "logs"],
    ROLE_MANAGER: ["dashboard", "students", "masterdata", "schedule", "attendance", "homework", "exams", "counseling", "payments", "announcements", "library"],
    ROLE_TEACHER: ["dashboard", "students", "schedule", "attendance", "homework", "exams", "counseling", "announcements", "library"],
    ROLE_PARENT: ["dashboard", "students", "attendance", "homework", "exams", "payments", "announcements"],
    ROLE_STUDENT: ["dashboard", "students", "attendance", "homework", "exams", "announcements"],
}
def t(key, lang=None):
    lang = lang or CURRENT_LANG
    table = I18N_TEXTS.get(lang, I18N_TEXTS["ko"])
    return table.get(key, I18N_TEXTS["ko"].get(key, key))
def menu_t(key, lang=None):
    token = NAV_LABELS.get(key, key)
    return t(token, lang)
def status_t(status, lang=None):
    return t(f"status.{status}", lang)

def attendance_status_t(status, lang=None):
    return t(f"status.{status}", lang)

def role_label(role, lang=None):
    return t(f"role.{role}", lang)

def safe_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default

def day_sort_value(day_text):
    days = {"mon":1,"monday":1,"tue":2,"tues":2,"tuesday":2,"wed":3,"wednesday":3,"thu":4,"thursday":4,"fri":5,"friday":5,"sat":6,"saturday":6,"sun":7,"sunday":7,
            "월":1,"화":2,"수":3,"목":4,"금":5,"토":6,"일":7,"周一":1,"周二":2,"周三":3,"周四":4,"周五":5,"周六":6,"周日":7,"星期一":1,"星期二":2,"星期三":3,"星期四":4,"星期五":5,"星期六":6,"星期日":7}
    k = (day_text or "").strip().lower()
    return days.get(k, 99)

def parse_hhmm(v):
    if not v or ":" not in v:
        return None
    try:
        h, m = v.split(":", 1)
        return int(h) * 60 + int(m)
    except Exception:
        return None


def is_valid_date(v):
    if not v:
        return False
    try:
        datetime.strptime(v, "%Y-%m-%d")
        return True
    except Exception:
        return False


def as_int(v):
    try:
        return int(str(v).strip())
    except Exception:
        return None


def as_float(v):
    try:
        return float(str(v).strip())
    except Exception:
        return None


def parse_bool_flag(v):
    vv = str(v or "").strip().lower()
    if vv in ("1", "true", "yes", "y", "on"):
        return 1
    if vv in ("0", "false", "no", "n", "off", ""):
        return 0
    return None


def get_class_credit_unit(conn, class_id):
    try:
        row = conn.execute("SELECT credit_unit FROM classes WHERE id=?", (class_id,)).fetchone()
        unit = as_float(row["credit_unit"]) if row and "credit_unit" in row.keys() else None
        return unit if unit is not None and unit > 0 else 1.0
    except Exception:
        return 1.0


def calc_credit_delta(status_v, absence_charge_type, credit_unit):
    if status_v in ("present", "late"):
        return -float(credit_unit)
    if status_v == "absent":
        return -float(credit_unit) if absence_charge_type == "deduct" else 0.0
    return 0.0


def apply_credit_adjustment(conn, student_user_id, old_delta, new_delta):
    old_v = as_float(old_delta) if old_delta is not None else 0.0
    new_v = as_float(new_delta) if new_delta is not None else 0.0
    if old_v is None:
        old_v = 0.0
    if new_v is None:
        new_v = 0.0
    adjustment = new_v - old_v
    if abs(adjustment) < 1e-9:
        return
    conn.execute(
        "UPDATE students SET remaining_credits=COALESCE(remaining_credits, 0) + ? WHERE user_id=?",
        (adjustment, student_user_id),
    )


def add_error(errors, field, msg):
    errors.append(f"- {field}: {msg}")


def format_errors(errors):
    if not errors:
        return ""
    return "<br>".join(errors)


def h(v):
    if v is None:
        return ""
    return html_lib.escape(str(v), quote=True)


def ensure_exists(conn, table, value, field="id", extra_where="", extra_params=()):
    if value is None or str(value).strip() == "":
        return False
    sql = f"SELECT 1 FROM {table} WHERE {field}=?"
    params = [value]
    if extra_where:
        sql += f" AND {extra_where}"
        params.extend(extra_params)
    return conn.execute(sql, tuple(params)).fetchone() is not None


def ensure_logs_table(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS app_logs (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      level TEXT NOT NULL,
      route TEXT,
      user_id INTEGER,
      message TEXT NOT NULL,
      detail TEXT,
      created_at TEXT NOT NULL,
      FOREIGN KEY(user_id) REFERENCES users(id)
    )""")


def ensure_logs_columns(conn):
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(app_logs)").fetchall()}
    if "route" not in cols:
        conn.execute("ALTER TABLE app_logs ADD COLUMN route TEXT")
    if "user_id" not in cols:
        conn.execute("ALTER TABLE app_logs ADD COLUMN user_id INTEGER")
    if "detail" not in cols:
        conn.execute("ALTER TABLE app_logs ADD COLUMN detail TEXT")
    if "created_at" not in cols:
        conn.execute("ALTER TABLE app_logs ADD COLUMN created_at TEXT")


def log_event(conn, level, route, message, detail="", user_id=None):
    try:
        conn.execute(
            "INSERT INTO app_logs(level, route, user_id, message, detail, created_at) VALUES(?,?,?,?,?,?)",
            (level, route, user_id, message, detail[:4000] if detail else None, now()),
        )
        conn.commit()
    except Exception:
        pass

def is_time_overlap(a_start, a_end, b_start, b_end):
    a1, a2, b1, b2 = parse_hhmm(a_start), parse_hhmm(a_end), parse_hhmm(b_start), parse_hhmm(b_end)
    if None in (a1, a2, b1, b2):
        return False
    return not (a2 <= b1 or b2 <= a1)

def ensure_schedule_columns(conn):
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(schedules)").fetchall()}
    if "classroom" not in cols:
        conn.execute("ALTER TABLE schedules ADD COLUMN classroom TEXT")
    if "status" not in cols:
        conn.execute("ALTER TABLE schedules ADD COLUMN status TEXT DEFAULT 'active'")
    if "note" not in cols:
        conn.execute("ALTER TABLE schedules ADD COLUMN note TEXT")
    if "teacher_id" not in cols:
        conn.execute("ALTER TABLE schedules ADD COLUMN teacher_id INTEGER")
    if "week_start_date" not in cols:
        conn.execute("ALTER TABLE schedules ADD COLUMN week_start_date TEXT")


def ensure_attendance_columns(conn):
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(attendance)").fetchall()}
    extra_cols = [
        ("participation_score", "INTEGER"),
        ("fluency_score", "INTEGER"),
        ("vocabulary_score", "INTEGER"),
        ("reading_score", "INTEGER"),
        ("homework_score", "INTEGER"),
        ("attitude_score", "INTEGER"),
        ("teacher_memo", "TEXT"),
        ("schedule_id", "INTEGER"),
    ]
    for col, typ in extra_cols:
        if col not in cols:
            conn.execute(f"ALTER TABLE attendance ADD COLUMN {col} {typ}")


def ensure_homework_columns(conn):
    hcols = {r["name"] for r in conn.execute("PRAGMA table_info(homework)").fetchall()}
    if "teacher_id" not in hcols:
        conn.execute("ALTER TABLE homework ADD COLUMN teacher_id INTEGER")
    if "description" not in hcols:
        conn.execute("ALTER TABLE homework ADD COLUMN description TEXT")
    if "status" not in hcols:
        conn.execute("ALTER TABLE homework ADD COLUMN status TEXT DEFAULT 'active'")
    if "updated_at" not in hcols:
        conn.execute("ALTER TABLE homework ADD COLUMN updated_at TEXT")
    if "created_by" in hcols and "teacher_id" in {r["name"] for r in conn.execute("PRAGMA table_info(homework)").fetchall()}:
        conn.execute("UPDATE homework SET teacher_id=COALESCE(teacher_id, created_by) WHERE teacher_id IS NULL")

    hscols = {r["name"] for r in conn.execute("PRAGMA table_info(homework_submissions)").fetchall()}
    if "feedback_teacher_id" not in hscols:
        conn.execute("ALTER TABLE homework_submissions ADD COLUMN feedback_teacher_id INTEGER")
    if "updated_at" not in hscols:
        conn.execute("ALTER TABLE homework_submissions ADD COLUMN updated_at TEXT")
    if "feedback_by" in hscols and "feedback_teacher_id" in {r["name"] for r in conn.execute("PRAGMA table_info(homework_submissions)").fetchall()}:
        conn.execute("UPDATE homework_submissions SET feedback_teacher_id=COALESCE(feedback_teacher_id, feedback_by) WHERE feedback_teacher_id IS NULL")


def ensure_user_columns(conn):
    ucols = {r["name"] for r in conn.execute("PRAGMA table_info(users)").fetchall()}
    if "name" not in ucols:
        conn.execute("ALTER TABLE users ADD COLUMN name TEXT")
        conn.execute("UPDATE users SET name=COALESCE(username, 'user_' || id) WHERE name IS NULL OR TRIM(name)=''")
    if "teacher_type" not in ucols:
        conn.execute("ALTER TABLE users ADD COLUMN teacher_type TEXT")


def ensure_teacher_table(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS teachers (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      user_id INTEGER NOT NULL UNIQUE,
      teacher_type TEXT NOT NULL DEFAULT 'foreign',
      memo TEXT,
      created_at TEXT NOT NULL,
      updated_at TEXT,
      FOREIGN KEY(user_id) REFERENCES users(id)
    )""")
    tcols = {r["name"] for r in conn.execute("PRAGMA table_info(teachers)").fetchall()}
    if "teacher_type" not in tcols:
        conn.execute("ALTER TABLE teachers ADD COLUMN teacher_type TEXT")
    if "memo" not in tcols:
        conn.execute("ALTER TABLE teachers ADD COLUMN memo TEXT")
    if "created_at" not in tcols:
        conn.execute("ALTER TABLE teachers ADD COLUMN created_at TEXT")
    if "updated_at" not in tcols:
        conn.execute("ALTER TABLE teachers ADD COLUMN updated_at TEXT")


def ensure_extended_columns(conn):
    ensure_user_columns(conn)

    ccols = {r["name"] for r in conn.execute("PRAGMA table_info(classes)").fetchall()}
    if "foreign_teacher_id" not in ccols:
        conn.execute("ALTER TABLE classes ADD COLUMN foreign_teacher_id INTEGER")
    if "chinese_teacher_id" not in ccols:
        conn.execute("ALTER TABLE classes ADD COLUMN chinese_teacher_id INTEGER")
    if "status" not in ccols:
        conn.execute("ALTER TABLE classes ADD COLUMN status TEXT DEFAULT 'active'")
    if "memo" not in ccols:
        conn.execute("ALTER TABLE classes ADD COLUMN memo TEXT")
    if "credit_unit" not in ccols:
        conn.execute("ALTER TABLE classes ADD COLUMN credit_unit REAL DEFAULT 1")

    acols = {r["name"] for r in conn.execute("PRAGMA table_info(attendance)").fetchall()}
    if "absence_charge_type" not in acols:
        conn.execute("ALTER TABLE attendance ADD COLUMN absence_charge_type TEXT")
    if "requires_makeup" not in acols:
        conn.execute("ALTER TABLE attendance ADD COLUMN requires_makeup INTEGER DEFAULT 0")
    if "makeup_completed" not in acols:
        conn.execute("ALTER TABLE attendance ADD COLUMN makeup_completed INTEGER DEFAULT 0")
    if "credit_delta" not in acols:
        conn.execute("ALTER TABLE attendance ADD COLUMN credit_delta REAL DEFAULT 0")
    if "makeup_attendance_id" not in acols:
        conn.execute("ALTER TABLE attendance ADD COLUMN makeup_attendance_id INTEGER")

    scols = {r["name"] for r in conn.execute("PRAGMA table_info(schedules)").fetchall()}
    if "room_id" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN room_id INTEGER")
    if "foreign_teacher_id" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN foreign_teacher_id INTEGER")
    if "chinese_teacher_id" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN chinese_teacher_id INTEGER")
    if "substitute_foreign_teacher_id" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN substitute_foreign_teacher_id INTEGER")
    if "substitute_chinese_teacher_id" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN substitute_chinese_teacher_id INTEGER")
    if "substitute_reason" not in scols:
        conn.execute("ALTER TABLE schedules ADD COLUMN substitute_reason TEXT")

    rcols = {r["name"] for r in conn.execute("PRAGMA table_info(classrooms)").fetchall()}
    if "room_code" not in rcols:
        conn.execute("ALTER TABLE classrooms ADD COLUMN room_code TEXT")
    if "room_name" not in rcols:
        conn.execute("ALTER TABLE classrooms ADD COLUMN room_name TEXT")
    if "status" not in rcols:
        conn.execute("ALTER TABLE classrooms ADD COLUMN status TEXT DEFAULT 'active'")
    if "memo" not in rcols:
        conn.execute("ALTER TABLE classrooms ADD COLUMN memo TEXT")

    stcols = {r["name"] for r in conn.execute("PRAGMA table_info(students)").fetchall()}
    if "homeroom_teacher_id" not in stcols:
        conn.execute("ALTER TABLE students ADD COLUMN homeroom_teacher_id INTEGER")


def ensure_master_tables(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS classrooms (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL UNIQUE,
      created_at TEXT NOT NULL
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS time_slots (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      label TEXT NOT NULL UNIQUE,
      start_time TEXT NOT NULL,
      end_time TEXT NOT NULL,
      created_at TEXT NOT NULL
    )""")


def repair_profile_integrity(conn):
    # 학생 role 사용자 -> students 프로필 보강
    student_users = conn.execute("SELECT id, name FROM users WHERE role=?", (ROLE_STUDENT,)).fetchall()
    for su in student_users:
        exists = conn.execute("SELECT id FROM students WHERE user_id=?", (su["id"],)).fetchone()
        if not exists:
            conn.execute(
                """INSERT INTO students(
                user_id, student_no, name_ko, status, created_at
                ) VALUES(?,?,?,?,?)""",
                (su["id"], f"S{su['id']:04d}", su["name"] or su["id"], "active", now()),
            )

    # 교사 role 사용자 -> teachers 프로필 보강 + teacher_type 정규화
    teacher_users = conn.execute("SELECT id, teacher_type FROM users WHERE role=?", (ROLE_TEACHER,)).fetchall()
    for tu in teacher_users:
        default_type = (tu["teacher_type"] or "").strip() or "foreign"
        if default_type not in ("foreign", "chinese"):
            default_type = "foreign"
        trow = conn.execute("SELECT id, teacher_type FROM teachers WHERE user_id=?", (tu["id"],)).fetchone()
        if not trow:
            conn.execute(
                "INSERT INTO teachers(user_id, teacher_type, created_at, updated_at) VALUES(?,?,?,?)",
                (tu["id"], default_type, now(), now()),
            )
        else:
            t_type = (trow["teacher_type"] or "").strip()
            if t_type not in ("foreign", "chinese"):
                conn.execute("UPDATE teachers SET teacher_type=?, updated_at=? WHERE id=?", (default_type, now(), trow["id"]))
        conn.execute("UPDATE users SET teacher_type=? WHERE id=?", (default_type, tu["id"]))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn
def hash_pw(pw: str) -> str:
    import hashlib
    import secrets

    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), salt.encode("utf-8"), 200000)
    return f"pbkdf2_sha256$200000${salt}${digest.hex()}"


def verify_pw(pw: str, stored_hash: str) -> bool:
    import hashlib
    import hmac

    stored_hash = str(stored_hash or "")
    if stored_hash.startswith("pbkdf2_sha256$"):
        try:
            _, iterations, salt, digest = stored_hash.split("$", 3)
            derived = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), salt.encode("utf-8"), int(iterations))
            return hmac.compare_digest(derived.hex(), digest)
        except Exception:
            return False
    return hmac.compare_digest(hashlib.sha256(pw.encode("utf-8")).hexdigest(), stored_hash)


def needs_password_rehash(stored_hash: str) -> bool:
    return not str(stored_hash or "").startswith("pbkdf2_sha256$")


def now():
    return datetime.now(timezone.utc).isoformat()


def parse_iso_datetime(value):
    if not value:
        return None
    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def session_ttl_seconds():
    try:
        return max(300, int(os.getenv("SESSION_TTL_SECONDS", str(60 * 60 * 12))))
    except ValueError:
        return 60 * 60 * 12


def session_cookie_secure(environ=None):
    if str(os.getenv("SESSION_COOKIE_SECURE", "")).strip() == "1":
        return True
    if environ is None:
        return False
    return (environ.get("wsgi.url_scheme") or "").lower() == "https"


def build_session_cookie(token, environ=None):
    parts = [
        f"session={token}",
        "Path=/",
        "HttpOnly",
        "SameSite=Lax",
        f"Max-Age={session_ttl_seconds()}",
    ]
    if session_cookie_secure(environ):
        parts.append("Secure")
    return "; ".join(parts)


def clear_session_cookie(environ=None):
    parts = ["session=", "Path=/", "HttpOnly", "SameSite=Lax", "Max-Age=0"]
    if session_cookie_secure(environ):
        parts.append("Secure")
    return "; ".join(parts)


def create_session(conn, user_id):
    token = str(uuid.uuid4())
    conn.execute("INSERT INTO sessions(user_id, token, created_at) VALUES(?,?,?)", (user_id, token, now()))
    return token


def invalidate_session(conn, token):
    if token:
        conn.execute("DELETE FROM sessions WHERE token=?", (token,))


def week_bounds_from_ref(ref_date_str=None, week_offset=0):
    try:
        base = datetime.strptime(ref_date_str, "%Y-%m-%d").date() if ref_date_str else datetime.utcnow().date()
    except Exception:
        base = datetime.utcnow().date()
    monday = base - timedelta(days=base.weekday()) + timedelta(days=week_offset * 7)
    sunday = monday + timedelta(days=6)
    iso_year, iso_week, _ = monday.isocalendar()
    return monday, sunday, iso_year, iso_week


def weekday_from_ref(ref_date_str=None, week_offset=0):
    try:
        base = datetime.strptime(ref_date_str, "%Y-%m-%d").date() if ref_date_str else datetime.utcnow().date()
    except Exception:
        base = datetime.utcnow().date()
    target = base + timedelta(days=week_offset * 7)
    return target.strftime("%a"), target.isoformat()

def iso_monday_str(ref_date_str=None, week_offset=0):
    monday, _, _, _ = week_bounds_from_ref(ref_date_str, week_offset)
    return monday.isoformat()


def copy_week_schedules(conn, source_week_start, target_week_start, class_ids=None):
    class_ids = class_ids or []
    if not class_ids:
        return {"copied": 0, "skipped": 0, "source_count": 0}
    placeholders = ",".join(["?"] * len(class_ids))
    src_rows = conn.execute(
        f"""SELECT * FROM schedules
        WHERE class_id IN ({placeholders})
          AND COALESCE(week_start_date, '') IN ('', ?)
        ORDER BY id""",
        (*class_ids, source_week_start),
    ).fetchall()
    copied = 0
    skipped = 0
    for row in src_rows:
        dup = conn.execute(
            """SELECT 1 FROM schedules
            WHERE class_id=? AND day_of_week=? AND start_time=? AND end_time=?
              AND COALESCE(classroom,'')=COALESCE(?, '')
              AND COALESCE(teacher_id,0)=COALESCE(?,0)
              AND COALESCE(week_start_date,'') IN ('', ?)
            LIMIT 1""",
            (row["class_id"], row["day_of_week"], row["start_time"], row["end_time"], row["classroom"], row["teacher_id"], target_week_start),
        ).fetchone()
        if dup:
            skipped += 1
            continue
        getv = lambda k: row[k] if k in row.keys() else None
        conn.execute(
            """INSERT INTO schedules(class_id, day_of_week, start_time, end_time, classroom, status, note, teacher_id, created_at, room_id, foreign_teacher_id, chinese_teacher_id, substitute_foreign_teacher_id, substitute_chinese_teacher_id, substitute_reason, week_start_date)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                row["class_id"], row["day_of_week"], row["start_time"], row["end_time"], row["classroom"],
                row["status"], row["note"], row["teacher_id"], now(),
                getv("room_id"), getv("foreign_teacher_id"), getv("chinese_teacher_id"), getv("substitute_foreign_teacher_id"),
                getv("substitute_chinese_teacher_id"), getv("substitute_reason"), target_week_start,
            ),
        )
        copied += 1
    return {"copied": copied, "skipped": skipped, "source_count": len(src_rows)}



def ensure_user_account(conn, name, username, role, password="1234", teacher_type=None):
    row = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if row:
        uid = row["id"]
        conn.execute("UPDATE users SET name=?, role=?, teacher_type=COALESCE(?, teacher_type) WHERE id=?", (name, role, teacher_type, uid))
    else:
        conn.execute(
            "INSERT INTO users(name, username, password_hash, role, teacher_type, created_at) VALUES(?,?,?,?,?,?)",
            (name, username, hash_pw(password), role, teacher_type, now()),
        )
        uid = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()["id"]
    return uid


def ensure_teacher_profile(conn, user_id, teacher_type):
    row = conn.execute("SELECT id FROM teachers WHERE user_id=?", (user_id,)).fetchone()
    if row:
        conn.execute("UPDATE teachers SET teacher_type=?, updated_at=? WHERE user_id=?", (teacher_type, now(), user_id))
    else:
        conn.execute(
            "INSERT INTO teachers(user_id, teacher_type, created_at, updated_at) VALUES(?,?,?,?)",
            (user_id, teacher_type, now(), now()),
        )


def ensure_course(conn, name):
    row = conn.execute("SELECT id FROM courses WHERE name=?", (name,)).fetchone()
    if row:
        return row["id"]
    conn.execute("INSERT INTO courses(name, created_at) VALUES(?,?)", (name, now()))
    return conn.execute("SELECT id FROM courses WHERE name=?", (name,)).fetchone()["id"]


def ensure_level(conn, course_id, name):
    row = conn.execute("SELECT id FROM levels WHERE course_id=? AND name=?", (course_id, name)).fetchone()
    if row:
        return row["id"]
    conn.execute("INSERT INTO levels(course_id, name, created_at) VALUES(?,?,?)", (course_id, name, now()))
    return conn.execute("SELECT id FROM levels WHERE course_id=? AND name=?", (course_id, name)).fetchone()["id"]


def ensure_classroom(conn, name):
    row = conn.execute("SELECT id FROM classrooms WHERE name=?", (name,)).fetchone()
    if row:
        return row["id"]
    conn.execute("INSERT INTO classrooms(name, created_at) VALUES(?,?)", (name, now()))
    return conn.execute("SELECT id FROM classrooms WHERE name=?", (name,)).fetchone()["id"]


def ensure_time_slot(conn, label, start_time, end_time):
    row = conn.execute("SELECT id FROM time_slots WHERE label=?", (label,)).fetchone()
    if row:
        conn.execute("UPDATE time_slots SET start_time=?, end_time=? WHERE id=?", (start_time, end_time, row["id"]))
        return row["id"]
    conn.execute("INSERT INTO time_slots(label, start_time, end_time, created_at) VALUES(?,?,?,?)", (label, start_time, end_time, now()))
    return conn.execute("SELECT id FROM time_slots WHERE label=?", (label,)).fetchone()["id"]


def ensure_class(conn, name, course_id, level_id, foreign_teacher_id, chinese_teacher_id, credit_unit, status="active", memo=""):
    row = conn.execute("SELECT id FROM classes WHERE name=?", (name,)).fetchone()
    teacher_id = foreign_teacher_id
    if row:
        conn.execute(
            """UPDATE classes SET course_id=?, level_id=?, teacher_id=?, foreign_teacher_id=?, chinese_teacher_id=?,
            credit_unit=?, status=?, memo=? WHERE id=?""",
            (course_id, level_id, teacher_id, foreign_teacher_id, chinese_teacher_id, credit_unit, status, memo, row["id"]),
        )
        return row["id"]
    conn.execute(
        """INSERT INTO classes(course_id, level_id, name, teacher_id, foreign_teacher_id, chinese_teacher_id, credit_unit, status, memo, created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?)""",
        (course_id, level_id, name, teacher_id, foreign_teacher_id, chinese_teacher_id, credit_unit, status, memo, now()),
    )
    return conn.execute("SELECT id FROM classes WHERE name=?", (name,)).fetchone()["id"]


def find_schedule_conflict(conn, class_id, day_of_week, start_time, end_time, classroom, teacher_id, week_start_date, ignore_id=None):
    rows = conn.execute(
        """SELECT sc.id, sc.class_id, sc.teacher_id, sc.start_time, sc.end_time, COALESCE(sc.classroom, '') AS classroom,
        c.teacher_id AS class_teacher_id, COALESCE(c.foreign_teacher_id, c.teacher_id) AS class_foreign_teacher_id
        FROM schedules sc
        LEFT JOIN classes c ON c.id=sc.class_id
        WHERE sc.day_of_week=?
          AND COALESCE(sc.week_start_date,'') IN ('', ?)""",
        (day_of_week, week_start_date),
    ).fetchall()
    for ex in rows:
        if ignore_id and str(ex["id"]) == str(ignore_id):
            continue
        if not is_time_overlap(start_time, end_time, ex["start_time"], ex["end_time"]):
            continue
        if str(ex["class_id"]) == str(class_id):
            return t("academics.validation_conflict_class")
        existing_teacher_id = ex["teacher_id"] or ex["class_foreign_teacher_id"] or ex["class_teacher_id"]
        if teacher_id and existing_teacher_id and str(existing_teacher_id) == str(teacher_id):
            return t("academics.validation_conflict_teacher")
        if classroom and ex["classroom"] and ex["classroom"].strip().lower() == classroom.strip().lower():
            return t("academics.validation_conflict_room")
    return None


def ensure_schedule_row(conn, class_id, day_of_week, start_time, end_time, classroom, teacher_id, week_start_date):
    row = conn.execute(
        """SELECT id FROM schedules WHERE class_id=? AND day_of_week=? AND start_time=? AND end_time=?
        AND COALESCE(classroom,'')=COALESCE(?, '') AND COALESCE(week_start_date,'') IN ('', ?)""",
        (class_id, day_of_week, start_time, end_time, classroom, week_start_date),
    ).fetchone()
    conflict = find_schedule_conflict(conn, class_id, day_of_week, start_time, end_time, classroom, teacher_id, week_start_date, ignore_id=(row["id"] if row else None))
    if conflict:
        raise ValueError(conflict)
    if row:
        conn.execute("UPDATE schedules SET teacher_id=?, status='active', classroom=?, week_start_date=? WHERE id=?", (teacher_id, classroom, week_start_date, row["id"]))
        return row["id"]
    conn.execute(
        """INSERT INTO schedules(class_id, day_of_week, start_time, end_time, classroom, status, note, teacher_id, created_at, week_start_date)
        VALUES(?,?,?,?,?,'active',NULL,?,?,?)""",
        (class_id, day_of_week, start_time, end_time, classroom, teacher_id, now(), week_start_date),
    )
    return conn.execute("SELECT id FROM schedules WHERE class_id=? AND day_of_week=? AND start_time=? AND end_time=? AND week_start_date=? ORDER BY id DESC LIMIT 1",
                        (class_id, day_of_week, start_time, end_time, week_start_date)).fetchone()["id"]


def seed_demo_data(conn, force=False, options=None):
    options = dict(options or {})
    app_env = (os.getenv("APP_ENV") or "").strip().lower()
    allow = force or app_env in ("staging", "demo") or os.getenv("SEED_DEMO_DATA") == "1"
    if not allow:
        return {"seeded": False, "reason": "disabled"}

    base_users = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
    if not force and base_users > 300:
        return {"seeded": False, "reason": "existing_data"}

    preset = str(options.get("preset", "")).strip().lower()
    preset_counts = {
        "large_school": {
            "manager_count": 4,
            "foreign_teacher_count": 10,
            "chinese_teacher_count": 10,
            "parent_count": 240,
            "student_count": 400,
            "course_count": 8,
            "level_count": 4,
            "class_count": 40,
            "classroom_count": 12,
            "book_count": 500,
            "announcement_count": 8,
        }
    }.get(preset, {})

    def opt_int(name, default, minimum=0):
        raw = options.get(name, preset_counts.get(name, default))
        try:
            value = int(raw)
        except (TypeError, ValueError):
            value = default
        return max(minimum, value)

    manager_count = opt_int("manager_count", 2, 1)
    foreign_teacher_count = opt_int("foreign_teacher_count", 4, 1)
    chinese_teacher_count = opt_int("chinese_teacher_count", 4, 1)
    student_count = opt_int("student_count", 30, 1)
    parent_count = opt_int("parent_count", max(12, (student_count + 1) // 2), 1)
    course_count = opt_int("course_count", 4, 1)
    level_count = opt_int("level_count", 3, 1)
    class_count = opt_int("class_count", 8, 1)
    classroom_count = opt_int("classroom_count", max(6, class_count // 4), 1)
    book_count = opt_int("book_count", max(24, class_count * 6), 0)
    announcement_count = opt_int("announcement_count", 4, 0)

    owner_id = ensure_user_account(conn, "Demo Owner", "demo_owner", ROLE_OWNER)
    for i in range(1, manager_count + 1):
        ensure_user_account(conn, f"Demo Manager {chr(64 + i)}", f"demo_manager_{i:02d}", ROLE_MANAGER)

    foreign_first = ["Alex", "Emma", "Liam", "Noah", "Olivia", "Mason", "Sophia", "Ethan", "Ava", "Lucas", "Mia", "James"]
    foreign_last = ["Carter", "Stone", "Foster", "Reed", "Brooks", "Parker", "Cole", "Hayes", "Price", "Kelly", "Turner", "Morris"]
    chinese_family = ["Wang", "Li", "Zhang", "Liu", "Chen", "Yang", "Zhao", "Huang", "Wu", "Zhou", "Xu", "Sun", "Ma", "Zhu", "Hu", "Guo"]
    chinese_given = ["Li", "Yu", "Min", "Fang", "Xin", "Jie", "Nan", "Yun", "Tao", "Rui", "Jing", "Han", "Yue", "Bo", "Qi", "Ling"]

    def teacher_name(index, kind):
        if kind == "foreign":
            return f"{foreign_first[(index - 1) % len(foreign_first)]} {foreign_last[((index - 1) // len(foreign_first)) % len(foreign_last)]}"
        return f"{chinese_family[(index - 1) % len(chinese_family)]} {chinese_given[((index - 1) // len(chinese_family)) % len(chinese_given)]}"

    foreign_ids = []
    chinese_ids = []
    for i in range(1, foreign_teacher_count + 1):
        uid = ensure_user_account(conn, teacher_name(i, "foreign"), f"demo_ft_{i:02d}", ROLE_TEACHER, teacher_type="foreign")
        ensure_teacher_profile(conn, uid, "foreign")
        foreign_ids.append(uid)
    for i in range(1, chinese_teacher_count + 1):
        uid = ensure_user_account(conn, teacher_name(i, "chinese"), f"demo_ct_{i:02d}", ROLE_TEACHER, teacher_type="chinese")
        ensure_teacher_profile(conn, uid, "chinese")
        chinese_ids.append(uid)

    parent_ids = []
    for i in range(1, parent_count + 1):
        pid = ensure_user_account(conn, f"Parent {i:03d}", f"demo_parent_{i:03d}", ROLE_PARENT)
        parent_ids.append(pid)

    base_courses = ["Phonics", "Reading", "Writing", "Speaking", "Grammar", "Vocabulary", "Debate", "Test Prep", "Literature", "Presentation"]
    course_names = base_courses[:course_count]
    if len(course_names) < course_count:
        for i in range(len(course_names) + 1, course_count + 1):
            course_names.append(f"Course {i}")

    level_labels = ["Starter", "Basic", "Intermediate", "Advanced", "Master"]
    selected_levels = level_labels[:level_count]
    if len(selected_levels) < level_count:
        for i in range(len(selected_levels) + 1, level_count + 1):
            selected_levels.append(f"Level {i}")

    course_ids = {}
    level_ids = {}
    for cname in course_names:
        course_ids[cname] = ensure_course(conn, cname)
        for level_name in selected_levels:
            level_ids[(cname, level_name)] = ensure_level(conn, course_ids[cname], level_name)

    classrooms = []
    for i in range(1, classroom_count + 1):
        floor = 100 + (((i - 1) // 4) + 1) * 100
        room_name = f"R{floor + ((i - 1) % 4) + 1}"
        ensure_classroom(conn, room_name)
        classrooms.append(room_name)

    slots = [
        ("16:25-17:20", "16:25", "17:20"),
        ("17:25-18:20", "17:25", "18:20"),
        ("18:30-19:25", "18:30", "19:25"),
        ("19:35-20:30", "19:35", "20:30"),
        ("20:35-21:30", "20:35", "21:30"),
    ]
    for label, st, et in slots:
        ensure_time_slot(conn, label, st, et)

    class_ids = []
    for i in range(1, class_count + 1):
        cname = course_names[(i - 1) % len(course_names)]
        level_name = selected_levels[((i - 1) // len(course_names)) % len(selected_levels)]
        foreign_teacher_id = foreign_ids[(i - 1) % len(foreign_ids)]
        chinese_teacher_id = chinese_ids[(i - 1) % len(chinese_ids)]
        class_name = f"{cname} {level_name} {chr(64 + ((i - 1) % 26) + 1)}"
        credit_unit = 1.0 if i % 3 else 1.5
        status_v = "active" if i % 11 else "inactive"
        memo = f"Auto-seeded demo class {i}"
        class_id = ensure_class(conn, class_name, course_ids[cname], level_ids[(cname, level_name)], foreign_teacher_id, chinese_teacher_id, credit_unit, status_v, memo)
        class_ids.append(class_id)

    english_given = ["Leo", "Hank", "Shawn", "Luna", "Momo", "Finn", "Lynn", "Ray", "Jane", "Mina", "Tom", "Yuri", "Hans", "Ethan", "Mark", "Cindy", "Liam", "Ann", "Yuna", "Jay", "Wade", "Nora", "Kiki", "Yoyo", "Bobby", "Nina", "Howard", "Lily", "Ruby", "Chris"]

    def student_name_pair(index):
        family = chinese_family[(index - 1) % len(chinese_family)]
        given = chinese_given[((index - 1) // len(chinese_family)) % len(chinese_given)]
        en = english_given[(index - 1) % len(english_given)]
        return f"{family} {given}", f"{en} {index:03d}"

    student_user_ids = []
    active_cut = max(1, int(student_count * 0.82))
    leave_cut = max(active_cut + 1, int(student_count * 0.92))
    for i in range(1, student_count + 1):
        name_ko, name_en = student_name_pair(i)
        uid = ensure_user_account(conn, name_ko, f"demo_student_{i:03d}", ROLE_STUDENT)
        student_user_ids.append(uid)
        guardian_idx = (i - 1) % len(parent_ids)
        guardian_user = conn.execute("SELECT name FROM users WHERE id=?", (parent_ids[guardian_idx],)).fetchone()
        class_id = class_ids[(i - 1) % len(class_ids)]
        if i <= active_cut:
            status_v = "active"
        elif i <= leave_cut:
            status_v = "leave"
        else:
            status_v = "ended"
        credits = float(6 + (i % 10))
        homeroom_teacher_id = foreign_ids[(i - 1) % len(foreign_ids)] if i % 2 else chinese_ids[(i - 1) % len(chinese_ids)]
        leave_start = "2026-01-10" if status_v == "leave" else None
        leave_end = "2026-02-01" if status_v == "leave" else None
        row = conn.execute("SELECT id FROM students WHERE user_id=?", (uid,)).fetchone()
        if row:
            conn.execute(
                """UPDATE students SET student_no=?, name_ko=?, name_en=?, phone=?, guardian_name=?, guardian_phone=?,
                current_class_id=?, homeroom_teacher_id=?, remaining_credits=?, status=?, enrolled_at=?, leave_start_date=?, leave_end_date=?, memo=?, updated_at=?
                WHERE user_id=?""",
                (
                    f"ST{i:04d}", name_ko, name_en, f"138{i:08d}"[-11:], guardian_user["name"], f"139{i:08d}"[-11:],
                    class_id, homeroom_teacher_id, credits, status_v, "2025-09-01", leave_start, leave_end,
                    "Auto-seeded demo student", now(), uid,
                ),
            )
        else:
            conn.execute(
                """INSERT INTO students(user_id, student_no, name_ko, name_en, phone, guardian_name, guardian_phone, current_class_id,
                homeroom_teacher_id, remaining_credits, status, enrolled_at, leave_start_date, leave_end_date, memo, created_at, updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    uid, f"ST{i:04d}", name_ko, name_en, f"138{i:08d}"[-11:], guardian_user["name"], f"139{i:08d}"[-11:],
                    class_id, homeroom_teacher_id, credits, status_v, "2025-09-01", leave_start, leave_end,
                    "Auto-seeded demo student", now(), now(),
                ),
            )

    week_start = iso_monday_str(datetime.utcnow().date().isoformat(), 0)
    weekday_cycle = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    slot_cycle = [(st, et) for _, st, et in slots[:4]]
    available_slots = [(day, st, et) for day in weekday_cycle for st, et in slot_cycle]

    if force and class_ids:
        placeholders = ",".join(["?"] * len(class_ids))
        conn.execute(f"DELETE FROM attendance WHERE class_id IN ({placeholders})", tuple(class_ids))
        conn.execute(f"DELETE FROM schedules WHERE class_id IN ({placeholders})", tuple(class_ids))

    schedule_rows = []
    teacher_slot_index = {tid: 0 for tid in foreign_ids}
    used_teacher_slots = set()
    used_room_slots = set()
    for idx, cls_id in enumerate(class_ids):
        lesson_count = 2 if idx % 4 == 0 else 3
        teacher_id = foreign_ids[idx % len(foreign_ids)]
        cursor = teacher_slot_index[teacher_id]
        for lesson_idx in range(lesson_count):
            assigned = None
            for slot_offset in range(len(available_slots)):
                day, st, et = available_slots[(cursor + slot_offset) % len(available_slots)]
                if (teacher_id, day, st, et) in used_teacher_slots:
                    continue
                for room_offset in range(len(classrooms)):
                    room = classrooms[(idx + lesson_idx + room_offset) % len(classrooms)]
                    if (room, day, st, et) in used_room_slots:
                        continue
                    assigned = (day, st, et, room, (cursor + slot_offset + 1) % len(available_slots))
                    break
                if assigned:
                    break
            if not assigned:
                raise ValueError(f"No free timetable slot for teacher {teacher_id} class {cls_id}")
            day, st, et, room, next_cursor = assigned
            sc_id = ensure_schedule_row(conn, cls_id, day, st, et, room, teacher_id, week_start)
            used_teacher_slots.add((teacher_id, day, st, et))
            used_room_slots.add((room, day, st, et))
            teacher_slot_index[teacher_id] = next_cursor
            schedule_rows.append((sc_id, cls_id, day, st, et, room))

    class_students = {}
    for r in conn.execute("SELECT user_id, current_class_id FROM students WHERE current_class_id IS NOT NULL ORDER BY id").fetchall():
        class_students.setdefault(r["current_class_id"], []).append(r["user_id"])

    recent_days = [0, 1, 2, 3]
    attendance_seeded = 0
    for sc_id, cls_id, day, st, et, room in schedule_rows:
        students = class_students.get(cls_id, [])[:12]
        class_credit = conn.execute("SELECT COALESCE(credit_unit,1) AS credit_unit FROM classes WHERE id=?", (cls_id,)).fetchone()["credit_unit"]
        for sid_idx, sid in enumerate(students):
            for back in recent_days:
                lesson_date = (datetime.utcnow().date() - timedelta(days=back + (sid_idx % 2))).isoformat()
                status_v = "present"
                charge = None
                requires = 0
                makeup_done = 0
                if (sid_idx + back) % 9 == 0:
                    status_v = "absent"
                    charge = "deduct" if (sid_idx % 2 == 0) else "no_deduct"
                    requires = 1
                    makeup_done = 1 if (sid_idx % 4 == 0) else 0
                elif (sid_idx + back) % 6 == 0:
                    status_v = "late"
                elif (sid_idx + back) % 13 == 0:
                    status_v = "makeup"
                delta = calc_credit_delta(status_v, charge, class_credit)
                exists = conn.execute("SELECT id FROM attendance WHERE schedule_id=? AND student_id=? AND lesson_date=?", (sc_id, sid, lesson_date)).fetchone()
                if exists:
                    conn.execute(
                        """UPDATE attendance SET status=?, class_id=?, note=?, created_by=?, participation_score=?, fluency_score=?,
                        vocabulary_score=?, reading_score=?, homework_score=?, attitude_score=?, teacher_memo=?, absence_charge_type=?,
                        requires_makeup=?, makeup_completed=?, credit_delta=? WHERE id=?""",
                        (
                            status_v, cls_id, "demo attendance", owner_id,
                            3 + ((sid_idx + 1) % 3), 2 + ((sid_idx + back) % 4), 3 + (sid_idx % 3), 2 + ((sid_idx + 2) % 4),
                            3 + ((sid_idx + 3) % 3), 3 + ((sid_idx + 4) % 3), "Demo lesson memo", charge, requires, makeup_done, delta, exists["id"],
                        ),
                    )
                else:
                    conn.execute(
                        """INSERT INTO attendance(class_id, student_id, lesson_date, status, note, created_by, created_at, schedule_id,
                        participation_score, fluency_score, vocabulary_score, reading_score, homework_score, attitude_score, teacher_memo,
                        absence_charge_type, requires_makeup, makeup_completed, credit_delta)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            cls_id, sid, lesson_date, status_v, "demo attendance", owner_id, now(), sc_id,
                            3 + ((sid_idx + 1) % 3), 2 + ((sid_idx + back) % 4), 3 + (sid_idx % 3), 2 + ((sid_idx + 2) % 4),
                            3 + ((sid_idx + 3) % 3), 3 + ((sid_idx + 4) % 3), "Demo lesson memo", charge, requires, makeup_done, delta,
                        ),
                    )
                    attendance_seeded += 1

    homework_seeded = 0
    for cls_id in class_ids[:min(class_count, 20)]:
        teacher_id = conn.execute("SELECT COALESCE(foreign_teacher_id, teacher_id) AS tid FROM classes WHERE id=?", (cls_id,)).fetchone()["tid"]
        for hidx in range(1, 3):
            title = f"Week Task {hidx} - Class {cls_id}"
            due_date = (datetime.utcnow().date() + timedelta(days=3 + hidx)).isoformat()
            hrow = conn.execute("SELECT id FROM homework WHERE class_id=? AND title=?", (cls_id, title)).fetchone()
            if hrow:
                hw_id = hrow["id"]
                conn.execute("UPDATE homework SET description=?, due_date=?, teacher_id=?, updated_at=? WHERE id=?", ("Demo homework content", due_date, teacher_id, now(), hw_id))
            else:
                conn.execute(
                    "INSERT INTO homework(class_id, teacher_id, title, description, due_date, status, created_by, created_at, updated_at) VALUES(?,?,?,?,?,'active',?,?,?)",
                    (cls_id, teacher_id, title, "Demo homework content", due_date, owner_id, now(), now()),
                )
                hw_id = conn.execute("SELECT id FROM homework WHERE class_id=? AND title=?", (cls_id, title)).fetchone()["id"]
            students = class_students.get(cls_id, [])[:12]
            for sidx, sid in enumerate(students):
                sub = conn.execute("SELECT id FROM homework_submissions WHERE homework_id=? AND student_id=?", (hw_id, sid)).fetchone()
                submitted = 1 if (sidx + hidx) % 3 != 0 else 0
                feedback = "Good progress" if submitted else "Please submit by next class"
                submitted_at = now() if submitted else None
                if sub:
                    conn.execute(
                        "UPDATE homework_submissions SET submitted=?, submitted_at=?, feedback=?, feedback_teacher_id=?, updated_at=? WHERE id=?",
                        (submitted, submitted_at, feedback, teacher_id, now(), sub["id"]),
                    )
                else:
                    conn.execute(
                        "INSERT INTO homework_submissions(homework_id, student_id, submitted, submitted_at, feedback, feedback_teacher_id, updated_at) VALUES(?,?,?,?,?,?,?)",
                        (hw_id, sid, submitted, submitted_at, feedback, teacher_id, now()),
                    )
                    homework_seeded += 1

    exam_seeded = 0
    for cls_id in class_ids[:min(class_count, 16)]:
        exam_name = f"Monthly Check {cls_id}"
        exam_date = (datetime.utcnow().date() - timedelta(days=7)).isoformat()
        ex = conn.execute("SELECT id FROM exams WHERE class_id=? AND name=?", (cls_id, exam_name)).fetchone()
        if ex:
            exam_id = ex["id"]
            conn.execute("UPDATE exams SET exam_date=? WHERE id=?", (exam_date, exam_id))
        else:
            conn.execute("INSERT INTO exams(class_id, name, exam_date, report, created_at) VALUES(?,?,?,?,?)", (cls_id, exam_name, exam_date, "Demo exam", now()))
            exam_id = conn.execute("SELECT id FROM exams WHERE class_id=? AND name=?", (cls_id, exam_name)).fetchone()["id"]
        students = class_students.get(cls_id, [])[:12]
        for sidx, sid in enumerate(students):
            score = float(65 + ((sidx * 7 + cls_id) % 31))
            erow = conn.execute("SELECT id FROM exam_scores WHERE exam_id=? AND student_id=?", (exam_id, sid)).fetchone()
            if erow:
                conn.execute("UPDATE exam_scores SET score=? WHERE id=?", (score, erow["id"]))
            else:
                conn.execute("INSERT INTO exam_scores(exam_id, student_id, score, created_at) VALUES(?,?,?,?)", (exam_id, sid, score, now()))
                exam_seeded += 1

    payment_seeded = 0
    for idx, sid in enumerate(student_user_ids):
        payment_cycles = 2 if idx % 3 else 3
        for pidx in range(payment_cycles):
            paid_date = (datetime.utcnow().date() - timedelta(days=20 - (idx % 10) - pidx * 14)).isoformat()
            amount = 1200 + (idx % 6) * 150 + pidx * 80
            row = conn.execute("SELECT id FROM payments WHERE student_id=? AND paid_date=? AND amount=?", (sid, paid_date, amount)).fetchone()
            if not row:
                conn.execute(
                    "INSERT INTO payments(student_id, paid_date, amount, package_hours, remaining_classes, created_at) VALUES(?,?,?,?,?,?)",
                    (sid, paid_date, amount, 24, 8 + (idx % 6), now()),
                )
                payment_seeded += 1

    counseling_seeded = 0
    for idx, sid in enumerate(student_user_ids[:max(20, student_count // 3)]):
        parent_id = parent_ids[idx % len(parent_ids)]
        memo = f"Seeded counseling note {idx + 1}"
        existing = conn.execute("SELECT id FROM counseling WHERE student_id=? AND memo=?", (sid, memo)).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO counseling(student_id, parent_id, memo, is_special_note, created_by, created_at) VALUES(?,?,?,?,?,?)",
                (sid, parent_id, memo, 1 if idx % 7 == 0 else 0, owner_id, now()),
            )
            counseling_seeded += 1

    book_seeded = 0
    loan_seeded = 0
    for i in range(1, book_count + 1):
        code = f"BK{i:04d}"
        title = f"Reading Book {i:04d}"
        brow = conn.execute("SELECT id, status FROM books WHERE code=?", (code,)).fetchone()
        if brow:
            book_id = brow["id"]
        else:
            conn.execute("INSERT INTO books(code, title, status, created_at) VALUES(?,?,?,?)", (code, title, "available", now()))
            book_id = conn.execute("SELECT id FROM books WHERE code=?", (code,)).fetchone()["id"]
            book_seeded += 1
        if i <= min(book_count, max(30, student_count // 4)):
            sid = student_user_ids[(i - 1) % len(student_user_ids)]
            existing_loan = conn.execute("SELECT id FROM book_loans WHERE book_id=? AND student_id=?", (book_id, sid)).fetchone()
            if not existing_loan:
                conn.execute(
                    "INSERT INTO book_loans(book_id, student_id, loaned_at, handled_by, created_at) VALUES(?,?,?,?,?)",
                    (book_id, sid, now(), owner_id, now()),
                )
                conn.execute("UPDATE books SET status='borrowed' WHERE id=?", (book_id,))
                loan_seeded += 1

    announcement_seeded = 0
    for i in range(1, announcement_count + 1):
        title = f"Demo Announcement {i:02d}"
        existing = conn.execute("SELECT id FROM announcements WHERE title=?", (title,)).fetchone()
        if not existing:
            conn.execute("INSERT INTO announcements(title, content, created_by, created_at) VALUES(?,?,?,?)", (title, f"Seeded announcement content {i}", owner_id, now()))
            announcement_seeded += 1

    return {
        "seeded": True,
        "preset": preset or "default",
        "teachers_foreign": len(foreign_ids),
        "teachers_chinese": len(chinese_ids),
        "parents": len(parent_ids),
        "students": len(student_user_ids),
        "classes": len(class_ids),
        "schedules": len(schedule_rows),
        "attendance_rows": attendance_seeded,
        "homework_rows": homework_seeded,
        "exam_rows": exam_seeded,
        "payment_rows": payment_seeded,
        "counseling_rows": counseling_seeded,
        "books": book_count,
        "book_rows": book_seeded,
        "loan_rows": loan_seeded,
        "announcement_rows": announcement_seeded,
    }
def init_db():
    conn = get_db()

    with open(os.path.join(BASE_DIR, "schema.sql"), "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    ensure_schedule_columns(conn)
    ensure_attendance_columns(conn)
    ensure_homework_columns(conn)
    ensure_master_tables(conn)
    ensure_extended_columns(conn)
    ensure_teacher_table(conn)
    ensure_logs_table(conn)
    ensure_logs_columns(conn)
    cur = conn.execute("SELECT COUNT(*) AS c FROM users")
    if cur.fetchone()["c"] == 0:
        users = [
            ("원장", "owner", "owner", "1234"),
            ("매니저", "manager", "manager", "1234"),
            ("강사", "teacher", "teacher", "1234"),
            ("학부모", "parent", "parent", "1234"),
            ("학생", "student", "student", "1234"),
        ]
        for name, username, role, pw in users:
            conn.execute(
                "INSERT INTO users(name, username, password_hash, role, created_at) VALUES(?,?,?,?,?)",
                (name, username, hash_pw(pw), role, now()),
            )
    repair_profile_integrity(conn)
    seed_demo_data(conn, force=False)
    conn.commit()
    conn.close()

def parse_query(environ):
    return {k: v[0] for k, v in parse_qs(environ.get("QUERY_STRING", "")).items()}
def get_lang(environ):
    q = parse_query(environ)
    lang = q.get("lang", "").strip().lower()
    if lang in SUPPORTED_LANGS:
        return lang
    cookies = parse_cookie(environ.get("HTTP_COOKIE", ""))
    cookie_lang = (cookies.get("lang") or "").strip().lower()
    if cookie_lang in SUPPORTED_LANGS:
        return cookie_lang
    return "en"
def parse_body(environ):
    try:
        length = int(environ.get("CONTENT_LENGTH") or 0)
    except ValueError:
        length = 0
    body = environ["wsgi.input"].read(length) if length > 0 else b""
    ctype = environ.get("CONTENT_TYPE", "")
    if "application/json" in ctype:
        return json.loads(body.decode("utf-8") or "{}")
    return {k: v[0] for k, v in parse_qs(body.decode("utf-8")).items()}

def parse_multipart_form(environ):
    data = {}
    files = {}
    try:
        length = int(environ.get("CONTENT_LENGTH") or 0)
    except ValueError:
        length = 0
    body = environ["wsgi.input"].read(length) if length > 0 else b""
    ctype = environ.get("CONTENT_TYPE", "")
    if not body or "multipart/form-data" not in ctype:
        return data, files

    header_bytes = (
        f"Content-Type: {ctype}\r\n"
        "MIME-Version: 1.0\r\n"
        "\r\n"
    ).encode("utf-8")
    message = BytesParser(policy=policy.default).parsebytes(header_bytes + body)
    if not message.is_multipart():
        return data, files

    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        filename = part.get_filename()
        content = part.get_payload(decode=True) or b""
        if filename:
            files[name] = {"filename": filename, "content": content}
        else:
            charset = part.get_content_charset() or "utf-8"
            data[name] = content.decode(charset, errors="replace")
    return data, files
def parse_cookie(cookie):
    out = {}
    if not cookie:
        return out
    for part in cookie.split(";"):
        if "=" in part:
            k, v = part.strip().split("=", 1)
            out[k] = v
    return out
def current_user(environ):
    cookies = parse_cookie(environ.get("HTTP_COOKIE", ""))
    token = cookies.get("session")
    if not token:
        return None
    conn = get_db()
    row = conn.execute(
        "SELECT s.created_at AS session_created_at, u.* FROM sessions s JOIN users u ON u.id=s.user_id WHERE s.token=?", (token,)
    ).fetchone()
    if not row:
        conn.close()
        return None
    created_at = parse_iso_datetime(row["session_created_at"])
    if created_at is None:
        invalidate_session(conn, token)
        conn.commit()
        conn.close()
        return None
    age = (datetime.now(timezone.utc) - created_at.astimezone(timezone.utc)).total_seconds()
    if age > session_ttl_seconds():
        invalidate_session(conn, token)
        conn.commit()
        conn.close()
        return None
    conn.close()
    return row

def render_html(title, body, user=None, lang=None, current_menu=None, flash_msg="", flash_type="success"):
    lang = lang or CURRENT_LANG
    css = """
    <style>
      html { -webkit-text-size-adjust: 100%; }
      * { box-sizing: border-box; }
      body { margin:0; font-family:Arial, sans-serif; background:#f5f7fb; color:#1f2937; line-height:1.5; font-size:16px; }
      .app { display:flex; min-height:100vh; }
      .sidebar { width:240px; background:#111827; color:#e5e7eb; padding:18px 14px; flex-shrink:0; }
      .brand { font-size:18px; font-weight:700; margin-bottom:18px; }
      .nav-link { display:block; color:#cbd5e1; text-decoration:none; padding:8px 10px; border-radius:8px; margin-bottom:6px; }
      .nav-link:hover { background:#1f2937; color:white; }
      .nav-link.active { background:#2563eb; color:white; }
      .main { flex:1; min-width:0; padding:18px 22px; }
      .page-container { max-width:1400px; margin:0 auto; }
      .topbar { background:white; border:1px solid #e5e7eb; border-radius:12px; padding:10px 14px; display:flex; justify-content:space-between; align-items:center; gap:10px; margin-bottom:14px; }
      .page-title { margin:0 0 14px 0; font-size:24px; overflow-wrap:anywhere; }
      .card { background:white; border:1px solid #e5e7eb; border-radius:12px; padding:14px; margin-bottom:14px; overflow:hidden; }
      .card h3, .card h4 { margin:0 0 10px 0; overflow-wrap:anywhere; }
      .card-header-row { display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:10px; }
      .card-header-row h4 { margin:0; }
      .mobile-stack { display:flex; flex-direction:column; gap:10px; }
      .form-row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
      .form-row label { display:flex; align-items:center; gap:6px; flex-wrap:wrap; }
      .filter-grid { display:grid; grid-template-columns:repeat(4,minmax(180px,1fr)); gap:10px; }
      .filter-grid > label, .filter-grid > div { display:flex; flex-direction:column; gap:6px; }
      .btn-row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
      .filter-row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
      .filter-row label { display:flex; align-items:center; gap:6px; flex-wrap:wrap; }
      input, select, textarea { max-width:100%; border:1px solid #d1d5db; border-radius:8px; padding:10px 12px; min-height:44px; line-height:1.4; }
      textarea { width:100%; }
      button, .btn { background:#2563eb; color:white; border:none; border-radius:8px; padding:11px 14px; min-height:44px; line-height:1.3; text-decoration:none; display:inline-flex; align-items:center; justify-content:center; }
      .btn.secondary { background:#6b7280; }
      .table-wrap { width:100%; overflow-x:auto; -webkit-overflow-scrolling:touch; border:1px solid #e5e7eb; border-radius:12px; background:white; }
      table { width:100%; border-collapse:collapse; background:white; }
      th, td { border:1px solid #e5e7eb; padding:10px; text-align:left; vertical-align:top; word-break:normal; overflow-wrap:normal; white-space:normal; }
      th { background:#f9fafb; }
      .card > table, .card > .admin-table { display:block; width:100%; overflow-x:auto; -webkit-overflow-scrolling:touch; border:1px solid #e5e7eb; border-radius:12px; }
      .card > table > tbody, .card > table > thead, .card > table > tfoot { width:100%; }
      .table-wrap table, .card > table { table-layout:auto; min-width:840px; }
      .table-wrap table th, .table-wrap table td, .card > table th, .card > table td { min-width:96px; }
      .table-wrap table th:first-child, .table-wrap table td:first-child, .card > table th:first-child, .card > table td:first-child { min-width:72px; }
      .col-class { min-width:160px; }
      .col-course { min-width:140px; }
      .col-level { min-width:120px; }
      .col-teacher { min-width:150px; }
      .col-students { min-width:220px; }
      .badge { display:inline-block; padding:3px 8px; border-radius:999px; font-size:12px; background:#e5e7eb; }
      .badge.active { background:#dcfce7; color:#166534; }
      .badge.leave { background:#fef3c7; color:#92400e; }
      .badge.ended { background:#fee2e2; color:#991b1b; }
      .empty-msg { color:#6b7280; font-style:italic; }
      .flash { border-radius:10px; padding:10px 12px; margin-bottom:12px; }
      .flash.success { background:#ecfdf5; color:#065f46; border:1px solid #a7f3d0; }
      .flash.error { background:#fef2f2; color:#991b1b; border:1px solid #fecaca; }
      .timetable-wrap { overflow:auto; border:1px solid #e5e7eb; border-radius:12px; background:white; }
      .timetable-grid { --schedule-row-width:150px; --schedule-col-min:170px; min-width:860px; display:grid; gap:0; }
      .tt-head { background:#f3f4f6; font-weight:600; padding:8px 10px; border-bottom:1px solid #e5e7eb; border-right:1px solid #e5e7eb; font-size:13px; }
      .tt-cell { min-height:92px; border-right:1px solid #e5e7eb; border-bottom:1px solid #e5e7eb; padding:4px; background:#fff; }
      .tt-rowhead { background:#f9fafb; padding:8px 10px; border-right:1px solid #e5e7eb; border-bottom:1px solid #e5e7eb; min-width:150px; font-size:13px; }
      .tt-rowhead strong { display:block; font-size:15px; line-height:1.2; }
      .lesson-block { border:1px solid #bfdbfe; background:#eff6ff; border-radius:8px; padding:8px; margin-bottom:4px; font-size:12px; line-height:1.3; }
      .lesson-title { font-size:13px; font-weight:700; line-height:1.2; margin-bottom:2px; }
      .lesson-meta { color:#475569; font-size:11px; line-height:1.25; }
      .student-line { color:#334155; font-size:11px; line-height:1.3; margin-top:4px; }
      .lesson-actions { display:flex; flex-wrap:wrap; gap:4px; margin-top:6px; }
      .lesson-main-actions, .lesson-sub-actions { display:flex; flex-wrap:wrap; gap:4px; margin-top:6px; }
      .lesson-main-actions .btn, .lesson-sub-actions .mini-link { min-height:28px; padding:5px 8px; border-radius:6px; font-size:11px; }
      .mini-link { font-size:11px; padding:5px 8px; border-radius:6px; text-decoration:none; background:#dbeafe; color:#1e3a8a; display:inline-block; }
      .schedule-editor-grid { display:grid; grid-template-columns:minmax(0,1.35fr) minmax(320px,0.65fr); gap:14px; }
      .print-only { display:none; }
      .schedule-print-head { margin-bottom:10px; padding-bottom:8px; border-bottom:1px solid #cbd5e1; }
      .schedule-print-title { font-size:18px; font-weight:700; margin-bottom:4px; }
      .schedule-print-meta { display:flex; flex-wrap:wrap; gap:14px; font-size:12px; color:#475569; }
      .print-room-grid { display:grid; grid-template-columns:repeat(2, minmax(0,1fr)); gap:10px; }
      .print-room-card { border:1px solid #94a3b8; border-radius:8px; overflow:hidden; background:#fff; break-inside:avoid; page-break-inside:avoid; }
      .print-room-title { background:#d9f99d; font-size:13px; font-weight:700; padding:5px 8px; border-bottom:1px solid #94a3b8; }
      .print-room-table { width:100%; border-collapse:collapse; table-layout:fixed; }
      .print-room-table th, .print-room-table td { border:1px solid #cbd5e1; padding:4px; vertical-align:top; font-size:11px; }
      .print-room-table th { background:#e0f2fe; text-align:center; }
      .print-lesson-cell { min-height:42px; }
      .print-class-name { font-weight:700; font-size:11px; line-height:1.15; }
      .print-class-meta { font-size:10px; color:#475569; line-height:1.15; }
      .print-students { font-size:10px; color:#334155; line-height:1.15; margin-top:2px; }
      .score-input { width:88px; min-width:88px; }
      .memo-input { min-width:220px; }
      .sticky-head thead th { position:sticky; top:0; z-index:2; background:#f9fafb; }
      .student-summary-grid { display:grid; grid-template-columns:repeat(2,minmax(220px,1fr)); gap:10px; }
      .student-summary-item { background:#f8fafc; border:1px solid #e5e7eb; border-radius:10px; padding:10px; }
      .student-summary-label { font-size:12px; color:#6b7280; margin-bottom:4px; }
      .student-summary-value { font-size:16px; font-weight:600; overflow-wrap:anywhere; }
      .two-col { display:grid; grid-template-columns:2fr 1fr; gap:14px; }
      .muted { color:#6b7280; font-size:12px; overflow-wrap:anywhere; }
      @media (max-width: 1100px) {
        .two-col { grid-template-columns:1fr; }
        .schedule-editor-grid { grid-template-columns:1fr; }
      }
      @media (max-width: 900px) {
        body { font-size:15px; line-height:1.5; }
        .app { flex-direction:column; }
        .sidebar { width:100%; padding:12px; display:flex; flex-wrap:wrap; gap:6px; align-items:center; }
        .brand { margin:0 8px 8px 0; width:100%; font-size:19px; }
        .nav-link { display:inline-block; margin:0; font-size:14px; padding:9px 11px; }
        .main { padding:12px; }
        .topbar { flex-direction:column; align-items:flex-start; gap:12px; }
        .topbar, .topbar select, .topbar .btn { font-size:14px; }
        .card { padding:16px; }
        .card h3, .card h4 { font-size:19px; line-height:1.35; }
        th, td { font-size:14px; line-height:1.45; }
        button, .btn, input, select, textarea { font-size:15px; }
        .tt-cell { min-height:110px; }
      }
      @media (max-width: 768px) {
        body { font-size:16px; line-height:1.55; }
        .page-title { font-size:24px; line-height:1.3; margin-bottom:16px; }
        h1, h2, h3 { line-height:1.3; }
        h1 { font-size:30px; }
        h2 { font-size:26px; }
        h3 { font-size:22px; }
        .card { padding:16px; margin-bottom:16px; }
        .card h3, .card h4 { font-size:20px; line-height:1.35; margin-bottom:12px; }
        .muted { font-size:14px; line-height:1.5; }
        .filter-row, .form-row { flex-direction:column; align-items:stretch; gap:12px; }
        .filter-row > *, .form-row > * { width:100%; }
        .filter-row label, .form-row label { width:100%; display:flex; flex-direction:column; align-items:stretch; gap:6px; font-size:15px; line-height:1.45; margin-bottom:2px; }
        .filter-grid { grid-template-columns:1fr; }
        .btn-row { flex-wrap:wrap; }
        .btn-row .btn, .btn-row button { flex:1 1 180px; }
        .filter-row input, .filter-row select, .filter-row textarea, .filter-row button, .filter-row .btn,
        .form-row input, .form-row select, .form-row textarea, .form-row button, .form-row .btn { width:100%; min-height:44px; font-size:16px; line-height:1.35; }
        th, td { white-space:nowrap; font-size:15px; line-height:1.5; padding:11px 9px; }
        .lesson-block { font-size:15px; line-height:1.5; padding:12px; }
        .lesson-main-actions .btn { min-width:140px; }
        .mini-link { font-size:13px; padding:7px 9px; }
        .flash { font-size:15px; line-height:1.45; }
        .timetable-wrap { overflow-x:auto; }
        .student-summary-grid { grid-template-columns:1fr; }
        .print-room-grid { grid-template-columns:1fr; }
      }
      @page { size:A4 landscape; margin:6mm; }
      @media print {
        body { background:#fff; color:#111827; font-size:9px; }
        .app, .main, .page-container { display:block; padding:0; margin:0; max-width:none; }
        .sidebar, .topbar, .page-title, .filter-print-hide, .schedule-editor-grid, .flash, .screen-only, details.card, .lesson-main-actions, .lesson-sub-actions, .lesson-actions, button, .btn { display:none !important; }
        .card { border:none; border-radius:0; padding:0; margin:0 0 5px 0; background:#fff; }
        .card-header-row { margin-bottom:4px; }
        .print-only { display:block !important; }
        .timetable-wrap, .table-wrap { overflow:visible; border:none; border-radius:0; }
        .timetable-grid { --schedule-row-width:64px; --schedule-col-min:54px; min-width:0; width:100%; page-break-inside:auto; }
        .schedule-print-head { margin-bottom:6px; padding-bottom:4px; }
        .schedule-print-title { font-size:13px; margin-bottom:2px; }
        .schedule-print-meta { gap:8px; font-size:9px; }
        .print-room-grid { grid-template-columns:repeat(2, minmax(0,1fr)); gap:5px; }
        .print-room-card { border:1px solid #94a3b8; border-radius:0; }
        .print-room-title { font-size:9px; padding:3px 4px; }
        .print-room-table th, .print-room-table td { padding:2px; font-size:7px; }
        .print-lesson-cell { min-height:28px; }
        .print-class-name { font-size:7px; }
        .print-class-meta, .print-students { font-size:6px; line-height:1.05; }
        .tt-head { padding:4px 5px; font-size:9px; }
        .tt-cell { min-height:52px; padding:2px; page-break-inside:auto; }
        .tt-rowhead { min-width:78px; padding:4px 5px; font-size:9px; page-break-inside:avoid; }
        .tt-rowhead strong { font-size:10px; }
        .muted { font-size:8px; }
        .lesson-block { background:#fff; border:1px solid #cbd5e1; padding:3px 4px; margin-bottom:2px; break-inside:avoid; page-break-inside:avoid; }
        .lesson-title { font-size:9px; margin-bottom:1px; }
        .lesson-meta, .student-line { font-size:7px; line-height:1.1; }
        .student-line { margin-top:2px; }
        .badge { padding:1px 4px; font-size:7px; }
      }
    </style>
    """

    layout = body
    if user:
        lang_options = "".join([
            f"<option value='{code}' {'selected' if lang == code else ''}>{label}</option>"
            for code, label in LANG_LABELS.items()
        ])
        keys = ROLE_MENU_KEYS.get(user["role"], ["dashboard"])
        menu_links = "".join([
            f"<a class='nav-link {'active' if current_menu==k else ''}' href='{NAV_PATHS[k]}?lang={lang}'>{menu_t(k, lang)}</a>"
            for k in keys
        ])
        flash_html = f"<div class='flash {flash_type}'>{flash_msg}</div>" if flash_msg else ""
        layout = f"""
        <div class='app'>
          <aside class='sidebar'>
            <div class='brand'>ReadingTown LMS</div>
            {menu_links}
          </aside>
          <main class='main'>
            <div class='page-container'>
            <div class='topbar'>
              <div>{menu_t('login_as', lang)}: <strong>{user['name']}</strong> ({role_label(user['role'], lang)})</div>
              <div>
                <span>{menu_t('lang', lang)}:</span>
                <select onchange="window.location.href=window.location.pathname+'?lang='+this.value">{lang_options}</select>
                <a class='btn secondary' href='/logout'>{menu_t('logout', lang)}</a>
              </div>
            </div>
            <h2 class='page-title'>{title}</h2>
            {flash_html}
            {body}
            </div>
          </main>
        </div>
        """
    else:
        layout = f"<div style='max-width:980px;margin:24px auto'><h2 class='page-title'>{title}</h2>{body}</div>"
    scroll_js = """
    <script>
    (function(){
      var SCROLL_NS = 'rtweb:scroll:';
      var CONTEXT_KEYS = ['load','md_view','week','ref_date','day','selected_class_id','selected_student_id','selected_teacher_id','selected_homework_id','schedule_id','student_id','parent_id','status','date_from','date_to','lesson_mode'];

      function parseUrl(url){
        try {
          return new URL(url, window.location.origin);
        } catch (e) {
          return null;
        }
      }

      function contextFromSearchParams(sp){
        var parts = [];
        CONTEXT_KEYS.forEach(function(k){
          var v = sp.get(k);
          if (v !== null && v !== '') {
            parts.push(k + '=' + v);
          }
        });
        return parts.join('&');
      }

      function keyForUrl(url){
        var parsed = parseUrl(url || window.location.href);
        if (!parsed) {
          return SCROLL_NS + window.location.pathname;
        }
        var context = contextFromSearchParams(parsed.searchParams);
        return SCROLL_NS + parsed.pathname + (context ? (':' + context) : '');
      }

      function saveScrollFor(url){
        var y = window.scrollY || window.pageYOffset || 0;
        try {
          sessionStorage.setItem(keyForUrl(url), String(y));
        } catch (e) {}
      }

      function shouldPreserveForLink(a){
        if (!a || !a.getAttribute) return false;
        var href = a.getAttribute('href') || '';
        if (!href || href.startsWith('#') || href.startsWith('javascript:')) return false;
        var parsed = parseUrl(href);
        if (!parsed || parsed.origin !== window.location.origin) return false;
        if (a.dataset.preserveScroll === '1' || a.classList.contains('preserve-scroll') || a.classList.contains('picker-link') || a.classList.contains('admin-action-link')) return true;
        var inWorkingArea = !!a.closest('.table-wrap, form.query-form, form.picker-form, .card ul, .picker-list, .btn-row');
        if (!inWorkingArea) return false;
        var samePath = parsed.pathname === window.location.pathname;
        var sameSection = parsed.pathname.split('/')[1] === window.location.pathname.split('/')[1];
        return samePath || sameSection;
      }

      function shouldPreserveForForm(form){
        if (!form) return false;
        if (form.dataset.preserveScroll === '1') return true;
        var method = (form.getAttribute('method') || 'get').toLowerCase();
        if (form.classList.contains('query-form') || form.classList.contains('picker-form')) return true;
        if (method === 'post' && (form.dataset.preserveScroll === '1' || form.classList.contains('preserve-scroll-form'))) return true;
        if (method !== 'get') return false;
        return !!form.closest('.table-wrap');
      }

      try {
        var currentKey = keyForUrl(window.location.href);
        var restore = sessionStorage.getItem(currentKey);
        if (restore !== null) {
          window.scrollTo(0, parseInt(restore, 10) || 0);
          sessionStorage.removeItem(currentKey);
        }
      } catch (e) {}

      document.querySelectorAll('form').forEach(function(form){
        if (!shouldPreserveForForm(form)) return;
        form.addEventListener('submit', function(){
          var action = form.getAttribute('action') || window.location.href;
          saveScrollFor(action);
        });
      });

      document.querySelectorAll('a').forEach(function(a){
        if (!shouldPreserveForLink(a)) return;
        a.addEventListener('click', function(){
          saveScrollFor(a.href);
        });
      });
    })();
    </script>
    """
    return f"<html><head><meta charset='utf-8'><title>{title}</title>{css}</head><body>{layout}{scroll_js}</body></html>".encode("utf-8")
def require_login(environ):
    user = current_user(environ)
    if not user:
        return None, redirect('/login')
    return user, None
def redirect(path, headers=None):
    hdrs = [("Location", path)]
    if headers:
        hdrs.extend(headers)
    return "302 Found", hdrs, b""
def has_role(user, roles):
    return user and user["role"] in roles
def route_allowed(user, route_key):
    if not user:
        return False
    return route_key in ROLE_MENU_KEYS.get(user["role"], [])

def op_code(prefix, value):
    try:
        return f"{prefix}{int(value):05d}"
    except Exception:
        return f"{prefix}00000"
def forbidden_html(user, msg=None):
    if msg is None:
        msg = t("common.forbidden")
    html = render_html("403 Forbidden", f"<p style='color:red'>{msg}</p>", user)
    return "403 Forbidden", [("Content-Type", "text/html; charset=utf-8")], html
def forbidden_json(msg=t('common.forbidden')):
    return json_resp({"error": msg}, "403 Forbidden")
def can_view_student_row(user, st):
    if user["role"] in (ROLE_OWNER, ROLE_MANAGER):
        return True
    if user["role"] == ROLE_TEACHER:
        return st["current_class_teacher_id"] == user["id"]
    if user["role"] == ROLE_PARENT:
        return (st["guardian_name"] or "") == user["name"]
    if user["role"] == ROLE_STUDENT:
        return st["user_id"] == user["id"]
    return False
def json_resp(data, status="200 OK"):
    return status, [("Content-Type", "application/json; charset=utf-8")], json.dumps(data, ensure_ascii=False).encode("utf-8")
def text_resp(text, status="200 OK"):
    return status, [("Content-Type", "text/html; charset=utf-8")], text
def fetch_student_candidates(conn, keyword, limit=10):
    kw = (keyword or "").strip()
    if not kw:
        return []
    like = f"%{kw}%"
    return conn.execute(
        """SELECT id, student_no, name_ko, phone FROM students
        WHERE name_ko LIKE ? OR student_no LIKE ? OR phone LIKE ?
        ORDER BY id DESC LIMIT ?""",
        (like, like, like, limit),
    ).fetchall()
def summarize_student_names(names_text, threshold=38):
    names = [x.strip() for x in (names_text or "").split(",") if x and x.strip()]
    if not names:
        return "-", ""
    full = ", ".join(names)
    if len(full) <= threshold or len(names) <= 2:
        return full, full
    return ", ".join(names[:2]) + f" +{len(names)-2}", full


def fetch_class_candidates(conn, keyword, limit=10, show_all_when_empty=False):
    kw = (keyword or "").strip()
    if not kw and not show_all_when_empty:
        return []
    params = []
    where_sql = ""
    if kw:
        like = f"%{kw}%"
        where_sql = "WHERE c.name LIKE ? OR co.name LIKE ? OR l.name LIKE ? OR uf.name LIKE ? OR uc.name LIKE ?"
        params.extend([like, like, like, like, like])
    params.append(limit)
    return conn.execute(
        f"""SELECT c.id, c.name, COALESCE(co.name, '') AS course_name, COALESCE(l.name, '') AS level_name,
        COALESCE(uf.name, '') AS foreign_teacher_name,
        COALESCE(uc.name, '') AS chinese_teacher_name,
        COALESCE(uf.name, '') AS teacher_name,
        (SELECT COUNT(*) FROM students s WHERE s.current_class_id=c.id) AS student_count,
        (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=c.id ORDER BY s2.id) AS student_names
        FROM classes c
        LEFT JOIN courses co ON co.id=c.course_id
        LEFT JOIN levels l ON l.id=c.level_id
        LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
        LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
        {where_sql}
        ORDER BY c.id DESC LIMIT ?""",
        tuple(params),
    ).fetchall()
def list_teacher_profiles(conn, limit=200):
    return conn.execute(
        """SELECT t.user_id AS id, u.name, u.username, t.teacher_type
        FROM teachers t
        JOIN users u ON u.id=t.user_id
        WHERE u.role='teacher'
        ORDER BY t.id DESC
        LIMIT ?""",
        (limit,),
    ).fetchall()


def fetch_teacher_by_id(conn, teacher_user_id):
    if not teacher_user_id or not str(teacher_user_id).isdigit():
        return None
    return conn.execute(
        """SELECT t.user_id AS id, u.name, u.username, t.teacher_type
        FROM teachers t
        JOIN users u ON u.id=t.user_id
        WHERE t.user_id=? AND u.role='teacher'""",
        (teacher_user_id,),
    ).fetchone()


def fetch_teacher_candidates(conn, keyword, limit=10):
    kw = (keyword or "").strip()
    params = []
    where_sql = ""
    if kw:
        like = f"%{kw}%"
        where_sql = "AND (u.name LIKE ? OR u.username LIKE ?)"
        params.extend([like, like])
    params.append(limit)
    return conn.execute(
        f"""SELECT t.user_id AS id, u.name, u.username, t.teacher_type
        FROM teachers t
        JOIN users u ON u.id=t.user_id
        WHERE u.role='teacher' {where_sql}
        ORDER BY t.id DESC LIMIT ?""",
        tuple(params),
    ).fetchall()
def homeroom_display_name(name):
    return (name or "-").strip() or "-"


def find_teacher_id_by_name(conn, teacher_name):
    name = (teacher_name or "").strip()
    if not name:
        return None
    row = conn.execute(
        "SELECT id FROM users WHERE role='teacher' AND LOWER(TRIM(name))=LOWER(TRIM(?)) ORDER BY id LIMIT 1",
        (name,),
    ).fetchone()
    return row["id"] if row else None


def parse_leave_period_range(value):
    v = (value or "").strip()
    if not v:
        return None, None
    v = v.replace(" to ", "~").replace(" - ", "~")
    parts = [x.strip() for x in v.split("~") if x.strip()]
    if len(parts) == 1 and is_valid_date(parts[0]):
        return parts[0], parts[0]
    if len(parts) >= 2 and is_valid_date(parts[0]) and is_valid_date(parts[1]):
        return parts[0], parts[1]
    return None, None


def build_students_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "student_no", "chinese_name", "english_name", "phone", "guardian_name", "guardian_phone",
        "current_class", "homeroom_teacher", "remaining_credits", "status", "enrollment_date", "leave_period", "memo"
    ])
    for r in rows:
        writer.writerow([
            r.get("student_no") or "",
            r.get("name_ko") or "",
            r.get("name_en") or "",
            r.get("phone") or "",
            r.get("guardian_name") or "",
            r.get("guardian_phone") or "",
            r.get("class_name") or "",
            homeroom_display_name(r.get("homeroom_teacher_name")),
            r.get("remaining_credits") if r.get("remaining_credits") is not None else "",
            status_t(r.get("status")) if r.get("status") else "",
            r.get("enrolled_at") or "",
            f"{r.get('leave_start_date') or ''} ~ {r.get('leave_end_date') or ''}".strip(" ~"),
            r.get("memo") or "",
        ])
    return output.getvalue().encode("utf-8-sig")


def build_students_xlsx(rows):
    try:
        import openpyxl
    except Exception:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append([
        "student_no", "chinese_name", "english_name", "phone", "guardian_name", "guardian_phone",
        "current_class", "homeroom_teacher", "remaining_credits", "status", "enrollment_date", "leave_period", "memo"
    ])
    for r in rows:
        ws.append([
            r.get("student_no") or "",
            r.get("name_ko") or "",
            r.get("name_en") or "",
            str(r.get("phone") or ""),
            r.get("guardian_name") or "",
            str(r.get("guardian_phone") or ""),
            r.get("class_name") or "",
            homeroom_display_name(r.get("homeroom_teacher_name")),
            float(r.get("remaining_credits") or 0),
            status_t(r.get("status")) if r.get("status") else "",
            r.get("enrolled_at") or "",
            f"{r.get('leave_start_date') or ''} ~ {r.get('leave_end_date') or ''}".strip(" ~"),
            r.get("memo") or "",
        ])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


UPLOAD_TEMPLATE_HEADERS = [
    "student_no", "chinese_name", "english_name", "phone", "guardian_name", "guardian_phone",
    "class_name", "homeroom_teacher_name", "status", "enrollment_date", "leave_period", "memo"
]


def build_upload_template_xlsx():
    try:
        import openpyxl
    except Exception:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "upload_template"
    ws.append(UPLOAD_TEMPLATE_HEADERS)
    ws.append(["ST9001", "Li Hua", "Lily", "13800009999", "Parent 01", "13910009999", "Phonics Starter A", "Alex Carter", "active", "2026-03-01", "", "new student"])
    ws.append(["ST9002", "Wang Ming", "Mike", "13800008888", "Parent 02", "13910008888", "Reading Starter A", "Wang Li", "leave", "2026-01-15", "2026-02-01~2026-02-20", "medical leave"])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_upload_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(UPLOAD_TEMPLATE_HEADERS)
    writer.writerow(["ST9001", "Li Hua", "Lily", "13800009999", "Parent 01", "13910009999", "Phonics Starter A", "Alex Carter", "active", "2026-03-01", "", "new student"])
    writer.writerow(["ST9002", "Wang Ming", "Mike", "13800008888", "Parent 02", "13910008888", "Reading Starter A", "Wang Li", "leave", "2026-01-15", "2026-02-01~2026-02-20", "medical leave"])
    return output.getvalue().encode("utf-8-sig")


def read_upload_rows(filename, content_bytes):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content_bytes.decode("utf-8-sig", errors="ignore")
        rdr = csv.DictReader(io.StringIO(text))
        rows = []
        for row in rdr:
            cleaned = {}
            for k, v in (row or {}).items():
                cleaned[(k or "").strip()] = (v or "").strip()
            if any(cleaned.values()):
                rows.append(cleaned)
        return rdr.fieldnames or [], rows

    try:
        import openpyxl
    except Exception as e:
        raise ValueError("xlsx upload requires openpyxl") from e

    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = [str(x).strip() if x is not None else "" for x in values[0]]
    rows = []
    for row in values[1:]:
        if row is None:
            continue
        row_dict = {}
        has_any = False
        for idx, hname in enumerate(headers):
            if not hname:
                continue
            cell = row[idx] if idx < len(row) else ""
            val = "" if cell is None else str(cell).strip()
            if val:
                has_any = True
            row_dict[hname] = val
        if has_any:
            rows.append(row_dict)
    return headers, rows


def render_picker_block(title, search_name, search_value, selected_name, selected_id, selected_label, candidates, base_path, lang, query_keep=None, query_flag_name="do_search", query_enabled=True):
    query_keep = query_keep or {}
    hidden = "".join([f"<input type='hidden' name='{k}' value='{v}'>" for k, v in query_keep.items() if v not in (None, "")])
    cand_rows = ""
    is_class_picker = any('course_name' in c.keys() for c in candidates) if candidates else False
    recent_ids = [x for x in str(query_keep.get("recent_class_ids", "")).split(",") if x]
    ordered = []
    if is_class_picker and recent_ids:
        by_id = {str(r['id']): r for r in candidates}
        for rid in recent_ids:
            if rid in by_id:
                ordered.append(by_id[rid])
    for c in candidates:
        if c not in ordered:
            ordered.append(c)

    for c in ordered:
        cid = c['id']
        label = c.get('label', '') if isinstance(c, dict) else ''
        if not label:
            if 'student_no' in c.keys():
                label = f"{c['name_ko']} ({c['student_no'] or '-'}, {c['phone'] or '-'})"
            elif 'course_name' in c.keys():
                student_short, _ = summarize_student_names(c['student_names'] if 'student_names' in c.keys() else '')
                label = f"{c['name']} / {c['course_name'] or '-'} / {c['level_name'] or '-'} / {student_short}"
            else:
                label = f"{c['name']} ({c['username']})"
        keep = dict(query_keep)
        if is_class_picker and selected_name in ("selected_form_class_id", "selected_class_id"):
            new_recent = [str(cid)] + [x for x in recent_ids if x != str(cid)]
            keep["recent_class_ids"] = ",".join(new_recent[:5])
        qp = "&".join([f"{k}={v}" for k, v in keep.items() if v not in (None, "")])
        sep = "&" if qp else ""
        cand_rows += f"<li><a class='picker-link' data-preserve-scroll='1' href='{base_path}?lang={lang}{sep}{qp}&{selected_name}={cid}'>{label}</a></li>"

    class_table = ""
    if is_class_picker:
        rows = ""
        for c in ordered:
            cid = c['id']
            keep = dict(query_keep)
            if selected_name in ("selected_form_class_id", "selected_class_id"):
                new_recent = [str(cid)] + [x for x in recent_ids if x != str(cid)]
                keep["recent_class_ids"] = ",".join(new_recent[:5])
            qp = "&".join([f"{k}={v}" for k, v in keep.items() if v not in (None, "")])
            sep = "&" if qp else ""
            student_short, student_full = summarize_student_names(c['student_names'] if 'student_names' in c.keys() else '')
            rows += f"<tr><td class='col-class'><a class='picker-link' data-preserve-scroll='1' href='{base_path}?lang={lang}{sep}{qp}&{selected_name}={cid}'>{op_code('C', c['id'])} ? {c['name']}</a></td><td class='col-course'>{c['course_name'] or '-'}</td><td class='col-level'>{c['level_name'] or '-'}</td><td class='col-teacher'>{c['foreign_teacher_name'] or '-'}</td><td class='col-teacher'>{c['chinese_teacher_name'] or '-'}</td><td class='col-students' title='{h(student_full)}'>{h(student_short)}</td><td>{c['student_count'] or 0}</td></tr>"
        class_table = f"""
        <div class='table-wrap'><table>
          <tr><th>{t('academics.class_name')}</th><th>{t('academics.course')}</th><th>{t('academics.level')}</th><th>{t('academics.foreign_teacher')}</th><th>{t('academics.chinese_teacher')}</th><th>{t('academics.students')}</th><th>{t('academics.student_count')}</th></tr>
          {rows or f"<tr><td colspan='7' class='empty-msg'>{t('common.no_data')}</td></tr>"}
        </table></div>
        """

    list_inner = cand_rows if cand_rows else ("<li class='empty-msg'>" + t('common.no_data') + "</li>")
    list_html = "<ul style='margin:0; padding-left:18px'>" + list_inner + "</ul>"

    query_flag_html = f"<input type='hidden' name='{query_flag_name}' value='1'>" if query_flag_name else ""
    helper = "" if query_enabled else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")
    return f"""
    <div class='card'>
      <h4>{title}</h4>
      <form method='get' class='mobile-stack query-form picker-form'>
        <input type='hidden' name='lang' value='{lang}'>
        {hidden}
        {query_flag_html}
        <input name='{search_name}' value='{search_value or ''}' placeholder='search'>
        <div class='btn-row'>
          <button>{t('common.search')}</button>
          <a class='btn secondary admin-action-link' data-preserve-scroll='1' href='{base_path}?lang={lang}'>{t('common.reset')}</a>
        </div>
      </form>
      <div style='margin:6px 0'>{t('common.selected')}: <strong>{selected_label or '-'}</strong> (ID: {selected_id or '-'})</div>
      {helper}
      {class_table if query_enabled else ''}
      {list_html if (query_enabled and not class_table) else ''}
    </div>
    """

def app(environ, start_response):
    global CURRENT_LANG
    query = parse_query(environ)
    CURRENT_LANG = get_lang(environ)
    _orig_start_response = start_response
    def start_response(status, headers, exc_info=None):
        if query.get("lang", "").strip().lower() in SUPPORTED_LANGS:
            headers.append(("Set-Cookie", f"lang={CURRENT_LANG}; Path=/; Max-Age=31536000"))
        return _orig_start_response(status, headers, exc_info)
    path = environ.get("PATH_INFO", "/")
    method = environ.get("REQUEST_METHOD", "GET")
    # auth/api routes extracted in Phase A modules
    api_result = handle_api_routes(path, method, environ, None, None, {
        "parse_body": parse_body,
        "get_db": get_db,
        "hash_pw": hash_pw,
        "verify_pw": verify_pw,
        "needs_password_rehash": needs_password_rehash,
        "create_session": create_session,
        "build_session_cookie": build_session_cookie,
        "clear_session_cookie": clear_session_cookie,
        "uuid": uuid,
        "now": now,
        "json_resp": json_resp,
        "t": t,
    })
    if api_result is not None:
        status, headers, body = api_result
        start_response(status, headers)
        return [body]

    auth_result = handle_auth_routes(path, method, environ, {
        "t": t,
        "render_html": render_html,
        "text_resp": text_resp,
        "parse_body": parse_body,
        "parse_cookie": parse_cookie,
        "get_db": get_db,
        "hash_pw": hash_pw,
        "verify_pw": verify_pw,
        "needs_password_rehash": needs_password_rehash,
        "create_session": create_session,
        "build_session_cookie": build_session_cookie,
        "clear_session_cookie": clear_session_cookie,
        "invalidate_session": invalidate_session,
        "uuid": uuid,
        "now": now,
        "redirect": redirect,
    })
    if auth_result is not None:
        status, headers, body = auth_result
        start_response(status, headers)
        return [body]
    if path == "/":
        status, headers, body = redirect('/dashboard')
        start_response(status, headers)
        return [body]
    user, resp = require_login(environ)
    if resp:
        status, headers, body = resp
        start_response(status, headers)
        return [body]
    # Dashboard
    if path == "/dashboard":
        stats = {}
        conn_dash = get_db()
        stats["users"] = conn_dash.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
        stats["students"] = conn_dash.execute("SELECT COUNT(*) c FROM students").fetchone()["c"]
        stats["classes"] = conn_dash.execute("SELECT COUNT(*) c FROM classes").fetchone()["c"]
        stats["attendance"] = conn_dash.execute("SELECT COUNT(*) c FROM attendance").fetchone()["c"]
        conn_dash.close()
        body_html = f"""
        <div class='card'>
          <h3>{t("dash.stats")}</h3>
          <div class='filter-row'>
            <div class='card' style='min-width:160px'><strong>{t('dash.users')}</strong><div>{stats['users']}</div></div>
            <div class='card' style='min-width:160px'><strong>{t('dash.students')}</strong><div>{stats['students']}</div></div>
            <div class='card' style='min-width:160px'><strong>{t('dash.classes')}</strong><div>{stats['classes']}</div></div>
            <div class='card' style='min-width:160px'><strong>{t('dash.attendance')}</strong><div>{stats['attendance']}</div></div>
          </div>
        </div>
        <div class='card'>
          <h3>{t("dash.quick")}</h3>
          <div class='filter-row'>
            <a class='btn' href='/students?lang={CURRENT_LANG}'>{menu_t('students')}</a>
            <a class='btn' href='/masterdata?lang={CURRENT_LANG}'>{menu_t('masterdata')}</a>
            <a class='btn' href='/schedule?lang={CURRENT_LANG}'>{menu_t('schedule')}</a>
            <a class='btn' href='/attendance?lang={CURRENT_LANG}'>{menu_t('attendance')}</a>
            <a class='btn' href='/homework?lang={CURRENT_LANG}'>{menu_t('homework')}</a>
          </div>
        </div>
        """
        html = render_html(t("dash.title"), body_html, user, current_menu="dashboard")
        status, headers, body = text_resp(html)
        start_response(status, headers)
        return [body]
    conn = get_db()

    notifications_result = handle_notifications_routes(path, method, environ, user, conn, {
        "query": query,
        "has_role": has_role,
        "ROLE_OWNER": ROLE_OWNER,
        "ROLE_MANAGER": ROLE_MANAGER,
        "ROLE_TEACHER": ROLE_TEACHER,
        "parse_body": parse_body,
        "now": now,
        "render_html": render_html,
        "t": t,
        "text_resp": text_resp,
    })
    if notifications_result is not None:
        status, headers, body = notifications_result
        conn.close()
        start_response(status, headers)
        return [body]

    logs_result = handle_logs_routes(path, user, conn, {
        "query": query,
        "has_role": has_role,
        "ROLE_OWNER": ROLE_OWNER,
        "forbidden_html": forbidden_html,
        "ensure_logs_table": ensure_logs_table,
        "ensure_logs_columns": ensure_logs_columns,
        "render_html": render_html,
        "menu_t": menu_t,
        "t": t,
        "text_resp": text_resp,
    })
    if logs_result is not None:
        status, headers, body = logs_result
        conn.close()
        start_response(status, headers)
        return [body]

    api_result = handle_api_routes(path, method, environ, user, conn, {
        "parse_body": parse_body,
        "get_db": get_db,
        "hash_pw": hash_pw,
        "uuid": uuid,
        "now": now,
        "json_resp": json_resp,
        "t": t,
    })
    if api_result is not None:
        status, headers, body = api_result
        conn.close()
        start_response(status, headers)
        return [body]
    # 사용자 관리
    if path == "/users":
        if not has_role(user, [ROLE_OWNER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        flash_msg = ""
        flash_type = "success"
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            data = parse_body(environ)
            errs = []
            name = (data.get("name") or "").strip()
            username = (data.get("username") or "").strip()
            role = (data.get("role") or "").strip()
            password = data.get("password", "1234")
            teacher_type = (data.get("teacher_type") or "foreign").strip() or "foreign"
            if not name:
                add_error(errs, "name", "필수값입니다")
            if not username:
                add_error(errs, "username", "필수값입니다")
            if role not in (ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT):
                add_error(errs, "role", "허용되지 않는 역할입니다")
            if role == ROLE_TEACHER and teacher_type not in ("foreign", "chinese"):
                add_error(errs, "teacher_type", "foreign/chinese 중 하나여야 합니다")
            if conn.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
                add_error(errs, "username", "이미 사용 중입니다")

            if errs:
                flash_msg = format_errors(errs)
                flash_type = "error"
            else:
                try:
                    conn.execute("BEGIN")
                    conn.execute(
                        "INSERT INTO users(name, username, password_hash, role, teacher_type, created_at) VALUES(?,?,?,?,?,?)",
                        (name, username, hash_pw(password), role, teacher_type if role == ROLE_TEACHER else None, now()),
                    )
                    user_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                    if role == ROLE_STUDENT:
                        conn.execute(
                            """INSERT INTO students(user_id, student_no, name_ko, status, created_at, updated_at)
                            VALUES(?,?,?,?,?,?)""",
                            (user_id, f"S{user_id:05d}", name, "active", now(), now()),
                        )
                    elif role == ROLE_TEACHER:
                        conn.execute(
                            "INSERT INTO teachers(user_id, teacher_type, created_at, updated_at) VALUES(?,?,?,?)",
                            (user_id, teacher_type, now(), now()),
                        )
                    conn.commit()
                    flash_msg = t("users.saved")
                except Exception:
                    conn.rollback()
                    flash_msg = "사용자 저장 실패: 입력값/중복값을 확인하세요"
                    flash_type = "error"
                    log_event(conn, "ERROR", path, "사용자 저장 실패", traceback.format_exc(), user["id"])
                    conn.commit()

        load_users = query.get("load", "") == "1"
        q_user_name = (query.get("q_name", "") or "").strip()
        q_role = (query.get("q_role", "") or "").strip()
        users = []
        if load_users:
            where = []
            params = []
            if q_user_name:
                where.append("(name LIKE ? OR username LIKE ?)")
                params.extend([f"%{q_user_name}%", f"%{q_user_name}%"])
            if q_role in (ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT):
                where.append("role=?")
                params.append(q_role)
            where_sql = (" WHERE " + " AND ".join(where)) if where else ""
            users = conn.execute(f"SELECT * FROM users{where_sql} ORDER BY id DESC", tuple(params)).fetchall()
        rows = "".join([
            f"<tr><td>{u['id']}</td><td>{u['name']}</td><td>{u['username']}</td><td><span class='badge'>{ROLE_LABELS.get(u['role'],u['role'])}</span></td><td>{(u['teacher_type'] or '-')}</td></tr>"
            for u in users
        ])
        form = ""
        if has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            form = f"""
            <div class='card'>
              <h3>{t("users.add")}</h3>
              <form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'>
                <label>{t('field.name')} <input name='name' required></label>
                <label>{t('login.username')} <input name='username' required></label>
                <label>{t('login.password')} <input name='password' type='password'></label>
                <label>{t('field.role')}
                <select name='role'>
                  <option value='owner'>{role_label('owner')}</option><option value='manager'>{role_label('manager')}</option><option value='teacher'>{role_label('teacher')}</option>
                  <option value='parent'>{role_label('parent')}</option><option value='student'>{role_label('student')}</option>
                </select></label>
                <label id='teacher-type-wrap'>{t('users.teacher_type')}
                  <select name='teacher_type'>
                    <option value='foreign'>{t('users.teacher_type.foreign')}</option>
                    <option value='chinese'>{t('users.teacher_type.chinese')}</option>
                  </select>
                </label>
                <button>{t("common.save")}</button>
              </form>
              <script>
                (function(){{
                  var roleSel=document.querySelector("select[name='role']");
                  var wrap=document.getElementById('teacher-type-wrap');
                  function sync(){{ if(!roleSel||!wrap) return; wrap.style.display=(roleSel.value==='teacher'?'':'none'); }}
                  if(roleSel){{ roleSel.addEventListener('change', sync); sync(); }}
                }})();
              </script>
            </div>
            """
        body_html = form + f"""
        <div class='card'>
          <h3>{t("users.list")}</h3>
          <form method='get' class='mobile-stack query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='load' value='1'>
            <div class='filter-grid'>
              <label>{t('field.name')} <input name='q_name' value='{h(q_user_name)}'></label>
              <label>{t('field.role')}
                <select name='q_role'>
                  <option value=''>{t('academics.day_all')}</option>
                  <option value='owner' {'selected' if q_role=='owner' else ''}>{role_label('owner')}</option>
                  <option value='manager' {'selected' if q_role=='manager' else ''}>{role_label('manager')}</option>
                  <option value='teacher' {'selected' if q_role=='teacher' else ''}>{role_label('teacher')}</option>
                  <option value='parent' {'selected' if q_role=='parent' else ''}>{role_label('parent')}</option>
                  <option value='student' {'selected' if q_role=='student' else ''}>{role_label('student')}</option>
                </select>
              </label>
            </div>
            <div class='btn-row'>
              <button>{t('common.search')}</button>
              <a class='btn secondary' href='/users?lang={CURRENT_LANG}'>{t('common.reset')}</a>
            </div>
          </form>
          {'' if load_users else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}
          <div class='table-wrap'><table>
            <tr><th>{t('field.id')}</th><th>{t('field.name')}</th><th>{t('login.username')}</th><th>{t('field.role')}</th><th>{t('users.teacher_type')}</th></tr>
            {rows if load_users else ''}
            {("<tr><td colspan='5' class='empty-msg'>" + t('common.no_data') + "</td></tr>") if (load_users and not rows) else ''}
          </table></div>
        </div>
        """
        html = render_html(t("users.page_title"), body_html, user, current_menu="users", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path.startswith("/students/"):
        student_id = path.split("/")[-1]
        if not student_id.isdigit():
            conn.close()
            start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
            return ["Not Found".encode("utf-8")]
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            d = parse_body(environ)
            action = d.get("action")
            if action == "save":
                errs = []
                credits = as_float(d.get("remaining_credits") or 0)
                if not (d.get("name_ko") or "").strip():
                    add_error(errs, "name_ko", "필수값입니다")
                if credits is None:
                    add_error(errs, "remaining_credits", "숫자 형식이어야 합니다")
                if (d.get("status") or "active") not in ("active", "leave", "ended"):
                    add_error(errs, "status", "허용되지 않는 상태값입니다")
                class_id_val = d.get("current_class_id") or None
                if class_id_val and not ensure_exists(conn, "classes", class_id_val):
                    add_error(errs, "current_class_id", "존재하지 않는 반입니다")
                homeroom_teacher_id = (d.get("homeroom_teacher_id") or "").strip() or None
                if homeroom_teacher_id and not ensure_exists(conn, "users", homeroom_teacher_id, extra_where="role='teacher'"):
                    add_error(errs, "homeroom_teacher_id", "invalid homeroom teacher")
                for date_field in ["enrolled_at", "leave_start_date", "leave_end_date"]:
                    dv = (d.get(date_field) or "").strip()
                    if dv and not is_valid_date(dv):
                        add_error(errs, date_field, "날짜 형식은 YYYY-MM-DD 여야 합니다")
                if errs:
                    status, headers, body = redirect(f"/students/{student_id}?lang={CURRENT_LANG}&msg=validation_error")
                else:
                    conn.execute(
                        """UPDATE students SET
                        student_no=?, name_ko=?, name_en=?, phone=?, guardian_name=?, guardian_phone=?,
                        current_class_id=?, homeroom_teacher_id=?, remaining_credits=?, status=?, enrolled_at=?, leave_start_date=?, leave_end_date=?, memo=?, updated_at=?
                        WHERE id=?""",
                        (
                            d.get("student_no"), d.get("name_ko"), d.get("name_en"), d.get("phone"), d.get("guardian_name"), d.get("guardian_phone"),
                            class_id_val, homeroom_teacher_id, credits, d.get("status") or "active", d.get("enrolled_at"),
                            d.get("leave_start_date"), d.get("leave_end_date"), d.get("memo"), now(), student_id,
                        ),
                    )
                    conn.commit()
                    status, headers, body = redirect(f"/students/{student_id}?lang={CURRENT_LANG}&msg=saved")
                conn.close()
                start_response(status, headers)
                return [body]
            if action == "password":
                new_password = d.get("new_password", "").strip()
                student_row = conn.execute("SELECT user_id FROM students WHERE id=?", (student_id,)).fetchone()
                if not student_row or not student_row["user_id"]:
                    status, headers, body = redirect(f"/students/{student_id}?lang={CURRENT_LANG}&msg=no_user")
                    conn.close()
                    start_response(status, headers)
                    return [body]
                if not new_password:
                    status, headers, body = redirect(f"/students/{student_id}?lang={CURRENT_LANG}&msg=empty_pw")
                    conn.close()
                    start_response(status, headers)
                    return [body]
                conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_pw(new_password), student_row["user_id"]))
                conn.commit()
                status, headers, body = redirect(f"/students/{student_id}?lang={CURRENT_LANG}&msg=pw_saved")
                conn.close()
                start_response(status, headers)
                return [body]
        student = conn.execute(
            """SELECT s.*, u.username, c.name AS class_name, c.teacher_id AS current_class_teacher_id,
            ht.name AS homeroom_teacher_name
            FROM students s
            LEFT JOIN users u ON u.id=s.user_id
            LEFT JOIN classes c ON c.id=s.current_class_id
            LEFT JOIN users ht ON ht.id=s.homeroom_teacher_id
            WHERE s.id=?""",
            (student_id,),
        ).fetchone()
        classes = conn.execute("SELECT id, name FROM classes ORDER BY id DESC").fetchall()
        if not student:
            conn.close()
            start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
            return ["Student Not Found".encode("utf-8")]
        if not can_view_student_row(user, student):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        msg_key = query.get("msg", "")
        msg_map = {
            "saved": t("students.msg.saved"),
            "validation_error": t("students.msg.validation_error"),
            "pw_saved": t("students.msg.pw_saved"),
            "no_user": t("students.msg.no_user"),
            "empty_pw": t("students.msg.empty_pw"),
        }
        message = msg_map.get(msg_key, "")
        detail_page_size = 10
        def parse_page_param(name):
            raw = query.get(name, "1")
            return int(raw) if raw.isdigit() and int(raw) > 0 else 1
        def fetch_paged(sql, params, page):
            offset = (page - 1) * detail_page_size
            rows = conn.execute(sql + " LIMIT ? OFFSET ?", (*params, detail_page_size + 1, offset)).fetchall()
            has_next = len(rows) > detail_page_size
            return rows[:detail_page_size], has_next
        student_user_id = student["user_id"]
        att_page = parse_page_param("att_page")
        attendance_rows, attendance_has_next = fetch_paged(
            """SELECT a.lesson_date, a.status, a.note, c.name AS class_name
            FROM attendance a LEFT JOIN classes c ON c.id=a.class_id
            WHERE a.student_id=? ORDER BY a.id DESC""",
            (student_user_id,),
            att_page,
        )
        eval_page = parse_page_param("eval_page")
        evaluation_rows, evaluation_has_next = fetch_paged(
            """SELECT a.lesson_date, c.name AS class_name, a.status,
            a.participation_score, a.fluency_score, a.vocabulary_score, a.reading_score, a.homework_score, a.attitude_score,
            COALESCE(a.teacher_memo, a.note) AS teacher_memo
            FROM attendance a LEFT JOIN classes c ON c.id=a.class_id
            WHERE a.student_id=? AND (
              a.participation_score IS NOT NULL OR a.fluency_score IS NOT NULL OR a.vocabulary_score IS NOT NULL OR
              a.reading_score IS NOT NULL OR a.homework_score IS NOT NULL OR a.attitude_score IS NOT NULL OR COALESCE(a.teacher_memo, a.note) IS NOT NULL
            )
            ORDER BY a.id DESC""",
            (student_user_id,),
            eval_page,
        )
        eval_avg = conn.execute(
            """SELECT AVG(participation_score) AS avg_participation,
            AVG(fluency_score) AS avg_fluency,
            AVG(vocabulary_score) AS avg_vocabulary,
            AVG(reading_score) AS avg_reading,
            AVG(homework_score) AS avg_homework,
            AVG(attitude_score) AS avg_attitude
            FROM (
              SELECT participation_score, fluency_score, vocabulary_score, reading_score, homework_score, attitude_score
              FROM attendance
              WHERE student_id=? AND (
                participation_score IS NOT NULL OR fluency_score IS NOT NULL OR vocabulary_score IS NOT NULL OR
                reading_score IS NOT NULL OR homework_score IS NOT NULL OR attitude_score IS NOT NULL
              )
              ORDER BY id DESC LIMIT 10
            ) recent_eval""",
            (student_user_id,),
        ).fetchone()
        hw_page = parse_page_param("hw_page")
        submission_rows, submission_has_next = fetch_paged(
            """SELECT hs.id, h.title, hs.submitted, hs.submitted_at, hs.feedback
            FROM homework_submissions hs LEFT JOIN homework h ON h.id=hs.homework_id
            WHERE hs.student_id=? ORDER BY hs.id DESC""",
            (student_user_id,),
            hw_page,
        )
        exam_page = parse_page_param("exam_page")
        exam_rows, exam_has_next = fetch_paged(
            """SELECT e.name AS exam_name, es.score, e.exam_date
            FROM exam_scores es LEFT JOIN exams e ON e.id=es.exam_id
            WHERE es.student_id=? ORDER BY es.id DESC""",
            (student_user_id,),
            exam_page,
        )
        counseling_page = parse_page_param("counseling_page")
        counseling_rows, counseling_has_next = fetch_paged(
            """SELECT c.created_at AS recorded_at, c.memo, c.is_special_note
            FROM counseling c WHERE c.student_id=? ORDER BY c.id DESC""",
            (student_user_id,),
            counseling_page,
        )
        payment_page = parse_page_param("payment_page")
        payment_rows, payment_has_next = fetch_paged(
            """SELECT paid_date, amount, package_hours, remaining_classes
            FROM payments WHERE student_id=? ORDER BY id DESC""",
            (student_user_id,),
            payment_page,
        )
        loan_page = parse_page_param("loan_page")
        loan_rows, loan_has_next = fetch_paged(
            """SELECT b.code, b.title, bl.loaned_at, bl.returned_at
            FROM book_loans bl LEFT JOIN books b ON b.id=bl.book_id
            WHERE bl.student_id=? ORDER BY bl.id DESC""",
            (student_user_id,),
            loan_page,
        )
        def rows_html(rows, cols, empty_colspan):
            if not rows:
                return f"<tr><td colspan='{empty_colspan}'>{t('common.no_data')}</td></tr>"
            out = ""
            for r in rows:
                values = []
                for c in cols:
                    v = r[c]
                    if c == "status" and v:
                        v = status_t(v)
                    elif c == "submitted":
                        v = t("common.yes") if str(v) == "1" else t("common.no")
                    elif c == "is_special_note":
                        v = t("common.yes") if str(v) == "1" else t("common.no")
                    if v in (None, ""):
                        v = "-"
                    values.append(f"<td>{v}</td>")
                out += "<tr>" + "".join(values) + "</tr>"
            return out
        def section_pager(param_name, page, has_next):
            page_keys = ["att_page", "eval_page", "hw_page", "exam_page", "counseling_page", "payment_page", "loan_page"]
            def mk_link(target_page):
                qp = {k: query.get(k, "1") for k in page_keys}
                qp[param_name] = str(target_page)
                qp["lang"] = CURRENT_LANG
                qstr = "&".join([f"{k}={v}" for k, v in qp.items()])
                return f"?{qstr}"
            prev_link = f"<a href='{mk_link(page-1)}'>{t('common.prev')}</a>" if page > 1 else ""
            next_link = f"<a href='{mk_link(page+1)}'>{t('common.next')}</a>" if has_next else ""
            if not prev_link and not next_link:
                return ""
            sep = " | " if prev_link and next_link else ""
            return f"<div style='margin:4px 0 12px 0'>{prev_link}{sep}{next_link}</div>"
        class_opts = ["<option value=''>-</option>"]
        for c in classes:
            selected = "selected" if student["current_class_id"] == c["id"] else ""
            class_opts.append(f"<option value='{c['id']}' {selected}>{c['name']}</option>")
        class_options = "".join(class_opts)
        teacher_opts = ["<option value=''>-</option>"]
        for tr in list_teacher_profiles(conn):
            selected = "selected" if str(student["homeroom_teacher_id"] or "") == str(tr["id"]) else ""
            teacher_opts.append(f"<option value='{tr['id']}' {selected}>{h(tr['name'])}</option>")
        homeroom_options = "".join(teacher_opts)
        edit_form = ""
        pw_form = ""
        if has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            edit_form = f"""
            <form id='student-edit' method='post' style='margin-bottom:16px'>
              <input type='hidden' name='action' value='save'>
              <h4>{t('common.edit')}</h4>
              {t('students.field.student_no')} <input name='student_no' value='{student['student_no'] or ''}'>
              {t('students.field.name_ko')} <input name='name_ko' value='{student['name_ko'] or ''}'>
              {t('students.field.name_en')} <input name='name_en' value='{student['name_en'] or ''}'><br>
              {t('students.field.phone')} <input name='phone' value='{student['phone'] or ''}'>
              {t('students.field.guardian_name')} <input name='guardian_name' value='{student['guardian_name'] or ''}'>
              {t('students.field.guardian_phone')} <input name='guardian_phone' value='{student['guardian_phone'] or ''}'><br>
              {t('students.field.class')} <select name='current_class_id'>{class_options}</select>
              {t('students.field.homeroom_teacher')} <select name='homeroom_teacher_id'>{homeroom_options}</select>
              {t('students.field.credits')} <input name='remaining_credits' value='{student['remaining_credits'] or 0}'>
              {t('students.field.status')} <select name='status'>
                <option value='active' {'selected' if student['status']=='active' else ''}>{status_t('active')}</option>
                <option value='leave' {'selected' if student['status']=='leave' else ''}>{status_t('leave')}</option>
                <option value='ended' {'selected' if student['status']=='ended' else ''}>{status_t('ended')}</option>
              </select><br>
              {t('students.field.enrolled_at')} <input name='enrolled_at' value='{student['enrolled_at'] or ''}'>
              {t('students.field.leave_start_date')} <input name='leave_start_date' value='{student['leave_start_date'] or ''}'>
              {t('students.field.leave_end_date')} <input name='leave_end_date' value='{student['leave_end_date'] or ''}'><br>
              {t('students.field.memo')} <input name='memo' style='width:500px' value='{student['memo'] or ''}'><br>
              <button>{t('common.save')}</button>
            </form>
            """
            pw_form = f"""
            <form method='post'>
              <input type='hidden' name='action' value='password'>
              <h4>{t('students.password_reset')}</h4>
              {t('students.new_password')} <input name='new_password' type='password'>
              <button>{t('common.edit')}</button>
            </form>
            """
        edit_button_html = ""
        if has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            edit_button_html = f'<div style="margin:8px 0"><button type="button" onclick="location.hash=\'student-edit\'">{t("students.detail.edit")}</button></div>'
        body_html = f"""
        <div><a href='/students?lang={CURRENT_LANG}'>← {t("students.detail.back")}</a></div>
        <h3>{student['name_ko']} ({student['student_no'] or '-'})</h3>{edit_button_html}
        <h4>{t('students.detail.section.basic')}</h4>
        <div class='card'>
          <div class='student-summary-grid'>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.student_no')}</div><div class='student-summary-value'>{student['student_no'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.name_ko')}</div><div class='student-summary-value'>{student['name_ko'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.name_en')}</div><div class='student-summary-value'>{student['name_en'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.class')}</div><div class='student-summary-value'>{student['class_name'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.homeroom_teacher')}</div><div class='student-summary-value'>{homeroom_display_name(student['homeroom_teacher_name'])}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.credits')}</div><div class='student-summary-value'>{student['remaining_credits'] or 0}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.status')}</div><div class='student-summary-value'>{status_t(student['status']) if student['status'] else '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.phone')}</div><div class='student-summary-value'>{student['phone'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.guardian_name')}</div><div class='student-summary-value'>{student['guardian_name'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.guardian_phone')}</div><div class='student-summary-value'>{student['guardian_phone'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.enrolled_at')}</div><div class='student-summary-value'>{student['enrolled_at'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.leave_period')}</div><div class='student-summary-value'>{student['leave_start_date'] or '-'} ~ {student['leave_end_date'] or '-'}</div></div>
            <div class='student-summary-item'><div class='student-summary-label'>{t('students.field.memo')}</div><div class='student-summary-value'>{student['memo'] or '-'}</div></div>
          </div>
        </div>
        <h4>{t('students.detail.section.attendance')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.lesson_date')}</th><th>{t('students.field.class')}</th><th>{t('students.field.status')}</th><th>{t('students.field.note')}</th></tr>
          {rows_html(attendance_rows, ['lesson_date', 'class_name', 'status', 'note'], 4)}
        </table></div>
        {section_pager('att_page', att_page, attendance_has_next)}
        <h4>{t('students.detail.section.evaluations')}</h4>
        <div class='muted'>
          {t('students.eval.avg.title')}: 
          {t('students.eval.avg.participation')} {f"{eval_avg['avg_participation']:.2f}" if eval_avg and eval_avg['avg_participation'] is not None else '-'} |
          {t('students.eval.avg.fluency')} {f"{eval_avg['avg_fluency']:.2f}" if eval_avg and eval_avg['avg_fluency'] is not None else '-'} |
          {t('students.eval.avg.vocabulary')} {f"{eval_avg['avg_vocabulary']:.2f}" if eval_avg and eval_avg['avg_vocabulary'] is not None else '-'} |
          {t('students.eval.avg.reading')} {f"{eval_avg['avg_reading']:.2f}" if eval_avg and eval_avg['avg_reading'] is not None else '-'} |
          {t('students.eval.avg.homework')} {f"{eval_avg['avg_homework']:.2f}" if eval_avg and eval_avg['avg_homework'] is not None else '-'} |
          {t('students.eval.avg.attitude')} {f"{eval_avg['avg_attitude']:.2f}" if eval_avg and eval_avg['avg_attitude'] is not None else '-'}
        </div>
        <table>
          <tr><th>{t('students.field.lesson_date')}</th><th>{t('students.field.class')}</th><th>{t('students.field.status')}</th><th>{t('lesson.score.participation')}</th><th>{t('lesson.score.fluency')}</th><th>{t('lesson.score.vocabulary')}</th><th>{t('lesson.score.reading')}</th><th>{t('lesson.score.homework')}</th><th>{t('lesson.score.attitude')}</th><th>{t('lesson.score.teacher_memo')}</th></tr>
          {rows_html(evaluation_rows, ['lesson_date', 'class_name', 'status', 'participation_score', 'fluency_score', 'vocabulary_score', 'reading_score', 'homework_score', 'attitude_score', 'teacher_memo'], 10)}
        </table></div>
        {section_pager('eval_page', eval_page, evaluation_has_next)}
        <h4>{t('students.detail.section.homework')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.homework')}</th><th>{t('students.field.submitted')}</th><th>{t('students.field.submitted_at')}</th><th>{t('students.field.feedback')}</th></tr>
          {rows_html(submission_rows, ['title', 'submitted', 'submitted_at', 'feedback'], 4)}
        </table></div>
        {section_pager('hw_page', hw_page, submission_has_next)}
        <h4>{t('students.detail.section.exams')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.exam_name')}</th><th>{t('students.field.score')}</th><th>{t('students.field.exam_date')}</th></tr>
          {rows_html(exam_rows, ['exam_name', 'score', 'exam_date'], 3)}
        </table></div>
        {section_pager('exam_page', exam_page, exam_has_next)}
        <h4>{t('students.detail.section.counseling')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.recorded_at')}</th><th>{t('students.field.memo')}</th><th>{t('students.field.special_note')}</th></tr>
          {rows_html(counseling_rows, ['recorded_at', 'memo', 'is_special_note'], 3)}
        </table></div>
        {section_pager('counseling_page', counseling_page, counseling_has_next)}
        <h4>{t('students.detail.section.payments')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.paid_date')}</th><th>{t('students.field.amount')}</th><th>{t('students.field.package_hours')}</th><th>{t('students.field.remaining_classes')}</th></tr>
          {rows_html(payment_rows, ['paid_date', 'amount', 'package_hours', 'remaining_classes'], 4)}
        </table></div>
        {section_pager('payment_page', payment_page, payment_has_next)}
        <h4>{t('students.detail.section.loans')}</h4>
        <div class='table-wrap'><table>
          <tr><th>{t('students.field.code')}</th><th>{t('students.field.title')}</th><th>{t('students.field.loaned_at')}</th><th>{t('students.field.returned_at')}</th></tr>
          {rows_html(loan_rows, ['code', 'title', 'loaned_at', 'returned_at'], 4)}
        </table></div>
        {section_pager('loan_page', loan_page, loan_has_next)}
        {edit_form}
        {pw_form}
        """
        html = render_html(t("students.detail.title"), body_html, user, current_menu="students", flash_msg=message)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/students":
        flash_msg = ""
        flash_type = "success"
        load_students = query.get("load", "") == "1"

        teacher_rows = list_teacher_profiles(conn)
        q_homeroom_teacher_id = (query.get("homeroom_teacher_id", "") or "").strip()
        homeroom_options = [f"<option value=''>{t('academics.day_all')}</option>"]
        for tr in teacher_rows:
            sel = "selected" if str(tr["id"]) == q_homeroom_teacher_id else ""
            homeroom_options.append(f"<option value='{tr['id']}' {sel}>{h(tr['name'])}</option>")

        class_rows_for_map = conn.execute("SELECT id, name FROM classes ORDER BY id DESC").fetchall()
        class_name_to_id = {((r['name'] or '').strip().lower()): r['id'] for r in class_rows_for_map}

        def build_student_filters(source):
            q_name = (source.get("name", "") or "").strip()
            q_student_no = (source.get("student_no", "") or "").strip()
            q_phone = (source.get("phone", "") or "").strip()
            q_status = (source.get("status", "") or "").strip()
            q_homeroom = (source.get("homeroom_teacher_id", "") or "").strip()
            where = []
            params = []
            if q_name:
                where.append("(s.name_ko LIKE ? OR s.name_en LIKE ?)")
                params += [f"%{q_name}%", f"%{q_name}%"]
            if q_student_no:
                where.append("s.student_no LIKE ?")
                params.append(f"%{q_student_no}%")
            if q_phone:
                where.append("s.phone LIKE ?")
                params.append(f"%{q_phone}%")
            if q_status in ("active", "leave", "ended"):
                where.append("COALESCE(s.status, 'active')=?")
                params.append(q_status)
            if q_homeroom.isdigit():
                where.append("COALESCE(s.homeroom_teacher_id,0)=?")
                params.append(q_homeroom)
            return q_name, q_student_no, q_phone, q_status, q_homeroom, where, params

        q_name, q_student_no, q_phone, q_status, q_homeroom_teacher_id, where, params = build_student_filters(query)

        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            ctype = environ.get("CONTENT_TYPE", "")
            files = {}
            if "multipart/form-data" in ctype:
                d, files = parse_multipart_form(environ)
            else:
                d = parse_body(environ)
            typ = (d.get("type") or "save").strip()

            if typ == "bulk_upload":
                upload_file = files.get("upload_file") if files else None
                required_headers = list(UPLOAD_TEMPLATE_HEADERS)
                if not upload_file or not upload_file.get("content"):
                    flash_msg = t('students.upload.file_required')
                    flash_type = "error"
                else:
                    try:
                        headers, upload_rows = read_upload_rows(upload_file.get("filename"), upload_file.get("content"))
                        normalized_headers = [h.strip() for h in headers if h and h.strip()]
                        missing_headers = sorted(set(required_headers) - set(normalized_headers))
                        if not normalized_headers:
                            flash_msg = t('students.upload.empty_file')
                            flash_type = "error"
                        elif missing_headers:
                            flash_msg = t('students.upload.invalid_headers').format(missing=(', '.join(missing_headers)))
                            flash_type = "error"
                        else:
                            total = len(upload_rows)
                            success = 0
                            failed = 0
                            row_errors = []
                            for idx, row in enumerate(upload_rows, start=2):
                                student_no = (row.get("student_no") or "").strip()
                                name_ko = (row.get("chinese_name") or "").strip()
                                name_en = (row.get("english_name") or "").strip()
                                phone = (row.get("phone") or "").strip()
                                guardian_name = (row.get("guardian_name") or "").strip()
                                guardian_phone = (row.get("guardian_phone") or "").strip()
                                class_name = (row.get("class_name") or "").strip()
                                homeroom_teacher_name = (row.get("homeroom_teacher_name") or "").strip()
                                status_v = (row.get("status") or "").strip().lower()
                                enrolled_at = (row.get("enrollment_date") or "").strip()
                                leave_period = (row.get("leave_period") or "").strip()
                                memo = (row.get("memo") or "").strip()

                                errs = []
                                if not student_no:
                                    errs.append(f"Row {idx}: student_no is missing")
                                if not (name_ko or name_en):
                                    errs.append(f"Row {idx}: chinese_name or english_name is required")
                                if not guardian_name:
                                    errs.append(f"Row {idx}: guardian_name is missing")
                                if not guardian_phone:
                                    errs.append(f"Row {idx}: guardian_phone is missing")
                                if not class_name:
                                    errs.append(f"Row {idx}: class_name is missing")
                                if not homeroom_teacher_name:
                                    errs.append(f"Row {idx}: homeroom_teacher_name is missing")
                                if status_v not in ("active", "leave", "ended"):
                                    errs.append(f"Row {idx}: status must be active/leave/ended")
                                if enrolled_at and not is_valid_date(enrolled_at):
                                    errs.append(f"Row {idx}: enrollment_date format must be YYYY-MM-DD")

                                class_id = class_name_to_id.get(class_name.lower())
                                if not class_id:
                                    errs.append(f"Row {idx}: class_name not found ({class_name})")
                                homeroom_teacher_id = find_teacher_id_by_name(conn, homeroom_teacher_name)
                                if homeroom_teacher_name and not homeroom_teacher_id:
                                    errs.append(f"Row {idx}: homeroom_teacher_name not found ({homeroom_teacher_name})")

                                leave_start, leave_end = parse_leave_period_range(leave_period)
                                if leave_period and (not leave_start or not leave_end):
                                    errs.append(f"Row {idx}: leave_period format should be YYYY-MM-DD~YYYY-MM-DD")

                                exists_student_no = conn.execute("SELECT id FROM students WHERE student_no=?", (student_no,)).fetchone() if student_no else None
                                if exists_student_no:
                                    errs.append(f"Row {idx}: student_no already exists ({student_no})")

                                if errs:
                                    failed += 1
                                    row_errors.extend(errs)
                                    continue

                                base_username = f"stu_{student_no.lower()}".replace(" ", "_")
                                username = base_username
                                seq = 1
                                while conn.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
                                    seq += 1
                                    username = f"{base_username}_{seq}"

                                conn.execute(
                                    "INSERT INTO users(name, username, password_hash, role, created_at) VALUES(?,?,?,?,?)",
                                    ((name_ko or name_en), username, hash_pw("1234"), ROLE_STUDENT, now()),
                                )
                                user_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                                conn.execute(
                                    """INSERT INTO students(
                                    user_id, student_no, name_ko, name_en, phone, guardian_name, guardian_phone,
                                    current_class_id, homeroom_teacher_id, remaining_credits, status, enrolled_at,
                                    leave_start_date, leave_end_date, memo, created_at, updated_at
                                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                    (
                                        user_id, student_no, (name_ko or name_en), name_en, phone, guardian_name, guardian_phone,
                                        class_id, homeroom_teacher_id, 0, status_v, enrolled_at or None,
                                        leave_start, leave_end, memo or None, now(), now(),
                                    ),
                                )
                                success += 1

                            conn.commit()
                            load_students = True
                            summary = [f"Total rows: {total}", f"Success: {success}", f"Failed: {failed}"]
                            if row_errors:
                                summary.append("Details:")
                                summary.extend(row_errors[:80])
                            flash_msg = "<br>".join(summary)
                            flash_type = "error" if failed else "success"
                    except Exception as e:
                        conn.rollback()
                        flash_msg = t('students.upload.failed').format(error=h(str(e)))
                        flash_type = "error"

            else:
                errs = []
                student_user_id = d.get("user_id")
                if not ensure_exists(conn, "users", student_user_id, extra_where="role='student'"):
                    add_error(errs, "user_id", "student role user only")
                if not (d.get("name_ko") or "").strip():
                    add_error(errs, "name_ko", "required")
                credits = as_float(d.get("remaining_credits") or 0)
                if credits is None:
                    add_error(errs, "remaining_credits", "number format")
                status_v = d.get("status") or "active"
                if status_v not in ("active", "leave", "ended"):
                    add_error(errs, "status", "invalid status")
                class_id_val = d.get("current_class_id") or None
                if class_id_val and not ensure_exists(conn, "classes", class_id_val):
                    add_error(errs, "current_class_id", "invalid class")
                homeroom_teacher_id = (d.get("homeroom_teacher_id") or "").strip() or None
                if homeroom_teacher_id and not ensure_exists(conn, "users", homeroom_teacher_id, extra_where="role='teacher'"):
                    add_error(errs, "homeroom_teacher_id", "invalid homeroom teacher")
                for date_field in ["enrolled_at", "leave_start_date", "leave_end_date"]:
                    dv = (d.get(date_field) or "").strip()
                    if dv and not is_valid_date(dv):
                        add_error(errs, date_field, "YYYY-MM-DD required")
                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                else:
                    exists = conn.execute("SELECT id FROM students WHERE user_id=?", (student_user_id,)).fetchone()
                    if exists:
                        conn.execute(
                            """UPDATE students SET
                            student_no=?, name_ko=?, name_en=?, phone=?, guardian_name=?, guardian_phone=?,
                            current_class_id=?, homeroom_teacher_id=?, remaining_credits=?, status=?, enrolled_at=?, leave_start_date=?, leave_end_date=?, memo=?, updated_at=?
                            WHERE user_id=?""",
                            (
                                (d.get("student_no") or f"S{int(student_user_id):05d}"), d.get("name_ko"), d.get("name_en"), d.get("phone"), d.get("guardian_name"), d.get("guardian_phone"),
                                class_id_val, homeroom_teacher_id, credits, status_v, d.get("enrolled_at"), d.get("leave_start_date"), d.get("leave_end_date"), d.get("memo"), now(), student_user_id,
                            ),
                        )
                    else:
                        conn.execute(
                            """INSERT INTO students(
                            user_id, student_no, name_ko, name_en, phone, guardian_name, guardian_phone,
                            current_class_id, homeroom_teacher_id, remaining_credits, status, enrolled_at, leave_start_date, leave_end_date, memo, created_at, updated_at
                            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (
                                student_user_id, d.get("student_no"), d.get("name_ko"), d.get("name_en"), d.get("phone"), d.get("guardian_name"), d.get("guardian_phone"),
                                class_id_val, homeroom_teacher_id, credits, status_v, d.get("enrolled_at"), d.get("leave_start_date"), d.get("leave_end_date"), d.get("memo"), now(), now(),
                            ),
                        )
                    conn.commit()
                    flash_msg = "saved"
                    load_students = True

        if query.get("template") == "1":
            xlsx_bytes = build_upload_template_xlsx()
            file_date = datetime.utcnow().date().isoformat()
            if xlsx_bytes:
                conn.close()
                start_response("200 OK", [("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("Content-Disposition", f"attachment; filename=students_upload_template_{file_date}.xlsx")])
                return [xlsx_bytes]
            csv_bytes = build_upload_template_csv()
            conn.close()
            start_response("200 OK", [("Content-Type", "text/csv; charset=utf-8"), ("Content-Disposition", f"attachment; filename=students_upload_template_{file_date}.csv")])
            return [csv_bytes]

        students = []
        if load_students:
            where_sql = ("WHERE " + " AND ".join(where)) if where else ""
            students_all = conn.execute(
                f"""SELECT s.*, c.name AS class_name, c.teacher_id AS current_class_teacher_id,
                ht.name AS homeroom_teacher_name
                FROM students s
                LEFT JOIN classes c ON c.id=s.current_class_id
                LEFT JOIN users ht ON ht.id=s.homeroom_teacher_id
                {where_sql}
                ORDER BY s.id DESC""",
                params,
            ).fetchall()
            students = [st for st in students_all if can_view_student_row(user, st)]

        if query.get("export") in ("xlsx", "csv"):
            if not load_students:
                flash_msg = t('students.export.query_first')
                flash_type = "error"
            else:
                file_date = datetime.utcnow().date().isoformat()
                as_dict = [dict(r) for r in students]
                if query.get("export") == "xlsx":
                    xlsx_bytes = build_students_xlsx(as_dict)
                    if xlsx_bytes:
                        conn.close()
                        start_response("200 OK", [("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("Content-Disposition", f"attachment; filename=students_{file_date}.xlsx")])
                        return [xlsx_bytes]
                csv_bytes = build_students_csv(as_dict)
                conn.close()
                start_response("200 OK", [("Content-Type", "text/csv; charset=utf-8"), ("Content-Disposition", f"attachment; filename=students_{file_date}.csv")])
                return [csv_bytes]

        rows = ""
        for st in students:
            rows += f"""
            <tr>
              <td>{h(st['student_no'] or '-')}</td>
              <td><a href='/students/{st['id']}?lang={CURRENT_LANG}'>{h(st['name_ko'] or '-')}</a></td>
              <td>{h(st['name_en'] or '-')}</td>
              <td>{h(st['phone'] or '-')}</td>
              <td>{h(st['guardian_name'] or '-')}</td>
              <td>{h(st['guardian_phone'] or '-')}</td>
              <td>{h(st['class_name'] or '-')}</td>
              <td>{h(homeroom_display_name(st['homeroom_teacher_name']))}</td>
              <td>{h(st['remaining_credits'] or 0)}</td>
              <td><span class='badge {h(st['status'] or '')}'>{h(status_t(st['status']) if st['status'] else '-')}</span></td>
            </tr>
            """

        html = render_html(t('menu.students'), f"""
        <div class='card'>
          <h3>{t("students.search")}</h3>
          <form method='get' class='mobile-stack query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='load' value='1'>
            <div class='filter-grid'>
              <label>{t('field.name')} <input name='name' value='{h(q_name)}'></label>
              <label>{t('students.field.student_no')} <input name='student_no' value='{h(q_student_no)}'></label>
              <label>{t('students.field.phone')} <input name='phone' value='{h(q_phone)}'></label>
              <label>{t('students.field.status')} <select name='status'>
                <option value=''>{t('academics.day_all')}</option>
                <option value='active' {'selected' if q_status=='active' else ''}>{status_t('active')}</option>
                <option value='leave' {'selected' if q_status=='leave' else ''}>{status_t('leave')}</option>
                <option value='ended' {'selected' if q_status=='ended' else ''}>{status_t('ended')}</option>
              </select></label>
              <label>{t('students.field.homeroom_teacher')} <select name='homeroom_teacher_id'>{''.join(homeroom_options)}</select></label>
            </div>
            <div class='btn-row'>
              <button>{t("common.search")}</button>
              <a class='btn secondary' href='/students?lang={CURRENT_LANG}'>{t('common.reset')}</a>
              {f"<a class='btn secondary' href='/students?lang={CURRENT_LANG}&load=1&name={h(q_name)}&student_no={h(q_student_no)}&phone={h(q_phone)}&status={h(q_status)}&homeroom_teacher_id={h(q_homeroom_teacher_id)}&export=xlsx'>{t('students.export_excel')}</a>" if load_students else f"<span class='muted'>{t('students.export.query_first')}</span>"}
            </div>
          </form>
          <div class='btn-row' style='margin-top:10px'>
            <a class='btn secondary' href='/students?lang={CURRENT_LANG}&template=1'>{t('students.upload_template')}</a>
          </div>
          <form method='post' enctype='multipart/form-data' class='form-row' style='margin-top:10px'>
            <input type='hidden' name='type' value='bulk_upload'>
            <label>{t('students.upload_students')} <input type='file' name='upload_file' accept='.xlsx,.csv'></label>
            <button>{t('students.upload_students')}</button>
          </form>
        </div>
        <div class='card'>
          <h3>{t("students.list")}</h3>
          {'' if load_students else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}
          <div class='table-wrap'><table>
            <tr>
              <th>{t('students.field.student_no')}</th><th>{t('students.field.name_ko')}</th><th>{t('students.field.name_en')}</th><th>{t('students.field.phone')}</th><th>{t('students.field.guardian_name')}</th><th>{t('students.field.guardian_phone')}</th><th>{t('students.field.class')}</th><th>{t('students.field.homeroom_teacher')}</th><th>{t('students.field.credits')}</th><th>{t('students.field.status')}</th>
            </tr>
            {rows if load_students else ''}
            {("<tr><td colspan='10' class='empty-msg'>" + t('common.no_data') + "</td></tr>") if (load_students and not rows) else ""}
          </table></div>
        </div>
        """, user, current_menu="students", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]

    if path.startswith("/classes/"):
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        class_id = path.split("/")[-1]
        if not class_id.isdigit():
            conn.close()
            start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
            return ["Not Found".encode("utf-8")]
        if has_role(user, [ROLE_TEACHER]):
            class_row = conn.execute(
                """SELECT c.*, co.name AS course_name, l.name AS level_name,
                uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name
                FROM classes c
                LEFT JOIN courses co ON co.id=c.course_id
                LEFT JOIN levels l ON l.id=c.level_id
                LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                WHERE c.id=? AND COALESCE(c.foreign_teacher_id, c.teacher_id)=?""",
                (class_id, user["id"]),
            ).fetchone()
        else:
            class_row = conn.execute(
                """SELECT c.*, co.name AS course_name, l.name AS level_name,
                uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name
                FROM classes c
                LEFT JOIN courses co ON co.id=c.course_id
                LEFT JOIN levels l ON l.id=c.level_id
                LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                WHERE c.id=?""",
                (class_id,),
            ).fetchone()
        if not class_row:
            conn.close()
            start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
            return ["Class Not Found".encode("utf-8")]

        flash_msg = ""
        flash_type = "success"
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            data = parse_body(environ)
            if data.get("type") == "class_edit":
                errs = []
                class_name = (data.get("name") or "").strip()
                course_id = (data.get("course_id") or "").strip()
                level_id = (data.get("level_id") or "").strip()
                foreign_teacher_id = (data.get("foreign_teacher_id") or "").strip()
                chinese_teacher_id = (data.get("chinese_teacher_id") or "").strip()
                class_status = (data.get("status") or "active").strip() or "active"
                class_memo = (data.get("memo") or "").strip()
                if not class_name:
                    add_error(errs, "class.name", "Required")
                if not ensure_exists(conn, "courses", course_id):
                    add_error(errs, "class.course_id", "Invalid course")
                if level_id and not ensure_exists(conn, "levels", level_id):
                    add_error(errs, "class.level_id", "Invalid level")
                if foreign_teacher_id and not ensure_exists(conn, "teachers", foreign_teacher_id, field="user_id"):
                    add_error(errs, "class.foreign_teacher_id", "Invalid foreign teacher")
                if chinese_teacher_id and not ensure_exists(conn, "teachers", chinese_teacher_id, field="user_id"):
                    add_error(errs, "class.chinese_teacher_id", "Invalid chinese teacher")
                if class_status not in ("active", "inactive"):
                    add_error(errs, "class.status", "Invalid status")
                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                else:
                    conn.execute(
                        """UPDATE classes
                        SET name=?, course_id=?, level_id=?, teacher_id=?, foreign_teacher_id=?, chinese_teacher_id=?, status=?, memo=?
                        WHERE id=?""",
                        (
                            class_name,
                            course_id,
                            level_id or None,
                            foreign_teacher_id or None,
                            foreign_teacher_id or None,
                            chinese_teacher_id or None,
                            class_status,
                            class_memo or None,
                            class_id,
                        ),
                    )
                    conn.commit()
                    flash_msg = t("common.save")
                class_row = conn.execute(
                    """SELECT c.*, co.name AS course_name, l.name AS level_name,
                    uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name
                    FROM classes c
                    LEFT JOIN courses co ON co.id=c.course_id
                    LEFT JOIN levels l ON l.id=c.level_id
                    LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                    LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                    WHERE c.id=?""",
                    (class_id,),
                ).fetchone()

        student_sort = query.get("student_sort", "name_ko")
        if student_sort not in ("name_ko", "student_no", "status"):
            student_sort = "name_ko"
        student_order = query.get("student_order", "asc").lower()
        if student_order not in ("asc", "desc"):
            student_order = "asc"
        students = conn.execute(
            f"""SELECT id, student_no, name_ko, phone, status
            FROM students WHERE current_class_id=?
            ORDER BY {student_sort} {student_order}, id DESC""",
            (class_id,),
        ).fetchall()
        schedules = conn.execute(
            """SELECT id, day_of_week, start_time, end_time, COALESCE(classroom,'-') AS classroom, COALESCE(status,'active') AS status, COALESCE(note,'') AS note
            FROM schedules WHERE class_id=? ORDER BY id DESC""",
            (class_id,),
        ).fetchall()
        attendance_rows = conn.execute(
            """SELECT a.lesson_date, s.name_ko AS student_name, a.status, a.note
            FROM attendance a LEFT JOIN students s ON s.user_id=a.student_id
            WHERE a.class_id=? ORDER BY a.id DESC LIMIT 20""",
            (class_id,),
        ).fetchall()
        homework_rows = conn.execute(
            """SELECT h.title, h.due_date,
            COUNT(hs.id) AS total_submissions,
            SUM(CASE WHEN hs.submitted=1 THEN 1 ELSE 0 END) AS submitted_count
            FROM homework h
            LEFT JOIN homework_submissions hs ON hs.homework_id=h.id
            WHERE h.class_id=?
            GROUP BY h.id
            ORDER BY h.id DESC LIMIT 20""",
            (class_id,),
        ).fetchall()
        exam_rows = conn.execute(
            """SELECT e.name AS exam_name, e.exam_date, ROUND(AVG(es.score),2) AS avg_score, COUNT(es.id) AS score_count
            FROM exams e
            LEFT JOIN exam_scores es ON es.exam_id=e.id
            WHERE e.class_id=?
            GROUP BY e.id
            ORDER BY e.id DESC LIMIT 20""",
            (class_id,),
        ).fetchall()
        export = query.get("export", "")
        if export in ("students_csv", "attendance_csv"):
            if export == "students_csv":
                lines = ["student_no,name_ko,phone,status"]
                for r in students:
                    lines.append(f'"{r["student_no"] or ""}","{r["name_ko"] or ""}","{r["phone"] or ""}","{r["status"] or ""}"')
                filename = f"class_{class_id}_students.csv"
            else:
                lines = ["lesson_date,student_name,status,note"]
                for r in attendance_rows:
                    lines.append(f'"{r["lesson_date"] or ""}","{r["student_name"] or ""}","{r["status"] or ""}","{r["note"] or ""}"')
                filename = f"class_{class_id}_attendance.csv"
            conn.close()
            start_response("200 OK", [
                ("Content-Type", "text/csv; charset=utf-8"),
                ("Content-Disposition", f"attachment; filename={filename}"),
            ])
            return ["\n".join(lines).encode("utf-8")]
        def rows_html(rows, cols):
            if not rows:
                return f"<tr><td colspan='{len(cols)}'>{t('common.no_data')}</td></tr>"
            out = ""
            for r in rows:
                out += "<tr>" + "".join([f"<td>{r[c] if r[c] not in (None, '') else '-'}</td>" for c in cols]) + "</tr>"
            return out
        student_rows = rows_html(students, ["student_no", "name_ko", "phone", "status"])
        schedule_rows = ""
        if not schedules:
            schedule_rows = f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>"
        else:
            for r in sorted(schedules, key=lambda x: (day_sort_value(x['day_of_week']), x['start_time'] or '')):
                schedule_rows += f"<tr><td>{r['day_of_week'] or '-'}</td><td>{r['start_time'] or '-'} ~ {r['end_time'] or '-'}</td><td>{r['classroom'] or '-'}</td><td><span class='badge {r['status'] or ''}'>{status_t(r['status']) if r['status'] else '-'}</span></td><td>{r['note'] or '-'}</td></tr>"
        attendance_html = rows_html(attendance_rows, ["lesson_date", "student_name", "status", "note"])
        homework_html = rows_html(homework_rows, ["title", "due_date", "submitted_count", "total_submissions"])
        exam_html = rows_html(exam_rows, ["exam_name", "exam_date", "avg_score", "score_count"])
        next_order = "desc" if student_order == "asc" else "asc"
        html = render_html(t('academics.class_detail.title'), f"""
        <div class='card'>
        <div><a href='/masterdata?lang={CURRENT_LANG}'>← {t('academics.back_to_list')}</a></div>
        <h3>{class_row['name']}</h3>
        <table>
          <tr><th>{t('academics.basic_info')}</th><td>{t('academics.class_name')}: {class_row['name']}</td></tr>
          <tr><th>{t('academics.course_level')}</th><td>{class_row['course_name'] or '-'} / {class_row['level_name'] or '-'}</td></tr>
          <tr><th>{t('academics.foreign_teacher')}</th><td>{class_row['foreign_teacher_name'] or '-'}</td></tr>
          <tr><th>{t('academics.chinese_teacher')}</th><td>{class_row['chinese_teacher_name'] or '-'}</td></tr>
          <tr><th>{t('academics.status')}</th><td>{status_t(class_row['status']) if class_row['status'] else '-'}</td></tr>
          <tr><th>{t('field.note')}</th><td>{class_row['memo'] or '-'}</td></tr>
          <tr><th>{t('academics.student_count')}</th><td>{len(students)}</td></tr>
        </table>
        </div>
        <div class='card'>
        <h4>{t('common.edit')}</h4>
        <form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'>
          <input type='hidden' name='type' value='class_edit'>
          <label>{t('academics.class_name')} <input name='name' value='{h(class_row['name'] or '')}'></label>
          <label>{t('academics.course')} <select name='course_id'><option value=''>-</option>{''.join([f"<option value='{co['id']}' {'selected' if str(class_row['course_id'] or '')==str(co['id']) else ''}>{co['name']}</option>" for co in conn.execute('SELECT id, name FROM courses ORDER BY name').fetchall()])}</select></label>
          <label>{t('academics.level')} <select name='level_id'><option value=''>-</option>{''.join([f"<option value='{lv['id']}' {'selected' if str(class_row['level_id'] or '')==str(lv['id']) else ''}>{lv['name']} ({lv['course_name'] or '-'})</option>" for lv in conn.execute('SELECT l.id, l.name, c.name AS course_name FROM levels l LEFT JOIN courses c ON c.id=l.course_id ORDER BY l.name').fetchall()])}</select></label>
          <label>{t('academics.foreign_teacher')} <select name='foreign_teacher_id'><option value=''>-</option>{''.join([f"<option value='{tr['id']}' {'selected' if str(class_row['foreign_teacher_id'] or class_row['teacher_id'] or '')==str(tr['id']) else ''}>{tr['name']}</option>" for tr in list_teacher_profiles(conn)])}</select></label>
          <label>{t('academics.chinese_teacher')} <select name='chinese_teacher_id'><option value=''>-</option>{''.join([f"<option value='{tr['id']}' {'selected' if str(class_row['chinese_teacher_id'] or '')==str(tr['id']) else ''}>{tr['name']}</option>" for tr in list_teacher_profiles(conn)])}</select></label>
          <label>{t('academics.status')} <select name='status'><option value='active' {'selected' if (class_row['status'] or 'active')=='active' else ''}>{status_t('active')}</option><option value='inactive' {'selected' if (class_row['status'] or '')=='inactive' else ''}>{t('status.ended')}</option></select></label>
          <label>{t('field.note')} <input name='memo' value='{h(class_row['memo'] or '')}'></label>
          <button>{t('common.save')}</button>
        </form>
        </div>
        <div class='card'>
        <h4>{t('academics.students')}</h4>
        <div style='margin-bottom:8px'>
          {t('academics.sort')}: {student_sort} ({student_order}) |
          <a href='?lang={CURRENT_LANG}&student_sort=name_ko&student_order={next_order}'>{t('academics.sort_name')}</a> |
          <a href='?lang={CURRENT_LANG}&student_sort=student_no&student_order={next_order}'>{t('academics.sort_student_no')}</a> |
          <a href='?lang={CURRENT_LANG}&student_sort=status&student_order={next_order}'>{t('academics.sort_status')}</a> |
          <a href='?lang={CURRENT_LANG}&student_sort={student_sort}&student_order={student_order}&export=students_csv'>{t('academics.export_students_csv')}</a>
        </div>
        <table>
          <tr><th>{t('students.field.student_no')}</th><th>{t('field.name')}</th><th>{t('students.field.phone')}</th><th>{t('field.status')}</th></tr>
          {student_rows}
        </table>
        </div>
        <div class='card'>
        <h4>{t('academics.schedule')}</h4>
        <table>
          <tr><th>{t('academics.day_of_week')}</th><th>{t('academics.time_slot')}</th><th>{t('academics.classroom')}</th><th>{t('academics.status')}</th><th>{t('field.note')}</th></tr>
          {schedule_rows}
        </table>
        </div>
        <div class='card'>
        <h4>{t('academics.recent_attendance')}</h4>
        <div style='margin-bottom:8px'>
          <a href='?lang={CURRENT_LANG}&student_sort={student_sort}&student_order={student_order}&export=attendance_csv'>{t('academics.export_attendance_csv')}</a>
        </div>
        <table>
          <tr><th>{t('field.date')}</th><th>{t('field.student')}</th><th>{t('field.status')}</th><th>{t('field.note')}</th></tr>
          {attendance_html}
        </table>
        </div>
        <div class='card'>
        <h4>{t('academics.recent_homework')}</h4>
        <table>
          <tr><th>{t('field.homework_title')}</th><th>{t('field.due_date')}</th><th>{t('field.submission_count')}</th><th>{t('field.total_targets')}</th></tr>
          {homework_html}
        </table>
        </div>
        <div class='card'>
        <h4>{t('academics.recent_exams')}</h4>
        <table>
          <tr><th>{t('field.exam_name')}</th><th>{t('field.exam_date')}</th><th>{t('field.avg_score')}</th><th>{t('field.score_entries')}</th></tr>
          {exam_html}
        </table>
        </div>
        """, user, current_menu="masterdata")
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/academics":
        status, headers, body = redirect(f"/schedule?lang={CURRENT_LANG}")
        start_response(status, headers)
        return [body]

    if path == "/masterdata":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        flash_msg = ""
        flash_type = "success"
        md_view = query.get("md_view", "classes")
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            data = parse_body(environ)
            typ = data.get("type")
            errs = []
            try:
                if typ == "course":
                    if not (data.get("name") or "").strip():
                        add_error(errs, "course.name", "필수값입니다")
                    else:
                        conn.execute("INSERT INTO courses(name, created_at) VALUES(?,?)", (data.get("name").strip(), now()))
                elif typ == "level":
                    if not ensure_exists(conn, "courses", data.get("course_id")):
                        add_error(errs, "level.course_id", "존재하지 않는 코스입니다")
                    if not (data.get("name") or "").strip():
                        add_error(errs, "level.name", "필수값입니다")
                    if not errs:
                        conn.execute("INSERT INTO levels(course_id, name, created_at) VALUES(?,?,?)", (data.get("course_id"), data.get("name").strip(), now()))
                elif typ == "class":
                    class_name = (data.get("name") or "").strip()
                    course_id = (data.get("course_id") or "").strip()
                    level_id = (data.get("level_id") or "").strip()
                    foreign_teacher_id = (data.get("foreign_teacher_id") or data.get("teacher_id") or "").strip()
                    chinese_teacher_id = (data.get("chinese_teacher_id") or "").strip()
                    class_status = (data.get("status") or "active").strip() or "active"
                    class_memo = (data.get("memo") or "").strip()
                    if not ensure_exists(conn, "courses", course_id):
                        add_error(errs, "class.course_id", "Invalid course")
                    if level_id and not ensure_exists(conn, "levels", level_id):
                        add_error(errs, "class.level_id", "Invalid level")
                    if foreign_teacher_id and not ensure_exists(conn, "teachers", foreign_teacher_id, field="user_id"):
                        add_error(errs, "class.foreign_teacher_id", "Invalid foreign teacher")
                    if chinese_teacher_id and not ensure_exists(conn, "teachers", chinese_teacher_id, field="user_id"):
                        add_error(errs, "class.chinese_teacher_id", "Invalid chinese teacher")
                    if class_status not in ("active", "inactive"):
                        add_error(errs, "class.status", "Invalid status")
                    if not class_name:
                        add_error(errs, "class.name", "Required")
                    if not errs:
                        conn.execute(
                            """INSERT INTO classes(course_id, level_id, name, teacher_id, foreign_teacher_id, chinese_teacher_id, status, memo, created_at)
                            VALUES(?,?,?,?,?,?,?,?,?)""",
                            (
                                course_id,
                                level_id or None,
                                class_name,
                                foreign_teacher_id or None,
                                foreign_teacher_id or None,
                                chinese_teacher_id or None,
                                class_status,
                                class_memo or None,
                                now(),
                            ),
                        )
                elif typ == "classroom":
                    room_name = (data.get("name") or "").strip()
                    if not room_name:
                        add_error(errs, "classroom.name", "필수값입니다")
                    else:
                        conn.execute("INSERT OR IGNORE INTO classrooms(name, room_name, room_code, status, created_at) VALUES(?,?,?,?,?)", (room_name, room_name, room_name, "active", now()))
                elif typ == "time_slot":
                    st = (data.get("start_time") or "").strip()
                    et = (data.get("end_time") or "").strip()
                    if not st or not et or parse_hhmm(st) is None or parse_hhmm(et) is None:
                        add_error(errs, "time_slot", "시작/종료 시간을 HH:MM 형식으로 입력하세요")
                    elif parse_hhmm(et) <= parse_hhmm(st):
                        add_error(errs, "time_slot", "종료시간은 시작시간보다 커야 합니다")
                    else:
                        label = f"{st}~{et}"
                        conn.execute("INSERT OR IGNORE INTO time_slots(label, start_time, end_time, created_at) VALUES(?,?,?,?)", (label, st, et, now()))
                elif typ in ("delete_course", "delete_level", "delete_class", "delete_classroom", "delete_time_slot"):
                    del_id = data.get("id")
                    table_map = {"delete_course": "courses", "delete_level": "levels", "delete_class": "classes", "delete_classroom": "classrooms", "delete_time_slot": "time_slots"}
                    table = table_map[typ]
                    if not ensure_exists(conn, table, del_id):
                        add_error(errs, "delete.id", "삭제할 데이터가 없습니다")
                    elif table == "courses" and conn.execute("SELECT 1 FROM classes WHERE course_id=? LIMIT 1", (del_id,)).fetchone():
                        add_error(errs, "course", "반에서 사용 중인 코스는 삭제할 수 없습니다")
                    elif table == "levels" and conn.execute("SELECT 1 FROM classes WHERE level_id=? LIMIT 1", (del_id,)).fetchone():
                        add_error(errs, "level", "반에서 사용 중인 레벨은 삭제할 수 없습니다")
                    elif table == "classes" and conn.execute("SELECT 1 FROM schedules WHERE class_id=? LIMIT 1", (del_id,)).fetchone():
                        add_error(errs, "class", "시간표에서 사용 중인 반은 삭제할 수 없습니다")
                    if not errs:
                        conn.execute(f"DELETE FROM {table} WHERE id=?", (del_id,))
                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                    log_event(conn, "ERROR", path, "마스터데이터 검증 실패", "\n".join(errs), user["id"])
                else:
                    conn.commit()
                    flash_msg = "삭제되었습니다" if str(typ).startswith("delete_") else t("common.save")
            except Exception:
                conn.rollback()
                flash_msg = "저장/삭제 처리 중 오류가 발생했습니다"
                flash_type = "error"
                log_event(conn, "ERROR", path, "마스터데이터 예외", traceback.format_exc(), user["id"])
                conn.commit()

        load_master = query.get("load", "") == "1"
        courses_ref = conn.execute("SELECT id, name FROM courses ORDER BY id DESC").fetchall()
        levels_ref = conn.execute("""SELECT l.id, l.name, c.name AS course_name, l.course_id
                               FROM levels l LEFT JOIN courses c ON c.id=l.course_id ORDER BY l.id DESC""").fetchall()
        teachers = list_teacher_profiles(conn)

        courses = []
        levels = []
        classes = []
        classrooms = []
        time_slots = []
        if load_master:
            courses = conn.execute("SELECT id, name, created_at FROM courses ORDER BY id DESC").fetchall()
            levels = conn.execute("""SELECT l.id, l.name, c.name AS course_name, l.course_id, l.created_at
                               FROM levels l LEFT JOIN courses c ON c.id=l.course_id ORDER BY l.id DESC""").fetchall()
            classes = conn.execute("""SELECT c.id, c.name, c.teacher_id, c.foreign_teacher_id, c.chinese_teacher_id,
                    COALESCE(c.status, 'active') AS status, COALESCE(c.memo, '') AS memo,
                    co.name AS course_name, l.name AS level_name,
                    uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name,
                    (SELECT COUNT(*) FROM students s WHERE s.current_class_id=c.id) AS student_count,
                    (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=c.id ORDER BY s2.id) AS student_names
                    FROM classes c
                    LEFT JOIN courses co ON co.id=c.course_id
                    LEFT JOIN levels l ON l.id=c.level_id
                    LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                    LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                    ORDER BY c.id DESC""").fetchall()
            classrooms = conn.execute("SELECT id, name, created_at FROM classrooms ORDER BY id DESC").fetchall()
            time_slots = conn.execute("SELECT id, label, start_time, end_time, created_at FROM time_slots ORDER BY id DESC").fetchall()
        teacher_options = "".join([f"<option value='{tr['id']}'>{tr['name']} [{op_code('T', tr['id'])}] ({tr['username']}, {tr['teacher_type'] or '-'})</option>" for tr in teachers])

        def rows_html(rows, cols):
            if not rows:
                return f"<tr><td colspan='{len(cols)}' class='empty-msg'>{t('common.no_data')}</td></tr>"
            out = ""
            for r in rows:
                out += "<tr>" + "".join([f"<td>{r[c] if r[c] not in (None, '') else '-'}</td>" for c in cols]) + "</tr>"
            return out

        section_nav = f"""
        <div class='card'><div class='filter-row'>
          <a class='btn {'secondary' if md_view!='classes' else ''}' href='/masterdata?lang={CURRENT_LANG}&md_view=classes'>{t('academics.class_list')}</a>
          <a class='btn {'secondary' if md_view!='courses' else ''}' href='/masterdata?lang={CURRENT_LANG}&md_view=courses'>{t('academics.course_level')}</a>
          <a class='btn {'secondary' if md_view!='teachers' else ''}' href='/masterdata?lang={CURRENT_LANG}&md_view=teachers'>{t('academics.teacher')}</a>
          <a class='btn {'secondary' if md_view!='rooms' else ''}' href='/masterdata?lang={CURRENT_LANG}&md_view=rooms'>{t('academics.classroom')} / {t('academics.time_slot')}</a>
        </div></div>
        """

        class_rows_md = "".join([
            f"<tr><td><a href='/classes/{c['id']}?lang={CURRENT_LANG}'>{op_code('C', c['id'])} · {c['name']}</a></td><td>{c['course_name'] or '-'}</td><td>{c['level_name'] or '-'}</td><td>{c['foreign_teacher_name'] or '-'}</td><td>{c['chinese_teacher_name'] or '-'}</td><td title='{h(summarize_student_names(c['student_names'])[1])}'>{h(summarize_student_names(c['student_names'])[0])}</td><td>{status_t(c['status']) if c['status'] else '-'}</td><td>{c['student_count'] or 0}</td><td><form method='post'><input type='hidden' name='type' value='delete_class'><input type='hidden' name='id' value='{c['id']}'><button class='btn secondary'>{t('common.delete')}</button></form></td></tr>"
            for c in classes
        ])
        course_rows_md = "".join([
            f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['created_at']}</td><td><form method='post'><input type='hidden' name='type' value='delete_course'><input type='hidden' name='id' value='{r['id']}'><button class='btn secondary'>{t('common.delete')}</button></form></td></tr>"
            for r in courses
        ])
        level_rows_md = "".join([
            f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['course_name'] or '-'}</td><td>{r['created_at']}</td><td><form method='post'><input type='hidden' name='type' value='delete_level'><input type='hidden' name='id' value='{r['id']}'><button class='btn secondary'>{t('common.delete')}</button></form></td></tr>"
            for r in levels
        ])
        room_rows_md = "".join([
            f"<tr><td>{r['id']}</td><td>{r['name'] or '-'}</td><td>{r['created_at']}</td><td><form method='post'><input type='hidden' name='type' value='delete_classroom'><input type='hidden' name='id' value='{r['id']}'><button class='btn secondary'>{t('common.delete')}</button></form></td></tr>"
            for r in classrooms
        ])
        slot_rows_md = "".join([
            f"<tr><td>{r['id']}</td><td>{r['label']}</td><td>{r['start_time']}</td><td>{r['end_time']}</td><td>{r['created_at']}</td><td><form method='post'><input type='hidden' name='type' value='delete_time_slot'><input type='hidden' name='id' value='{r['id']}'><button class='btn secondary'>{t('common.delete')}</button></form></td></tr>"
            for r in time_slots
        ])
        empty9 = f"<tr><td colspan='9' class='empty-msg'>{t('common.no_data')}</td></tr>"
        empty6 = f"<tr><td colspan='6' class='empty-msg'>{t('common.no_data')}</td></tr>"
        empty5 = f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>"
        empty4 = f"<tr><td colspan='4' class='empty-msg'>{t('common.no_data')}</td></tr>"
        empty3 = f"<tr><td colspan='3' class='empty-msg'>{t('common.no_data')}</td></tr>"
        helper_html = "" if load_master else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")
        classes_card = f"<div class='card'><h4>{t('academics.class_list')}</h4>{helper_html}<div class='table-wrap'><table><tr><th>{t('academics.class_name')}</th><th>{t('academics.course')}</th><th>{t('academics.level')}</th><th>{t('academics.foreign_teacher')}</th><th>{t('academics.chinese_teacher')}</th><th>{t('academics.students')}</th><th>{t('academics.status')}</th><th>{t('academics.student_count')}</th><th>{t('common.delete')}</th></tr>{class_rows_md if load_master else ''}{empty9 if (load_master and not class_rows_md) else ''}</table></div></div>"
        course_cards = f"<div class='card'><h4>{t('academics.course')}</h4>{helper_html}<div class='table-wrap'><table><tr><th>{t('field.id')}</th><th>{t('academics.course_name')}</th><th>{t('field.created_at')}</th><th>{t('common.delete')}</th></tr>{course_rows_md if load_master else ''}{empty4 if (load_master and not course_rows_md) else ''}</table></div></div><div class='card'><h4>{t('academics.level')}</h4><div class='table-wrap'><table><tr><th>{t('field.id')}</th><th>{t('academics.level_name')}</th><th>{t('academics.course')}</th><th>{t('field.created_at')}</th><th>{t('common.delete')}</th></tr>{level_rows_md if load_master else ''}{empty5 if (load_master and not level_rows_md) else ''}</table></div></div>"
        teacher_rows_md = ''.join([f"<tr><td>{op_code('T', r['id'])}</td><td>{r['name']}</td><td>{r['username']}</td></tr>" for r in teachers])
        teacher_card = f"<div class='card'><h4>{t('academics.teacher')}</h4><table><tr><th>{t('field.id')}</th><th>{t('field.name')}</th><th>{t('login.username')}</th></tr>{teacher_rows_md or empty3}</table></div>"
        room_cards = f"<div class='card'><h4>{t('academics.classroom')}</h4><table><tr><th>{t('field.id')}</th><th>{t('academics.classroom')}</th><th>{t('field.created_at')}</th><th>{t('common.delete')}</th></tr>{room_rows_md or empty4}</table></div><div class='card'><h4>{t('academics.time_slot')}</h4><table><tr><th>{t('field.id')}</th><th>{t('academics.time_slot')}</th><th>{t('academics.start_time')}</th><th>{t('academics.end_time')}</th><th>{t('field.created_at')}</th><th>{t('common.delete')}</th></tr>{slot_rows_md or empty6}</table></div>"
        section_body = classes_card if md_view == 'classes' else course_cards if md_view == 'courses' else teacher_card if md_view == 'teachers' else room_cards

        html = render_html(t('menu.masterdata'), f"""
        <div class='card'><h4>{t('menu.masterdata')}</h4><div class='muted'>{t('academics.go_structure')}</div></div>
        {section_nav}
        <div class='card'>
          <form method='get' class='btn-row query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='md_view' value='{md_view}'>
            <input type='hidden' name='load' value='1'>
            <button>{t('common.search')}</button>
            <a class='btn secondary' href='/masterdata?lang={CURRENT_LANG}&md_view={md_view}'>{t('common.reset')}</a>
          </form>
        </div>
        <div class='card'>
          <h4>{t('academics.register')}</h4>
          <form method='post' class='form-row'><input type='hidden' name='type' value='course'><label>{t('academics.course_name')} <input name='name'></label><button>{t('common.add')}</button></form>
          <form method='post' class='form-row'><input type='hidden' name='type' value='level'><label>{t('academics.level_name')} <input name='name'></label><label>{t('academics.course')} <select name='course_id'><option value=''>-</option>{''.join([f"<option value='{c['id']}'>{c['name']}</option>" for c in courses_ref])}</select></label><button>{t('common.add')}</button></form>
          <form method='post' class='form-row'><input type='hidden' name='type' value='class'><label>{t('academics.class_name')} <input name='name'></label><label>{t('academics.course')} <select name='course_id'><option value=''>-</option>{''.join([f"<option value='{c['id']}'>{c['name']}</option>" for c in courses_ref])}</select></label><label>{t('academics.level')} <select name='level_id'><option value=''>-</option>{''.join([f"<option value='{lv['id']}'>{lv['name']} ({lv['course_name'] or '-'})</option>" for lv in levels_ref])}</select></label><label>{t('academics.foreign_teacher')} <select name='foreign_teacher_id'><option value=''>-</option>{teacher_options}</select></label><label>{t('academics.chinese_teacher')} <select name='chinese_teacher_id'><option value=''>-</option>{teacher_options}</select></label><label>{t('academics.status')} <select name='status'><option value='active'>{status_t('active')}</option><option value='inactive'>{t('status.ended')}</option></select></label><label>{t('field.note')} <input name='memo'></label><button>{t('common.add')}</button></form>
          <form method='post' class='form-row'><input type='hidden' name='type' value='classroom'><label>{t('academics.classroom')} <input name='name'></label><button>{t('common.add')}</button></form>
          <form method='post' class='form-row'><input type='hidden' name='type' value='time_slot'><label>{t('academics.start_time')} <input type='time' name='start_time'></label><label>{t('academics.end_time')} <input type='time' name='end_time'></label><button>{t('common.add')}</button></form>
        </div>
        {section_body}
        """, user, current_menu="masterdata", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]

    if path == "/schedule":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]

        week_offset = safe_int(query.get("week", "0"), 0)
        ref_date_str = (query.get("ref_date", "") or "").strip()
        selected_day_query = (query.get("day", "") or "").strip().lower()
        selected_teacher_id = query.get("teacher_id", "")
        selected_room = query.get("classroom", "").strip()
        selected_course_level = ""
        selected_class_q = ""
        selected_schedule_id = query.get("schedule_id", "")
        selected_form_class_id = query.get("selected_form_class_id", "")
        recent_class_ids = query.get("recent_class_ids", "")

        flash_msg = ""
        flash_type = "success"
        week_start, week_end, week_year, week_no = week_bounds_from_ref(ref_date_str, week_offset)
        if not ref_date_str:
            ref_date_str = datetime.utcnow().date().isoformat()
        selected_weekday, selected_view_date = weekday_from_ref(ref_date_str, week_offset)
        selected_day = selected_day_query.capitalize() if selected_day_query.capitalize() in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun") else selected_weekday
        print_mode = query.get("print", "") == "1"

        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            data = parse_body(environ)
            typ = data.get("type")
            if typ == "copy_week":
                source_week_start = iso_monday_str(ref_date_str, week_offset)
                target_week_start = (datetime.strptime(source_week_start, "%Y-%m-%d").date() + timedelta(days=7)).isoformat()
                copy_class_rows = conn.execute("SELECT id FROM classes").fetchall()
                copy_class_ids = [str(r["id"]) for r in copy_class_rows]
                result = copy_week_schedules(conn, source_week_start, target_week_start, class_ids=copy_class_ids)
                conn.commit()
                if result["source_count"] == 0:
                    flash_msg = t("academics.copy_week.none")
                    flash_type = "error"
                else:
                    flash_msg = f"{t('academics.copy_week.done')}: {t('academics.copy_week.copied')} {result['copied']}, {t('academics.copy_week.skipped')} {result['skipped']}"
            elif typ == "course":
                conn.execute("INSERT INTO courses(name, created_at) VALUES(?,?)", (data.get("name"), now()))
                conn.commit()
            elif typ == "level":
                conn.execute("INSERT INTO levels(course_id, name, created_at) VALUES(?,?,?)", (data.get("course_id"), data.get("name"), now()))
                conn.commit()
            elif typ == "class":
                conn.execute("INSERT INTO classes(course_id, level_id, name, teacher_id, created_at) VALUES(?,?,?,?,?)", (data.get("course_id"), data.get("level_id"), data.get("name"), data.get("teacher_id"), now()))
                conn.commit()
            elif typ == "schedule":
                schedule_id = data.get("schedule_id", "").strip()
                class_id = (data.get("class_id") or data.get("selected_form_class_id") or "").strip()
                day_of_week = (data.get("day_of_week") or "").strip()
                time_slot_value = (data.get("time_slot") or "").strip()
                start_time = (data.get("start_time") or "").strip()
                end_time = (data.get("end_time") or "").strip()
                if time_slot_value and "|" in time_slot_value:
                    sp = time_slot_value.split("|", 1)
                    start_time, end_time = sp[0].strip(), sp[1].strip()
                classroom = (data.get("classroom") or "").strip()
                sc_status = (data.get("status") or "active").strip() or "active"
                note = (data.get("note") or "").strip()
                selected_teacher = (data.get("teacher_id") or "").strip()
                form_teacher_id = selected_teacher if selected_teacher else None

                if not class_id:
                    flash_msg = t("academics.validation_class_required")
                    flash_type = "error"
                elif not ensure_exists(conn, "classes", class_id):
                    flash_msg = "class_id: invalid class"
                    flash_type = "error"
                elif form_teacher_id and not ensure_exists(conn, "teachers", form_teacher_id, field="user_id"):
                    flash_msg = "teacher_id: 존재하지 않는 강사입니다"
                    flash_type = "error"
                elif not classroom:
                    flash_msg = f"{t('academics.classroom')}: required"
                    flash_type = "error"
                elif not ensure_exists(conn, "classrooms", classroom, field="name") and not ensure_exists(conn, "classrooms", classroom, field="room_name"):
                    flash_msg = "classroom: invalid room"
                    flash_type = "error"
                elif not day_of_week:
                    flash_msg = f"{t('academics.day_of_week')}: required"
                    flash_type = "error"
                elif day_of_week not in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"):
                    flash_msg = "day_of_week: invalid value"
                    flash_type = "error"
                elif not time_slot_value and (not start_time or not end_time):
                    flash_msg = f"{t('academics.time_slot')}: required"
                    flash_type = "error"
                elif not start_time or not end_time or parse_hhmm(start_time) is None or parse_hhmm(end_time) is None:
                    flash_msg = "start_time/end_time: invalid HH:MM"
                    flash_type = "error"
                elif parse_hhmm(end_time) is not None and parse_hhmm(start_time) is not None and parse_hhmm(end_time) <= parse_hhmm(start_time):
                    flash_msg = t("academics.validation_end_before_start")
                    flash_type = "error"
                else:
                    class_row = conn.execute("SELECT id, teacher_id FROM classes WHERE id=?", (class_id,)).fetchone()
                    class_teacher_id = class_row["teacher_id"] if class_row else None
                    form_teacher_id = selected_teacher if selected_teacher else class_teacher_id
                    ignore_id = schedule_id if str(schedule_id).isdigit() else "0"
                    target_week_start = iso_monday_str(ref_date_str, week_offset)
                    conflict = find_schedule_conflict(conn, class_id, day_of_week, start_time, end_time, classroom, form_teacher_id, target_week_start, ignore_id=ignore_id)
                    if conflict:
                        flash_msg = conflict
                        flash_type = "error"
                        log_event(conn, "ERROR", path, "시간표 중복 검증 실패", conflict, user["id"])
                    else:
                        if str(schedule_id).isdigit():
                            conn.execute(
                                """UPDATE schedules SET class_id=?, day_of_week=?, start_time=?, end_time=?, classroom=?, status=?, note=?, teacher_id=?
                                WHERE id=?""",
                                (class_id, day_of_week, start_time, end_time, classroom or None, sc_status, note or None, form_teacher_id, schedule_id),
                            )
                            flash_msg = t("academics.updated")
                        else:
                            conn.execute(
                                """INSERT INTO schedules(class_id, day_of_week, start_time, end_time, classroom, status, note, teacher_id, created_at, week_start_date)
                                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                                (class_id, day_of_week, start_time, end_time, classroom or None, sc_status, note or None, form_teacher_id, now(), iso_monday_str(ref_date_str, week_offset)),
                            )
                            flash_msg = t("academics.saved")
                        conn.commit()
            if flash_type == "error" and flash_msg:
                log_event(conn, "ERROR", path, "시간표 저장 실패", flash_msg, user["id"])

        courses = conn.execute("SELECT id, name, created_at FROM courses ORDER BY id DESC").fetchall()
        levels = conn.execute("""SELECT l.id, l.name, l.course_id, c.name AS course_name, l.created_at
                               FROM levels l LEFT JOIN courses c ON c.id=l.course_id
                               ORDER BY l.id DESC""").fetchall()
        if has_role(user, [ROLE_TEACHER]):
            classes = conn.execute(
                """SELECT c.id, c.name, c.teacher_id, c.foreign_teacher_id, c.chinese_teacher_id,
                co.name AS course_name, l.name AS level_name,
                uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name,
                (SELECT COUNT(*) FROM students s WHERE s.current_class_id=c.id) AS student_count,
                (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=c.id ORDER BY s2.id) AS student_names
                FROM classes c
                LEFT JOIN courses co ON co.id=c.course_id
                LEFT JOIN levels l ON l.id=c.level_id
                LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                WHERE COALESCE(c.foreign_teacher_id, c.teacher_id)=?
                ORDER BY c.id DESC""",
                (user["id"],),
            ).fetchall()
        else:
            classes = conn.execute(
                """SELECT c.id, c.name, c.teacher_id, c.foreign_teacher_id, c.chinese_teacher_id,
                co.name AS course_name, l.name AS level_name,
                uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name,
                (SELECT COUNT(*) FROM students s WHERE s.current_class_id=c.id) AS student_count,
                (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=c.id ORDER BY s2.id) AS student_names
                FROM classes c
                LEFT JOIN courses co ON co.id=c.course_id
                LEFT JOIN levels l ON l.id=c.level_id
                LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
                LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
                ORDER BY c.id DESC"""
            ).fetchall()

        class_ids = [str(c["id"]) for c in classes]
        if class_ids:
            schedule_sql = f"""SELECT sc.id, sc.class_id, c.name AS class_name, c.teacher_id AS class_teacher_id,
            c.foreign_teacher_id AS class_foreign_teacher_id, c.chinese_teacher_id AS class_chinese_teacher_id,
            sc.teacher_id, sc.foreign_teacher_id, sc.chinese_teacher_id,
            COALESCE(sc.teacher_id, sc.foreign_teacher_id, c.foreign_teacher_id, c.teacher_id) AS effective_teacher_id,
            COALESCE(uf2.name, u2.name, uf.name) AS foreign_teacher_name,
            uc.name AS chinese_teacher_name,
            co.name AS course_name, l.name AS level_name, sc.day_of_week, sc.start_time, sc.end_time,
            COALESCE(sc.classroom,'') AS classroom, COALESCE(sc.status,'active') AS status, COALESCE(sc.note,'') AS note,
            (SELECT COUNT(*) FROM students s WHERE s.current_class_id=sc.class_id) AS student_count,
            (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=sc.class_id ORDER BY s2.id) AS student_names
            FROM schedules sc
            LEFT JOIN classes c ON c.id=sc.class_id
            LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
            LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
            LEFT JOIN users u2 ON u2.id=sc.teacher_id
            LEFT JOIN users uf2 ON uf2.id=sc.foreign_teacher_id
            LEFT JOIN courses co ON co.id=c.course_id
            LEFT JOIN levels l ON l.id=c.level_id
            WHERE sc.class_id IN ({','.join(['?']*len(class_ids))})
              AND COALESCE(sc.week_start_date,'') IN ('', ?)"""
            schedules = conn.execute(schedule_sql, [*class_ids, week_start.isoformat()]).fetchall()
        else:
            schedules = []

        schedules = [r for r in schedules if (r["day_of_week"] or "") == selected_day]
        if selected_teacher_id:
            schedules = [r for r in schedules if str(r["effective_teacher_id"] or "") == selected_teacher_id]
        if selected_room:
            schedules = [r for r in schedules if selected_room.lower() in (r["classroom"] or "").lower()]

        slot_rows_grid = conn.execute("SELECT start_time, end_time FROM time_slots ORDER BY start_time, end_time").fetchall()
        time_slots = [f"{r['start_time']}~{r['end_time']}" for r in slot_rows_grid if r['start_time'] and r['end_time']]
        if not time_slots:
            time_slots = ["16:25~17:20", "17:25~18:20", "18:30~19:25", "19:35~20:30"]

        grouped = {}
        teacher_rooms = {}
        for r in schedules:
            slot = f"{r['start_time']}~{r['end_time']}"
            rowkey = str(r["effective_teacher_id"] or "")
            grouped.setdefault((rowkey, slot), []).append(r)
            teacher_rooms.setdefault(rowkey, set())
            if (r["classroom"] or "").strip():
                teacher_rooms[rowkey].add((r["classroom"] or "").strip())

        row_headers = []
        row_seen = set()
        for r in schedules:
            rowkey = str(r["effective_teacher_id"] or "")
            if rowkey in row_seen:
                continue
            row_seen.add(rowkey)
            rooms = ", ".join(sorted(teacher_rooms.get(rowkey) or []))
            row_headers.append((rowkey, r["foreign_teacher_name"] or "-", rooms))
        row_headers.sort(key=lambda x: (x[1], x[2]))

        selected_schedule = None
        if str(selected_schedule_id).isdigit():
            for r in schedules:
                if str(r["id"]) == selected_schedule_id:
                    selected_schedule = r
                    break

        class_rows = ""
        if classes:
            for c in classes:
                class_rows += f"""
                <tr>
                  <td><a href='/classes/{c['id']}?lang={CURRENT_LANG}'>{op_code('C', c['id'])} · {c['name']}</a></td>
                  <td>{c['course_name'] or '-'}</td>
                  <td>{c['level_name'] or '-'}</td>
                  <td>{c['foreign_teacher_name'] or '-'}</td>
                  <td>{c['chinese_teacher_name'] or '-'}</td>
                  <td>{(c['student_names'] or '-')}</td>
                </tr>
                """
        else:
            class_rows = f"<tr><td colspan='6' class='empty-msg'>{t('common.no_data')}</td></tr>"

        def rows_html(rows, cols):
            if not rows:
                return f"<tr><td colspan='{len(cols)}'>{t('common.no_data')}</td></tr>"
            out = ""
            for r in rows:
                out += "<tr>" + "".join([f"<td>{r[c] if r[c] not in (None, '') else '-'}</td>" for c in cols]) + "</tr>"
            return out

        course_rows = rows_html(courses, ["id", "name", "created_at"])
        level_rows = rows_html(levels, ["id", "name", "course_name", "created_at"])

        def compact_student_names(raw_names, limit=3):
            names = [n.strip() for n in (raw_names or "").split(",") if n.strip()]
            if not names:
                return "-"
            if len(names) <= limit:
                return ", ".join(names)
            return ", ".join(names[:limit]) + f" +{len(names) - limit}"

        room_grouped = {}
        room_names = []
        for r in schedules:
            room_name = (r["classroom"] or "-").strip() or "-"
            if room_name not in room_grouped:
                room_grouped[room_name] = {}
                room_names.append(room_name)
            slot = f"{r['start_time']}~{r['end_time']}"
            room_grouped[room_name][slot] = r
        room_names.sort()

        print_room_sections = ""
        for room_name in room_names:
            slot_cells = ""
            for slot in time_slots:
                les = room_grouped.get(room_name, {}).get(slot)
                if les:
                    students_compact = compact_student_names(les["student_names"], limit=2)
                    teacher_meta = f"{les['foreign_teacher_name'] or '-'}"
                    if les["chinese_teacher_name"]:
                        teacher_meta += f" / {les['chinese_teacher_name']}"
                    slot_cells += f"""
                    <td>
                      <div class='print-lesson-cell'>
                        <div class='print-class-name'>{les['class_name'] or '-'}</div>
                        <div class='print-class-meta'>{les['course_name'] or '-'} / {les['level_name'] or '-'}</div>
                        <div class='print-class-meta'>{teacher_meta}</div>
                        <div class='print-students'>{students_compact}</div>
                      </div>
                    </td>
                    """
                else:
                    slot_cells += "<td><div class='print-lesson-cell'></div></td>"
            print_room_sections += f"""
            <section class='print-room-card'>
              <div class='print-room-title'>{t('academics.classroom')}: {room_name}</div>
              <table class='print-room-table'>
                <tr>{''.join([f"<th>{slot}</th>" for slot in time_slots])}</tr>
                <tr>{slot_cells}</tr>
              </table>
            </section>
            """

        timetable_cols = ["<div class='tt-head'>" + t("academics.teacher_room") + "</div>"] + [f"<div class='tt-head'>{slot}</div>" for slot in time_slots]
        timetable_cells = ""
        for rowkey, tname, room in row_headers:
            room_meta = f"<div class='muted'>{t('academics.classroom')}: {room}</div>" if room else ""
            timetable_cells += f"<div class='tt-rowhead'><strong>{tname}</strong>{room_meta}</div>"
            for slot in time_slots:
                lessons = grouped.get((rowkey, slot), [])
                blocks = ""
                for les in lessons:
                    students_label = (les['student_names'] or '').strip()
                    meta_bits = [
                        f"{les['course_name'] or '-'} / {les['level_name'] or '-'}",
                        f"{t('academics.classroom')}: {les['classroom'] or '-'}",
                    ]
                    if les['chinese_teacher_name']:
                        meta_bits.append(les['chinese_teacher_name'])
                    meta_html = " | ".join(meta_bits)
                    status_html = f"<span class='badge {les['status'] or ''}'>{status_t(les['status']) if les['status'] else '-'}</span>"
                    blocks += f"""
                    <div class='lesson-block'>
                      <div class='lesson-title'>{les['class_name'] or '-'}</div>
                      <div class='lesson-meta'>{meta_html}</div>
                      <div class='lesson-meta'>{les['day_of_week'] or '-'} {les['start_time'] or '-'}~{les['end_time'] or '-'} {status_html}</div>
                      <div class='student-line'>{students_label or '-'}</div>
                      <div class='lesson-main-actions'>
                        <a class='btn secondary' href='/attendance?lang={CURRENT_LANG}&lesson_mode=1&schedule_id={les['id']}&class_id={les['class_id']}&lesson_date={selected_view_date}&teacher_id={les['effective_teacher_id'] or ''}'>{t('academics.action.attendance_eval')}</a>
                        <a class='mini-link' href='/homework?lang={CURRENT_LANG}&selected_class_id={les['class_id']}'>{t('academics.go_homework')}</a>
                        <a class='mini-link' href='/exams?lang={CURRENT_LANG}&selected_class_id={les['class_id']}'>{t('academics.go_exams')}</a>
                        <a class='mini-link' href='/classes/{les['class_id']}?lang={CURRENT_LANG}'>{t('academics.view_class')}</a>
                        <a class='mini-link' href='/schedule?lang={CURRENT_LANG}&schedule_id={les['id']}&week={week_offset}&ref_date={ref_date_str}'>{t('common.edit')}</a>
                      </div>
                    </div>
                    """
                timetable_cells += f"<div class='tt-cell'>{blocks}</div>"

        week_label = f"{week_year}-{week_start.month:02d} W{week_no}"
        week_range_label = f"{week_start.isoformat()} ~ {week_end.isoformat()}"
        print_button_label = {"ko": "출력", "en": "Print", "zh": "打印"}.get(CURRENT_LANG, "Print")
        print_title_label = {"ko": "주간 시간표", "en": "Weekly Timetable", "zh": "周课表"}.get(CURRENT_LANG, "Weekly Timetable")
        print_day_label = t("academics.selected_day")
        print_date_label = t("academics.selected_date")
        print_week_label = t("academics.week_label")
        teacher_rows = list_teacher_profiles(conn)
        selected_teacher_options = [f"<option value=''>{t('academics.day_all')}</option>"]
        for tr in teacher_rows:
            selected = "selected" if str(tr["id"]) == selected_teacher_id else ""
            selected_teacher_options.append(f"<option value='{tr['id']}' {selected}>{tr['name']}</option>")

        day_options = []
        for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
            sel = "selected" if d == selected_day else ""
            day_options.append(f"<option value='{d}' {sel}>{d}</option>")

        class_query_enabled = query.get("form_class_load", "") == "1" or bool((query.get("form_class_q", "") or "").strip()) or bool(selected_form_class_id)
        class_candidates = fetch_class_candidates(conn, query.get("form_class_q", ""), limit=20, show_all_when_empty=class_query_enabled)
        if selected_schedule and not selected_form_class_id:
            selected_form_class_id = str(selected_schedule['class_id'])
            class_query_enabled = True
        selected_form_class = conn.execute(
            """SELECT c.id, c.name, c.teacher_id, c.foreign_teacher_id, c.chinese_teacher_id,
            co.name AS course_name, l.name AS level_name,
            uf.name AS foreign_teacher_name, uc.name AS chinese_teacher_name,
            (SELECT COUNT(*) FROM students s WHERE s.current_class_id=c.id) AS student_count,
            (SELECT GROUP_CONCAT(name_ko, ', ') FROM students s2 WHERE s2.current_class_id=c.id ORDER BY s2.id) AS student_names
            FROM classes c
            LEFT JOIN courses co ON co.id=c.course_id
            LEFT JOIN levels l ON l.id=c.level_id
            LEFT JOIN users uf ON uf.id=COALESCE(c.foreign_teacher_id, c.teacher_id)
            LEFT JOIN users uc ON uc.id=c.chinese_teacher_id
            WHERE c.id=?""",
            (selected_form_class_id,),
        ).fetchone() if selected_form_class_id else None

        selected_schedule_day = selected_schedule['day_of_week'] if selected_schedule else ''
        selected_schedule_status = selected_schedule['status'] if selected_schedule else 'active'
        selected_teacher_form = str((selected_schedule['teacher_id'] if selected_schedule and selected_schedule['teacher_id'] else ((selected_form_class['foreign_teacher_id'] or selected_form_class['teacher_id']) if selected_form_class else '')) or '')
        teacher_select_options = ["<option value=''>-</option>"]
        for tr in teacher_rows:
            sel = "selected" if str(tr['id']) == selected_teacher_form else ""
            teacher_select_options.append(f"<option value='{tr['id']}' {sel}>{tr['name']}</option>")

        selected_room_form = (selected_schedule['classroom'] if selected_schedule else '') or ''
        master_rooms = [r['name'] for r in conn.execute("SELECT name FROM classrooms ORDER BY name").fetchall()]
        existing_rooms = sorted(set(master_rooms) | {(r['classroom'] or '').strip() for r in schedules if (r['classroom'] or '').strip()} | ({selected_room_form} if selected_room_form else set()))
        room_options = ["<option value=''>-</option>"]
        for r in existing_rooms:
            sel = "selected" if r == selected_room_form else ""
            room_options.append(f"<option value='{r}' {sel}>{r}</option>")
        day_select_options = ["<option value=''>-</option>"]
        for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
            sel = "selected" if selected_schedule_day == d else ""
            day_select_options.append(f"<option value='{d}' {sel}>{d}</option>")
        slot_rows = conn.execute("SELECT label, start_time, end_time FROM time_slots ORDER BY start_time, end_time").fetchall()
        slot_pairs = [(r['start_time'], r['end_time']) for r in slot_rows if r['start_time'] and r['end_time']]
        if not slot_pairs:
            slot_pairs = [("16:25", "17:20"), ("17:25", "18:20"), ("18:30", "19:25"), ("19:35", "20:30")]
        selected_start = selected_schedule['start_time'] if selected_schedule else ''
        selected_end = selected_schedule['end_time'] if selected_schedule else ''
        selected_slot = f"{selected_start}|{selected_end}" if selected_start and selected_end else ""
        slot_options = ["<option value=''>-</option>"]
        for st, et in slot_pairs:
            val = f"{st}|{et}"
            sel = "selected" if val == selected_slot else ""
            slot_options.append(f"<option value='{val}' {sel}>{st}-{et}</option>")

        form_class_picker = render_picker_block(
            t("picker.class"),
            "form_class_q",
            query.get("form_class_q", ""),
            "selected_form_class_id",
            selected_form_class_id,
            (selected_form_class['name'] if selected_form_class else ""),
            class_candidates,
            "/schedule",
            CURRENT_LANG,
            {
                "week": str(week_offset), "ref_date": ref_date_str, "day": selected_day, "teacher_id": selected_teacher_id,
                "classroom": selected_room,
                "schedule_id": selected_schedule_id, "recent_class_ids": recent_class_ids,
            },
            query_flag_name="form_class_load",
            query_enabled=class_query_enabled,
        )

        room_param = quote(selected_room) if selected_room else ""
        print_query = f"/schedule?lang={CURRENT_LANG}&print=1&week={week_offset}&ref_date={ref_date_str}&day={selected_day}&teacher_id={selected_teacher_id}&classroom={room_param}"
        schedule_query = f"/schedule?lang={CURRENT_LANG}&week={week_offset}&ref_date={ref_date_str}&day={selected_day}&teacher_id={selected_teacher_id}&classroom={room_param}"

        detail_html = f"<div class='card'><h4>{t('academics.lesson_detail')}</h4><p class='empty-msg'>{t('common.no_data')}</p></div>"
        register_forms = f"<div class='card'><a class='btn secondary' href='/masterdata?lang={CURRENT_LANG}'>{t('academics.go_structure')}</a></div>"
        if selected_schedule:
            stu_rows = conn.execute("SELECT name_ko FROM students WHERE current_class_id=? ORDER BY id LIMIT 20", (selected_schedule['class_id'],)).fetchall()
            stu_text = ", ".join([r['name_ko'] for r in stu_rows]) if stu_rows else "-"
            detail_html = f"""
            <div class='card'>
              <h4>{t('academics.lesson_detail')}</h4>
              <table>
                <tr><th>{t('academics.class_name')}</th><td>{selected_schedule['class_name'] or '-'}</td></tr>
                <tr><th>{t('academics.course_level')}</th><td>{selected_schedule['course_name'] or '-'} / {selected_schedule['level_name'] or '-'}</td></tr>
                <tr><th>{t('academics.foreign_teacher')}</th><td>{selected_schedule['foreign_teacher_name'] or '-'}</td></tr>
                <tr><th>{t('academics.chinese_teacher')}</th><td>{selected_schedule['chinese_teacher_name'] or '-'}</td></tr>
                <tr><th>{t('academics.time_slot')}</th><td>{selected_schedule['day_of_week'] or '-'} {selected_schedule['start_time'] or '-'}~{selected_schedule['end_time'] or '-'}</td></tr>
                <tr><th>{t('academics.classroom')}</th><td>{selected_schedule['classroom'] or '-'}</td></tr>
                <tr><th>{t('academics.status')}</th><td><span class='badge {selected_schedule['status'] or ''}'>{status_t(selected_schedule['status']) if selected_schedule['status'] else '-'}</span></td></tr>
                <tr><th>{t('field.note')}</th><td>{selected_schedule['note'] or '-'}</td></tr>
                <tr><th>{t('academics.students')}</th><td>{stu_text}</td></tr>
              </table>
              <div class='lesson-actions' style='margin-top:10px'>
                <a class='btn' href='/classes/{selected_schedule['class_id']}?lang={CURRENT_LANG}'>{t('academics.view_class')}</a>
                <a class='btn' href='/attendance?lang={CURRENT_LANG}&lesson_mode=1&schedule_id={selected_schedule['id']}&class_id={selected_schedule['class_id']}&lesson_date={selected_view_date}&teacher_id={selected_schedule['effective_teacher_id'] or ''}'>{t('academics.action.attendance_eval')}</a>
                <a class='btn' href='/homework?lang={CURRENT_LANG}&selected_class_id={selected_schedule['class_id']}'>{t('academics.go_homework')}</a>
                <a class='btn' href='/exams?lang={CURRENT_LANG}&selected_class_id={selected_schedule['class_id']}'>{t('academics.go_exams')}</a>
              </div>
            </div>
            """

        if print_mode:
            html = render_html(t('academics.timetable_title'), f"""
            <div class='card filter-print-hide'>
              <div class='btn-row'>
                <a class='btn secondary' href='{schedule_query}'>Back</a>
                <button type='button' onclick='window.print()'>{print_button_label}</button>
              </div>
            </div>
            <div class='card schedule-print-card'>
              <div class='schedule-print-head'>
                <div class='schedule-print-title'>{print_title_label}</div>
                <div class='schedule-print-meta'>
                  <span>{print_week_label}: {week_label}</span>
                  <span>{print_day_label}: {selected_day}</span>
                  <span>{print_date_label}: {selected_view_date}</span>
                  <span>{t('academics.week_range')}: {week_range_label}</span>
                </div>
              </div>
              <div class='print-room-grid'>
                {print_room_sections or f"<div class='empty-msg'>{t('common.no_data')}</div>"}
              </div>
            </div>
            """, user, current_menu="schedule", flash_msg=flash_msg, flash_type=flash_type)
        else:
            html = render_html(t('academics.timetable_title'), f"""
        <div class='card'>
          <div class='muted'>{t('academics.timetable_desc')}</div>
        </div>
        <div class='card filter-print-hide'>
          <h4>{t('academics.filter')}</h4>
          <form method='get' class='filter-row'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='week' value='{week_offset}'>
            <input type='hidden' name='ref_date' value='{ref_date_str}'>
            <input type='hidden' name='recent_class_ids' value='{recent_class_ids}'>
            <div class='filter-grid'>
            <label>{t('academics.week_label')} <strong>{week_label}</strong></label>
            <label>{t('academics.week_range')} <strong>{week_range_label}</strong></label>
            <label>{t('academics.selected_day')} <strong>{selected_day}</strong></label>
            <label>{t('academics.selected_date')} <strong>{selected_view_date}</strong></label>
            <label>{t('academics.day_filter')} <select name='day' onchange='this.form.submit()'>{''.join(day_options)}</select></label>
            <label>{t('academics.teacher_filter')} <select name='teacher_id'>{''.join(selected_teacher_options)}</select></label>
            <label>{t('academics.classroom_filter')} <input name='classroom' value='{selected_room}'></label>
          </div>
          <div class='btn-row'>
            <a class='btn secondary' href='/schedule?lang={CURRENT_LANG}&ref_date={ref_date_str}&week={week_offset-1}&day={selected_day}'>{t('academics.week_prev')}</a>
            <a class='btn secondary' href='/schedule?lang={CURRENT_LANG}&ref_date={datetime.utcnow().date().isoformat()}&week=0&day={selected_day}'>{t('academics.week_current')}</a>
            <a class='btn secondary' href='/schedule?lang={CURRENT_LANG}&ref_date={ref_date_str}&week={week_offset+1}&day={selected_day}'>{t('academics.week_next')}</a>
            <button>{t('academics.search')}</button>
            <a class='btn secondary' href='/schedule?lang={CURRENT_LANG}'>{t('common.reset')}</a>
            <a class='btn' href='#schedule-form'>{t('academics.add_lesson')}</a>
          </div>
          </form>
          <form method='post' class='mobile-stack' style='margin-top:10px'>
            <input type='hidden' name='type' value='copy_week'>
            <div class='muted'>{t('academics.copy_week.desc')}</div>
            <div>{t('academics.copy_week.source')}: <strong>{week_range_label}</strong></div>
            <div>{t('academics.copy_week.target')}: <strong>{(week_start + timedelta(days=7)).isoformat()} ~ {(week_end + timedelta(days=7)).isoformat()}</strong></div>
            <button type='submit'>{t('academics.copy_week.button')}</button>
          </form>
        </div>

        <div class='card schedule-print-card'>
          <div class='card-header-row'>
            <h4>{t('academics.timetable')}</h4>
            <a class='btn secondary screen-only' href='{print_query}' target='_blank'>{print_button_label}</a>
          </div>
          <div class='schedule-print-head print-only'>
            <div class='schedule-print-title'>{print_title_label}</div>
            <div class='schedule-print-meta'>
              <span>{print_week_label}: {week_label}</span>
              <span>{print_day_label}: {selected_day}</span>
              <span>{print_date_label}: {selected_view_date}</span>
              <span>{t('academics.week_range')}: {week_range_label}</span>
            </div>
          </div>
          <div class='table-wrap timetable-wrap'>
            <div class='timetable-grid' style='grid-template-columns: var(--schedule-row-width, 150px) repeat({len(time_slots)}, minmax(var(--schedule-col-min, 170px),1fr));'>
              {''.join(timetable_cols)}
              {timetable_cells}
            </div>
          </div>
        </div>

        <div class='schedule-editor-grid'>
          <div>
            <div class='card' id='schedule-form'>
              <h4>{t('academics.schedule_form')}</h4>
              {form_class_picker}
              <form method='post' class='form-row'>
                <input type='hidden' name='type' value='schedule'>
                <input type='hidden' name='schedule_id' value='{selected_schedule['id'] if selected_schedule else ''}'>
                <input type='hidden' name='selected_form_class_id' value='{selected_form_class_id}'>
                <label>{t('academics.schedule_pick_class')} <input value='{selected_form_class['name'] if selected_form_class else '-'}' readonly></label>
                <label>{t('academics.course')} <input value='{selected_form_class['course_name'] if selected_form_class else '-'}' readonly></label>
                <label>{t('academics.level')} <input value='{selected_form_class['level_name'] if selected_form_class else '-'}' readonly></label>
                <label>{t('academics.foreign_teacher')} <select name='teacher_id'>{''.join(teacher_select_options)}</select></label>
                <label>{t('academics.chinese_teacher')} <input value='{selected_form_class['chinese_teacher_name'] if selected_form_class else '-'}' readonly></label>
                <label>{t('academics.schedule_pick_student')} <input value='{(selected_form_class['student_names'] if selected_form_class and selected_form_class['student_names'] else '-')}' readonly></label>
                <label>{t('academics.day_of_week')} <select name='day_of_week'>{''.join(day_select_options)}</select></label>
                <label>{t('academics.time_slot')} <select name='time_slot'>{''.join(slot_options)}</select></label>
                <label>{t('academics.classroom')} <select name='classroom'>{''.join(room_options)}</select></label>
                <label>{t('academics.status')} <select name='status'>
                  <option value='active' {'selected' if selected_schedule_status=='active' else ''}>{status_t('active')}</option>
                  <option value='makeup' {'selected' if selected_schedule_status=='makeup' else ''}>{status_t('makeup')}</option>
                  <option value='cancelled' {'selected' if selected_schedule_status=='cancelled' else ''}>{status_t('cancelled')}</option>
                </select></label>
                <label>{t('field.note')} <input name='note' value='{selected_schedule['note'] if selected_schedule else ''}'></label>
                <button>{t('common.save')}</button>
              </form>
              <div class='muted'>{t('academics.schedule_autofill')} / {t('academics.schedule_teacher_auto')}</div>
            </div>
          </div>
          <div>
            {detail_html}
            {register_forms}
            <details class='card' open>
              <summary><strong>{t('academics.class_list')}</strong></summary>
              <div style='margin-top:8px'>
              <h4>{t('academics.class_list')}</h4>
              <div class='table-wrap'><table>
                <tr><th>{t('academics.class_name')}</th><th>{t('academics.course')}</th><th>{t('academics.level')}</th><th>{t('academics.foreign_teacher')}</th><th>{t('academics.chinese_teacher')}</th><th>{t('academics.students')}</th></tr>
                {class_rows}
              </table></div>
              </div>
            </details>
          </div>
        </div>
        """, user, current_menu="schedule", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    # 출결
    if path == "/attendance":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]

        lesson_mode = query.get("lesson_mode", "") == "1"
        selected_student_id = query.get("selected_student_id", "")
        selected_class_id = query.get("selected_class_id", "")
        selected_teacher_id = query.get("selected_teacher_id", "")
        schedule_id = query.get("schedule_id", "")
        class_id_for_lesson = query.get("class_id", "") or selected_class_id
        lesson_date = (query.get("lesson_date", "") or datetime.utcnow().date().isoformat()).strip()
        flash_msg = ""
        flash_type = "success"

        # 운영 동선: 시간표 -> 출결 및 평가(일괄 입력)
        if lesson_mode:
            if not class_id_for_lesson or not str(class_id_for_lesson).isdigit():
                conn.close()
                html = render_html(t("attendance.title"), f"<div class='card'><div class='flash error'>{t('lesson.record.invalid_class')}</div></div>", user, current_menu="attendance")
                status, headers, body = text_resp(html)
                start_response(status, headers)
                return [body]
            if not is_valid_date(lesson_date):
                lesson_date = datetime.utcnow().date().isoformat()

            class_info = conn.execute(
                """SELECT c.id, c.name, c.teacher_id, COALESCE(c.credit_unit, 1) AS credit_unit, co.name AS course_name, l.name AS level_name, u.name AS teacher_name
                FROM classes c
                LEFT JOIN courses co ON co.id=c.course_id
                LEFT JOIN levels l ON l.id=c.level_id
                LEFT JOIN users u ON u.id=c.teacher_id
                WHERE c.id=?""",
                (class_id_for_lesson,),
            ).fetchone()
            if not class_info:
                conn.close()
                html = render_html(t("attendance.title"), f"<div class='card'><div class='flash error'>{t('lesson.record.not_found_class')}</div></div>", user, current_menu="attendance")
                status, headers, body = text_resp(html)
                start_response(status, headers)
                return [body]

            if has_role(user, [ROLE_TEACHER]) and str(class_info["teacher_id"] or "") != str(user["id"]):
                conn.close()
                status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                start_response(status, headers)
                return [body]

            schedule_info = None
            if str(schedule_id).isdigit():
                schedule_info = conn.execute(
                    "SELECT id, day_of_week, start_time, end_time, classroom, teacher_id FROM schedules WHERE id=? AND class_id=?",
                    (schedule_id, class_id_for_lesson),
                ).fetchone()

            students_in_class = conn.execute(
                "SELECT id, user_id, student_no, name_ko FROM students WHERE current_class_id=? ORDER BY name_ko, id",
                (class_id_for_lesson,),
            ).fetchall()

            score_fields = ["participation_score", "fluency_score", "vocabulary_score", "reading_score", "homework_score", "attitude_score"]
            score_labels = {
                "participation_score": t("lesson.score.participation"),
                "fluency_score": t("lesson.score.fluency"),
                "vocabulary_score": t("lesson.score.vocabulary"),
                "reading_score": t("lesson.score.reading"),
                "homework_score": t("lesson.score.homework"),
                "attitude_score": t("lesson.score.attitude"),
            }

            target_schedule_id = schedule_info["id"] if schedule_info else (schedule_id if str(schedule_id).isdigit() else None)

            if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
                d = parse_body(environ)
                post_lesson_date = (d.get("lesson_date") or lesson_date).strip()
                errs = []
                if not is_valid_date(post_lesson_date):
                    add_error(errs, "lesson_date", t("validation.date"))
                if schedule_info:
                    expected_weekday = schedule_info["day_of_week"] or ""
                    if expected_weekday and datetime.strptime(post_lesson_date, "%Y-%m-%d").strftime("%a") != expected_weekday:
                        add_error(errs, "lesson_date", t("lesson.record.weekday_mismatch"))
                if not students_in_class:
                    add_error(errs, "students", t("lesson.record.no_students"))

                row_errors = []
                parsed = []
                for st in students_in_class:
                    sid = str(st["user_id"])
                    status_v = (d.get(f"status_{sid}") or "present").strip()
                    if status_v not in ("present", "late", "absent", "makeup"):
                        row_errors.append(f"{st['name_ko']}: {t('validation.status_invalid')}")
                    absence_charge_type = (d.get(f"absence_charge_type_{sid}") or "").strip()
                    requires_makeup = parse_bool_flag(d.get(f"requires_makeup_{sid}"))
                    if requires_makeup is None:
                        row_errors.append(f"{st['name_ko']}: {t('attendance.validation.bool_invalid')}")
                        requires_makeup = 0
                    if status_v == "absent":
                        if absence_charge_type not in ("deduct", "no_deduct"):
                            row_errors.append(f"{st['name_ko']}: {t('attendance.validation.charge_required')}")
                    else:
                        absence_charge_type = ""
                        requires_makeup = 0
                    row = {
                        "student_user_id": st["user_id"],
                        "status": status_v,
                        "teacher_memo": (d.get(f"teacher_memo_{sid}") or "").strip(),
                        "absence_charge_type": absence_charge_type,
                        "requires_makeup": requires_makeup,
                        "makeup_completed": 0,
                    }
                    for sf in score_fields:
                        raw = (d.get(f"{sf}_{sid}") or "").strip()
                        if raw == "":
                            row[sf] = None
                        else:
                            v = as_int(raw)
                            if v is None or v < 1 or v > 5:
                                row_errors.append(f"{st['name_ko']}: {sf} {t('lesson.record.score_range')}")
                            row[sf] = v
                    parsed.append(row)

                if errs or row_errors:
                    flash_msg = format_errors(errs + [f"- row: {x}" for x in row_errors])
                    flash_type = "error"
                    log_event(conn, "ERROR", path, "수업기록 저장 검증 실패", "\n".join(errs + row_errors), user["id"])
                else:
                    class_credit_unit = get_class_credit_unit(conn, class_id_for_lesson)
                    for row in parsed:
                        row_credit_delta = calc_credit_delta(row["status"], row["absence_charge_type"], class_credit_unit)
                        if target_schedule_id is not None:
                            existing = conn.execute(
                                "SELECT id, credit_delta FROM attendance WHERE schedule_id=? AND student_id=? AND lesson_date=?",
                                (target_schedule_id, row["student_user_id"], post_lesson_date),
                            ).fetchone()
                        else:
                            existing = conn.execute(
                                "SELECT id, credit_delta FROM attendance WHERE class_id=? AND schedule_id IS NULL AND student_id=? AND lesson_date=?",
                                (class_id_for_lesson, row["student_user_id"], post_lesson_date),
                            ).fetchone()

                        if existing:
                            keep_id = existing["id"]
                            conn.execute(
                                """UPDATE attendance SET status=?, note=?, created_by=?, schedule_id=?,
                                participation_score=?, fluency_score=?, vocabulary_score=?, reading_score=?, homework_score=?, attitude_score=?, teacher_memo=?,
                                absence_charge_type=?, requires_makeup=?, makeup_completed=?, credit_delta=?
                                WHERE id=?""",
                                (
                                    row["status"], row["teacher_memo"] or None, user["id"], target_schedule_id,
                                    row["participation_score"], row["fluency_score"], row["vocabulary_score"], row["reading_score"], row["homework_score"], row["attitude_score"], row["teacher_memo"] or None,
                                    (row["absence_charge_type"] or None), row["requires_makeup"], row["makeup_completed"], row_credit_delta,
                                    keep_id,
                                ),
                            )
                            apply_credit_adjustment(conn, row["student_user_id"], existing["credit_delta"], row_credit_delta)
                        else:
                            cur = conn.execute(
                                """INSERT INTO attendance(
                                class_id, student_id, lesson_date, status, note, created_by, created_at, schedule_id,
                                participation_score, fluency_score, vocabulary_score, reading_score, homework_score, attitude_score, teacher_memo,
                                absence_charge_type, requires_makeup, makeup_completed, credit_delta
                                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                (
                                    class_id_for_lesson, row["student_user_id"], post_lesson_date, row["status"], row["teacher_memo"] or None, user["id"], now(),
                                    target_schedule_id,
                                    row["participation_score"], row["fluency_score"], row["vocabulary_score"], row["reading_score"], row["homework_score"], row["attitude_score"], row["teacher_memo"] or None,
                                    (row["absence_charge_type"] or None), row["requires_makeup"], row["makeup_completed"], row_credit_delta,
                                ),
                            )
                            keep_id = cur.lastrowid
                            apply_credit_adjustment(conn, row["student_user_id"], 0.0, row_credit_delta)
                        if target_schedule_id is not None:
                            conn.execute("DELETE FROM attendance WHERE id<>? AND schedule_id=? AND student_id=? AND lesson_date=?", (keep_id, target_schedule_id, row["student_user_id"], post_lesson_date))
                        else:
                            conn.execute("DELETE FROM attendance WHERE id<>? AND class_id=? AND schedule_id IS NULL AND student_id=? AND lesson_date=?", (keep_id, class_id_for_lesson, row["student_user_id"], post_lesson_date))
                        if row["status"] == "absent":
                            conn.execute("INSERT INTO notifications(type, target_user_id, payload, created_at) VALUES(?,?,?,?)",
                                         ("absence", row["student_user_id"], json.dumps({"student_id": row["student_user_id"], "date": post_lesson_date}, ensure_ascii=False), now()))
                    conn.commit()
                    flash_msg = t("lesson.record.saved")
                    lesson_date = post_lesson_date

            existing_map = {}
            if target_schedule_id is not None:
                ex_rows = conn.execute(
                    "SELECT * FROM attendance WHERE schedule_id=? AND lesson_date=?",
                    (target_schedule_id, lesson_date),
                ).fetchall()
            else:
                ex_rows = conn.execute(
                    "SELECT * FROM attendance WHERE class_id=? AND schedule_id IS NULL AND lesson_date=?",
                    (class_id_for_lesson, lesson_date),
                ).fetchall()
            for r in ex_rows:
                existing_map[str(r["student_id"])] = r

            student_rows_html = ""
            for st in students_in_class:
                sid = str(st["user_id"])
                ex = existing_map.get(sid)
                def score_cell(field):
                    val = ex[field] if ex and field in ex.keys() else None
                    return f"<input class='score-input' name='{field}_{sid}' value='{h(val if val is not None else '')}'>"
                status_val = (ex["status"] if ex else "present")
                charge_val = (ex["absence_charge_type"] if ex and 'absence_charge_type' in ex.keys() and ex["absence_charge_type"] else 'deduct')
                requires_makeup_val = 1 if ex and 'requires_makeup' in ex.keys() and str(ex['requires_makeup']) in ('1','true','True') else 0
                student_rows_html += f"""
                <tr>
                  <td>{h(st['student_no'] or '-')}</td>
                  <td>{h(st['name_ko'] or '-')}</td>
                  <td><select name='status_{sid}'>
                    <option value='present' {'selected' if status_val=='present' else ''}>{attendance_status_t('present')}</option>
                    <option value='late' {'selected' if status_val=='late' else ''}>{attendance_status_t('late')}</option>
                    <option value='absent' {'selected' if status_val=='absent' else ''}>{attendance_status_t('absent')}</option>
                    <option value='makeup' {'selected' if status_val=='makeup' else ''}>{attendance_status_t('makeup')}</option>
                  </select></td>
                  <td><select name='absence_charge_type_{sid}'>
                    <option value='deduct' {'selected' if charge_val=='deduct' else ''}>{t('attendance.charge.deduct')}</option>
                    <option value='no_deduct' {'selected' if charge_val=='no_deduct' else ''}>{t('attendance.charge.no_deduct')}</option>
                  </select></td>
                  <td><select name='requires_makeup_{sid}'>
                    <option value='0' {'selected' if requires_makeup_val==0 else ''}>{t('common.no')}</option>
                    <option value='1' {'selected' if requires_makeup_val==1 else ''}>{t('common.yes')}</option>
                  </select></td>
                  <td>{score_cell('participation_score')}</td>
                  <td>{score_cell('fluency_score')}</td>
                  <td>{score_cell('vocabulary_score')}</td>
                  <td>{score_cell('reading_score')}</td>
                  <td>{score_cell('homework_score')}</td>
                  <td>{score_cell('attitude_score')}</td>
                  <td><input class='memo-input' name='teacher_memo_{sid}' value='{h((ex['teacher_memo'] if ex and 'teacher_memo' in ex.keys() else (ex['note'] if ex else '')) or '')}'></td>
                </tr>
                """

            lesson_info = f"{h(class_info['name'])} / {h(class_info['course_name'] or '-')} / {h(class_info['level_name'] or '-')} · {h(class_info['teacher_name'] or '-') }"
            time_info = f"{h(schedule_info['day_of_week'])} {h(schedule_info['start_time'])}~{h(schedule_info['end_time'])}" if schedule_info else "-"
            room_info = h(schedule_info['classroom']) if schedule_info and schedule_info['classroom'] else "-"

            empty_student_notice = f"<div class='flash error'>{t('lesson.record.empty_students')}</div>" if not students_in_class else ""
            html = render_html(t("lesson.record.page_title"), f"""
            <div class='card'>
              <h4>{t('lesson.record.header')}</h4>
              <div class='muted'>{t('lesson.record.desc')}</div>
            </div>
            <div class='card'>
              <h4>{t('lesson.record.class_info')}</h4>
              <div><strong>{t('academics.course_level')} / {t('academics.teacher')}:</strong> {lesson_info}</div>
              <div><strong>{t('field.date')}:</strong> {h(lesson_date)}</div>
              <div><strong>{t('academics.time_slot')}:</strong> {time_info}</div>
              <div><strong>{t('academics.classroom')}:</strong> {room_info}</div>
              <div class='btn-row' style='margin-top:8px'><a class='btn secondary' href='/schedule?lang={CURRENT_LANG}'>{t('lesson.record.back_schedule')}</a></div>
            </div>
            <div class='card'>
              <h4>{t('lesson.record.input_title')}</h4>
              <form method='post'>
                <input type='hidden' name='lesson_date' value='{h(lesson_date)}'>
                {empty_student_notice}
                <div class='table-wrap'><table class='sticky-head'>
                  <thead><tr>
                    <th>{t('students.field.student_no')}</th><th>{t('field.name')}</th><th>{t('field.status')}</th><th>{t('attendance.absence_charge_type')}</th><th>{t('attendance.requires_makeup')}</th>
                    <th>{score_labels['participation_score']}</th><th>{score_labels['fluency_score']}</th><th>{score_labels['vocabulary_score']}</th><th>{score_labels['reading_score']}</th><th>{score_labels['homework_score']}</th><th>{score_labels['attitude_score']}</th><th>{t('lesson.score.teacher_memo')}</th>
                  </tr></thead>
                  <tbody>{student_rows_html or f"<tr><td colspan='12' class='empty-msg'>{t('common.no_data')}</td></tr>"}</tbody>
                </table></div>
                <div class='btn-row' style='margin-top:10px'><button>{t('common.save')}</button></div>
              </form>
            </div>
            """, user, current_menu="attendance", flash_msg=flash_msg, flash_type=flash_type)
            status, headers, body = text_resp(html)
            conn.close()
            start_response(status, headers)
            return [body]

        # 출결 관리/정정 화면
        selected_student_id = (query.get("selected_student_id", "") or "").strip()
        selected_class_id = (query.get("selected_class_id", "") or "").strip()
        status_filter = (query.get("status", "") or "").strip()
        deductible_filter = (query.get("deductible", "") or "").strip()
        requires_makeup_filter = (query.get("requires_makeup", "") or "").strip()
        makeup_completed_filter = (query.get("makeup_completed", "") or "").strip()
        date_from = (query.get("date_from", "") or "").strip()
        date_to = (query.get("date_to", "") or "").strip()
        makeup_needed_only = query.get("makeup_needed", "") == "1"

        student_candidates = fetch_student_candidates(conn, query.get("student_q", ""), limit=10)
        class_candidates = fetch_class_candidates(conn, query.get("class_q", ""), limit=10, show_all_when_empty=True)
        selected_student = conn.execute("SELECT id, name_ko, student_no FROM students WHERE id=?", (selected_student_id,)).fetchone() if selected_student_id else None
        selected_class = conn.execute("SELECT id, name FROM classes WHERE id=?", (selected_class_id,)).fetchone() if selected_class_id else None

        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            typ = (d.get("type") or "manual").strip()
            errs = []
            if typ == "update_attendance":
                attendance_id = (d.get("attendance_id") or "").strip()
                if not attendance_id.isdigit():
                    add_error(errs, "attendance_id", "Invalid attendance id")
                row = conn.execute(
                    """SELECT a.id, a.student_id, a.credit_delta, a.class_id, c.teacher_id
                    FROM attendance a LEFT JOIN classes c ON c.id=a.class_id WHERE a.id=?""",
                    (attendance_id,),
                ).fetchone() if attendance_id.isdigit() else None
                if not row:
                    add_error(errs, "attendance_id", "Attendance not found")
                else:
                    if has_role(user, [ROLE_TEACHER]) and str(row["teacher_id"] or "") != str(user["id"]):
                        conn.close()
                        status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                        start_response(status, headers)
                        return [body]
                status_v = (d.get("status") or "").strip()
                if status_v not in ("present", "late", "absent", "makeup"):
                    add_error(errs, "status", t("validation.status_invalid"))
                charge_type = (d.get("absence_charge_type") or "").strip()
                if status_v == "absent":
                    if charge_type not in ("deduct", "no_deduct"):
                        add_error(errs, "absence_charge_type", t("attendance.validation.charge_required"))
                else:
                    charge_type = ""
                requires_makeup = parse_bool_flag(d.get("requires_makeup"))
                makeup_completed = parse_bool_flag(d.get("makeup_completed"))
                if requires_makeup is None or makeup_completed is None:
                    add_error(errs, "makeup", t("attendance.validation.bool_invalid"))
                if status_v != "absent":
                    requires_makeup = 0
                if requires_makeup == 0:
                    makeup_completed = 0
                note_v = (d.get("note") or "").strip()
                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                else:
                    unit = get_class_credit_unit(conn, row["class_id"])
                    new_delta = calc_credit_delta(status_v, charge_type, unit)
                    conn.execute(
                        """UPDATE attendance SET status=?, note=?, absence_charge_type=?, requires_makeup=?, makeup_completed=?, credit_delta=? WHERE id=?""",
                        (status_v, note_v or None, charge_type or None, requires_makeup, makeup_completed, new_delta, attendance_id),
                    )
                    apply_credit_adjustment(conn, row["student_id"], row["credit_delta"], new_delta)
                    conn.commit()
                    flash_msg = t("attendance.updated")
            elif typ == "mark_makeup_completed":
                attendance_id = (d.get("attendance_id") or "").strip()
                row = conn.execute(
                    """SELECT a.id, a.class_id, c.teacher_id FROM attendance a
                    LEFT JOIN classes c ON c.id=a.class_id WHERE a.id=?""",
                    (attendance_id,),
                ).fetchone() if attendance_id.isdigit() else None
                if not row:
                    flash_msg = "Invalid attendance id"
                    flash_type = "error"
                elif has_role(user, [ROLE_TEACHER]) and str(row["teacher_id"] or "") != str(user["id"]):
                    conn.close()
                    status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                    start_response(status, headers)
                    return [body]
                else:
                    conn.execute("UPDATE attendance SET makeup_completed=1 WHERE id=?", (attendance_id,))
                    conn.commit()
                    flash_msg = t("attendance.makeup_marked")
            else:
                # emergency manual entry remains supported
                class_id = d.get("class_id") or selected_class_id
                student_input_id = d.get("student_id") or selected_student_id
                student_row = conn.execute("SELECT user_id FROM students WHERE id=?", (student_input_id,)).fetchone() if str(student_input_id).isdigit() else None
                student_id = student_row["user_id"] if student_row else student_input_id
                if not ensure_exists(conn, "classes", class_id):
                    add_error(errs, "class_id", "Invalid class")
                if not ensure_exists(conn, "users", student_id, extra_where="role='student'"):
                    add_error(errs, "student_id", "Invalid student")
                status_v = (d.get("status") or "").strip()
                if status_v not in ("present", "late", "absent", "makeup"):
                    add_error(errs, "status", t("validation.status_invalid"))
                lesson_date_v = (d.get("lesson_date") or "").strip()
                if not is_valid_date(lesson_date_v):
                    add_error(errs, "lesson_date", t("validation.date"))
                charge_type = (d.get("absence_charge_type") or "").strip()
                if status_v == "absent" and charge_type not in ("deduct", "no_deduct"):
                    add_error(errs, "absence_charge_type", t("attendance.validation.charge_required"))
                if status_v != "absent":
                    charge_type = ""
                requires_makeup = parse_bool_flag(d.get("requires_makeup"))
                if requires_makeup is None:
                    add_error(errs, "requires_makeup", t("attendance.validation.bool_invalid"))
                    requires_makeup = 0
                if status_v != "absent":
                    requires_makeup = 0
                created_by = user["id"]
                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                else:
                    unit = get_class_credit_unit(conn, class_id)
                    credit_delta = calc_credit_delta(status_v, charge_type, unit)
                    cur = conn.execute(
                        """INSERT INTO attendance(class_id, student_id, lesson_date, status, note, created_by, created_at,
                        absence_charge_type, requires_makeup, makeup_completed, credit_delta)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (class_id, student_id, lesson_date_v, status_v, d.get("note"), created_by, now(), charge_type or None, requires_makeup, 0, credit_delta),
                    )
                    apply_credit_adjustment(conn, student_id, 0.0, credit_delta)
                    if status_v == "absent":
                        conn.execute("INSERT INTO notifications(type, target_user_id, payload, created_at) VALUES(?,?,?,?)",
                                     ("absence", student_id, json.dumps({"student_id": student_id, "date": lesson_date_v}, ensure_ascii=False), now()))
                    conn.commit()
                    flash_msg = t("common.save")

        where = []
        params = []
        if has_role(user, [ROLE_TEACHER]):
            where.append("COALESCE(c.foreign_teacher_id, c.teacher_id)=?")
            params.append(user["id"])
        elif has_role(user, [ROLE_STUDENT]):
            where.append("a.student_id=?")
            params.append(user["id"])
        elif has_role(user, [ROLE_PARENT]):
            where.append("s.guardian_name=?")
            params.append(user["name"])
        if selected_student_id and selected_student_id.isdigit():
            stu = conn.execute("SELECT user_id FROM students WHERE id=?", (selected_student_id,)).fetchone()
            if stu:
                where.append("a.student_id=?")
                params.append(stu["user_id"])
        if selected_class_id and selected_class_id.isdigit():
            where.append("a.class_id=?")
            params.append(selected_class_id)
        if status_filter in ("present", "late", "absent", "makeup"):
            where.append("a.status=?")
            params.append(status_filter)
        if deductible_filter in ("deduct", "no_deduct"):
            where.append("a.absence_charge_type=?")
            params.append(deductible_filter)
        if requires_makeup_filter in ("0", "1"):
            where.append("COALESCE(a.requires_makeup,0)=?")
            params.append(requires_makeup_filter)
        if makeup_completed_filter in ("0", "1"):
            where.append("COALESCE(a.makeup_completed,0)=?")
            params.append(makeup_completed_filter)
        if is_valid_date(date_from):
            where.append("a.lesson_date>=?")
            params.append(date_from)
        if is_valid_date(date_to):
            where.append("a.lesson_date<=?")
            params.append(date_to)
        if makeup_needed_only:
            where.append("COALESCE(a.requires_makeup,0)=1 AND COALESCE(a.makeup_completed,0)=0")

        base_sql = """SELECT a.*, c.name AS class_name, st.name_ko AS student_name, st.student_no,
        COALESCE(c.credit_unit,1) AS class_credit_unit
        FROM attendance a
        LEFT JOIN classes c ON c.id=a.class_id
        LEFT JOIN students st ON st.user_id=a.student_id
        LEFT JOIN students s ON s.user_id=a.student_id
        """
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        rows = conn.execute(base_sql + where_sql + " ORDER BY a.lesson_date DESC, a.id DESC LIMIT 500", tuple(params)).fetchall()

        if query.get("export") == "csv":
            lines = ["id,date,student_no,student_name,class,status,absence_charge_type,requires_makeup,makeup_completed,credit_delta,note"]
            for r in rows:
                safe_note = str(r["note"] or "").replace('"', "'")
                lines.append(
                    f'"{r["id"]}","{r["lesson_date"] or ""}","{r["student_no"] or ""}","{r["student_name"] or ""}","{r["class_name"] or ""}","{r["status"] or ""}","{r["absence_charge_type"] or ""}","{r["requires_makeup"] or 0}","{r["makeup_completed"] or 0}","{r["credit_delta"] or 0}","{safe_note}"'
                )
            conn.close()
            start_response("200 OK", [("Content-Type", "text/csv; charset=utf-8"), ("Content-Disposition", "attachment; filename=attendance.csv")])
            return ["\n".join(lines).encode("utf-8")]

        student_picker = render_picker_block(t("picker.student"), "student_q", query.get("student_q", ""), "selected_student_id", selected_student_id,
                                            (f"{selected_student['name_ko']} ({selected_student['student_no'] or '-'})" if selected_student else ""),
                                            student_candidates, "/attendance", CURRENT_LANG,
                                            {"selected_class_id": selected_class_id, "status": status_filter, "date_from": date_from, "date_to": date_to, "deductible": deductible_filter, "requires_makeup": requires_makeup_filter, "makeup_completed": makeup_completed_filter})
        class_picker = render_picker_block(t("picker.class"), "class_q", query.get("class_q", ""), "selected_class_id", selected_class_id,
                                          (selected_class["name"] if selected_class else ""),
                                          class_candidates, "/attendance", CURRENT_LANG,
                                          {"selected_student_id": selected_student_id, "status": status_filter, "date_from": date_from, "date_to": date_to, "deductible": deductible_filter, "requires_makeup": requires_makeup_filter, "makeup_completed": makeup_completed_filter})

        row_html = ""
        for r in rows:
            row_html += f"""
            <tr>
              <td>{r['lesson_date'] or '-'}</td>
              <td>{h(r['student_name'] or '-')}</td>
              <td>{h(r['class_name'] or '-')}</td>
              <td>{attendance_status_t(r['status']) if r['status'] else '-'}</td>
              <td>{t('attendance.charge.deduct') if r['absence_charge_type']=='deduct' else (t('attendance.charge.no_deduct') if r['absence_charge_type']=='no_deduct' else '-')}</td>
              <td>{t('common.yes') if str(r['requires_makeup']) in ('1','True','true') else t('common.no')}</td>
              <td>{t('common.yes') if str(r['makeup_completed']) in ('1','True','true') else t('common.no')}</td>
              <td>{r['credit_delta'] if r['credit_delta'] is not None else 0}</td>
              <td>{h(r['note'] or '-')}</td>
              <td>
                <form method='post' class='form-row'>
                  <input type='hidden' name='type' value='update_attendance'>
                  <input type='hidden' name='attendance_id' value='{r['id']}'>
                  <select name='status'>
                    <option value='present' {'selected' if r['status']=='present' else ''}>{attendance_status_t('present')}</option>
                    <option value='late' {'selected' if r['status']=='late' else ''}>{attendance_status_t('late')}</option>
                    <option value='absent' {'selected' if r['status']=='absent' else ''}>{attendance_status_t('absent')}</option>
                    <option value='makeup' {'selected' if r['status']=='makeup' else ''}>{attendance_status_t('makeup')}</option>
                  </select>
                  <select name='absence_charge_type'>
                    <option value='deduct' {'selected' if r['absence_charge_type']=='deduct' else ''}>{t('attendance.charge.deduct')}</option>
                    <option value='no_deduct' {'selected' if r['absence_charge_type']=='no_deduct' else ''}>{t('attendance.charge.no_deduct')}</option>
                  </select>
                  <select name='requires_makeup'>
                    <option value='0' {'selected' if str(r['requires_makeup']) not in ('1','True','true') else ''}>{t('common.no')}</option>
                    <option value='1' {'selected' if str(r['requires_makeup']) in ('1','True','true') else ''}>{t('common.yes')}</option>
                  </select>
                  <select name='makeup_completed'>
                    <option value='0' {'selected' if str(r['makeup_completed']) not in ('1','True','true') else ''}>{t('common.no')}</option>
                    <option value='1' {'selected' if str(r['makeup_completed']) in ('1','True','true') else ''}>{t('common.yes')}</option>
                  </select>
                  <input name='note' value='{h(r['note'] or '')}' style='min-width:120px'>
                  <button>{t('attendance.correction')}</button>
                </form>
                <form method='post' class='btn-row' style='margin-top:6px'>
                  <input type='hidden' name='type' value='mark_makeup_completed'>
                  <input type='hidden' name='attendance_id' value='{r['id']}'>
                  <button class='btn secondary'>{t('attendance.makeup_completed')}</button>
                </form>
              </td>
            </tr>
            """

        html = render_html(t("attendance.title"), f"""
        {student_picker}
        {class_picker}
        <div class='card'>
          <h4>{t('attendance.list')}</h4>
          <div class='muted'>{t('attendance.list_desc')}</div>
          <form method='get' class='mobile-stack' style='margin-top:8px'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='selected_student_id' value='{selected_student_id}'>
            <input type='hidden' name='selected_class_id' value='{selected_class_id}'>
            <div class='filter-grid'>
            <label>{t('attendance.filter.date_from')} <input type='date' name='date_from' value='{date_from}'></label>
            <label>{t('attendance.filter.date_to')} <input type='date' name='date_to' value='{date_to}'></label>
            <label>{t('field.status')} <select name='status'><option value=''>{t('academics.day_all')}</option><option value='present' {'selected' if status_filter=='present' else ''}>{attendance_status_t('present')}</option><option value='late' {'selected' if status_filter=='late' else ''}>{attendance_status_t('late')}</option><option value='absent' {'selected' if status_filter=='absent' else ''}>{attendance_status_t('absent')}</option><option value='makeup' {'selected' if status_filter=='makeup' else ''}>{attendance_status_t('makeup')}</option></select></label>
            <label>{t('attendance.filter.deductible')} <select name='deductible'><option value=''>{t('academics.day_all')}</option><option value='deduct' {'selected' if deductible_filter=='deduct' else ''}>{t('attendance.charge.deduct')}</option><option value='no_deduct' {'selected' if deductible_filter=='no_deduct' else ''}>{t('attendance.charge.no_deduct')}</option></select></label>
            <label>{t('attendance.requires_makeup')} <select name='requires_makeup'><option value=''>{t('academics.day_all')}</option><option value='1' {'selected' if requires_makeup_filter=='1' else ''}>{t('common.yes')}</option><option value='0' {'selected' if requires_makeup_filter=='0' else ''}>{t('common.no')}</option></select></label>
            <label>{t('attendance.makeup_completed')} <select name='makeup_completed'><option value=''>{t('academics.day_all')}</option><option value='1' {'selected' if makeup_completed_filter=='1' else ''}>{t('common.yes')}</option><option value='0' {'selected' if makeup_completed_filter=='0' else ''}>{t('common.no')}</option></select></label>
          </div>
          <div class='btn-row'>
            <button>{t('common.search')}</button>
            <a class='btn secondary' href='/attendance?lang={CURRENT_LANG}&makeup_needed=1'>{t('attendance.students_needing_makeup')}</a>
            <a class='btn secondary' href='/attendance?lang={CURRENT_LANG}&selected_student_id={selected_student_id}&selected_class_id={selected_class_id}&status={status_filter}&date_from={date_from}&date_to={date_to}&deductible={deductible_filter}&requires_makeup={requires_makeup_filter}&makeup_completed={makeup_completed_filter}&export=csv'>{t('attendance.export_csv')}</a>
          </div>
          </form>
          <div class='table-wrap' style='margin-top:10px'><table class='sticky-head'>
            <tr><th>{t('field.date')}</th><th>{t('field.student')}</th><th>{t('students.field.class')}</th><th>{t('field.status')}</th><th>{t('attendance.absence_charge_type')}</th><th>{t('attendance.requires_makeup')}</th><th>{t('attendance.makeup_completed')}</th><th>{t('attendance.credit_impact')}</th><th>{t('field.note')}</th><th>{t('common.edit')}</th></tr>
            {row_html or f"<tr><td colspan='10' class='empty-msg'>{t('common.no_data')}</td></tr>"}
          </table></div>
        </div>
        <div class='card'>
          <h4>{t('attendance.manual_input_title')}</h4>
          <form method='post' class='form-row'>
            <input type='hidden' name='type' value='manual'>
            <input type='hidden' name='student_id' value='{selected_student_id}'>
            <input type='hidden' name='class_id' value='{selected_class_id}'>
            <label>{t('field.student_id')} <input value='{selected_student_id}' readonly></label>
            <label>{t('field.class_id')} <input value='{selected_class_id}' readonly></label>
            <label>{t('field.date')} <input name='lesson_date' placeholder='2026-03-06'></label>
            <label>{t('field.status')} <select name='status'><option value='present'>{attendance_status_t('present')}</option><option value='late'>{attendance_status_t('late')}</option><option value='absent'>{attendance_status_t('absent')}</option><option value='makeup'>{attendance_status_t('makeup')}</option></select></label>
            <label>{t('attendance.absence_charge_type')} <select name='absence_charge_type'><option value='deduct'>{t('attendance.charge.deduct')}</option><option value='no_deduct'>{t('attendance.charge.no_deduct')}</option></select></label>
            <label>{t('attendance.requires_makeup')} <select name='requires_makeup'><option value='0'>{t('common.no')}</option><option value='1'>{t('common.yes')}</option></select></label>
            <label>{t('field.note')} <input name='note'></label>
            <button>{t("common.save")}</button>
          </form>
        </div>
        """, user, current_menu="attendance", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/homework":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]

        flash_msg = ""
        flash_type = "success"
        selected_class_id = (query.get("selected_class_id", "") or "").strip()
        selected_homework_id = (query.get("selected_homework_id", "") or "").strip()
        load_homework = query.get("load", "") == "1" or bool(selected_class_id) or bool(selected_homework_id)
        class_query_enabled = query.get("do_search", "") == "1" or bool((query.get("class_q", "") or "").strip()) or bool(selected_class_id)

        class_candidates = fetch_class_candidates(conn, query.get("class_q", ""), limit=20, show_all_when_empty=class_query_enabled)
        selected_class = conn.execute(
            """SELECT c.id, c.name, c.teacher_id, co.name AS course_name, l.name AS level_name, u.name AS teacher_name
            FROM classes c
            LEFT JOIN courses co ON co.id=c.course_id
            LEFT JOIN levels l ON l.id=c.level_id
            LEFT JOIN users u ON u.id=c.teacher_id
            WHERE c.id=?""",
            (selected_class_id,),
        ).fetchone() if selected_class_id and selected_class_id.isdigit() else None

        if has_role(user, [ROLE_TEACHER]) and selected_class and str(selected_class["teacher_id"] or "") != str(user["id"]):
            conn.close()
            status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
            start_response(status, headers)
            return [body]

        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            typ = d.get("type")
            errs = []
            if typ == "homework_create":
                class_id = (d.get("class_id") or selected_class_id or "").strip()
                title = (d.get("title") or "").strip()
                description = (d.get("description") or "").strip()
                due_date = (d.get("due_date") or "").strip()
                hw_status = (d.get("status") or "active").strip() or "active"
                teacher_id = (d.get("teacher_id") or user["id"])

                if not class_id:
                    add_error(errs, "class_id", t("homework.class_required"))
                elif not ensure_exists(conn, "classes", class_id):
                    add_error(errs, "class_id", "invalid class")
                if has_role(user, [ROLE_TEACHER]):
                    class_ok = conn.execute("SELECT id FROM classes WHERE id=? AND teacher_id=?", (class_id, user["id"])).fetchone()
                    if not class_ok:
                        conn.close()
                        status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                        start_response(status, headers)
                        return [body]
                if not title:
                    add_error(errs, "title", t("homework.validation.title_required"))
                if due_date and not is_valid_date(due_date):
                    add_error(errs, "due_date", t("homework.validation.due_date"))
                if hw_status not in ("active", "closed"):
                    add_error(errs, "status", "invalid status")

                if errs:
                    flash_msg = format_errors(errs)
                    flash_type = "error"
                else:
                    conn.execute(
                        """INSERT INTO homework(class_id, teacher_id, title, description, due_date, status, created_by, created_at, updated_at)
                        VALUES(?,?,?,?,?,?,?,?,?)""",
                        (class_id, teacher_id, title, description or None, due_date or None, hw_status, teacher_id, now(), now()),
                    )
                    homework_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
                    # safer/simpler choice: initialize all class students immediately
                    class_students = conn.execute("SELECT user_id FROM students WHERE current_class_id=? ORDER BY id", (class_id,)).fetchall()
                    if not class_students:
                        add_error(errs, "targets", t("homework.validation.no_targets"))
                    for st in class_students:
                        exists = conn.execute("SELECT id FROM homework_submissions WHERE homework_id=? AND student_id=?", (homework_id, st["user_id"])).fetchone()
                        if not exists:
                            conn.execute(
                                """INSERT INTO homework_submissions(homework_id, student_id, submitted, submitted_at, feedback, feedback_teacher_id, updated_at)
                                VALUES(?,?,?,?,?,?,?)""",
                                (homework_id, st["user_id"], 0, None, None, None, now()),
                            )
                    conn.execute("INSERT INTO notifications(type, target_user_id, payload, created_at) VALUES(?,?,?,?)", ("homework", None, json.dumps({"homework_id": homework_id, "title": title}, ensure_ascii=False), now()))
                    conn.commit()
                    flash_msg = t("homework.saved")
                    selected_homework_id = str(homework_id)
                    selected_class_id = str(class_id)
                    if errs:
                        flash_msg = f"{flash_msg} ({t('homework.validation.no_targets')})"

            elif typ == "submission_bulk":
                homework_id = (d.get("homework_id") or selected_homework_id or "").strip()
                if not homework_id or not homework_id.isdigit():
                    add_error(errs, "homework_id", t("homework.validation.homework_required"))
                hw_row = conn.execute("SELECT id, class_id, teacher_id FROM homework WHERE id=?", (homework_id,)).fetchone() if homework_id.isdigit() else None
                if not hw_row:
                    add_error(errs, "homework", "invalid homework")
                elif has_role(user, [ROLE_TEACHER]) and str(hw_row["teacher_id"] or "") != str(user["id"]):
                    conn.close()
                    status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                    start_response(status, headers)
                    return [body]

                if not errs:
                    class_students = conn.execute("SELECT user_id, student_no, name_ko FROM students WHERE current_class_id=? ORDER BY name_ko, id", (hw_row["class_id"],)).fetchall()
                    valid_student_ids = {str(st["user_id"]) for st in class_students}
                    for sid in valid_student_ids:
                        submitted = 1 if d.get(f"submitted_{sid}") else 0
                        feedback = (d.get(f"feedback_{sid}") or "").strip()
                        submitted_at = now() if submitted else None
                        existing = conn.execute("SELECT id FROM homework_submissions WHERE homework_id=? AND student_id=?", (homework_id, sid)).fetchone()
                        if existing:
                            conn.execute(
                                """UPDATE homework_submissions
                                SET submitted=?, submitted_at=COALESCE(?, submitted_at), feedback=?, feedback_teacher_id=?, updated_at=?
                                WHERE id=?""",
                                (submitted, submitted_at, feedback or None, user["id"], now(), existing["id"]),
                            )
                        else:
                            conn.execute(
                                """INSERT INTO homework_submissions(homework_id, student_id, submitted, submitted_at, feedback, feedback_teacher_id, updated_at)
                                VALUES(?,?,?,?,?,?,?)""",
                                (homework_id, sid, submitted, submitted_at, feedback or None, user["id"], now()),
                            )
                    conn.commit()
                    flash_msg = t("homework.rows_saved")
                    selected_homework_id = str(homework_id)
                    selected_class_id = str(hw_row["class_id"])
                else:
                    flash_msg = format_errors(errs)
                    flash_type = "error"

        if method == "POST":
            load_homework = True

        # access scoped homework list
        hw_params = []
        hw_where = []
        if selected_class_id and selected_class_id.isdigit():
            hw_where.append("h.class_id=?")
            hw_params.append(selected_class_id)
        if has_role(user, [ROLE_TEACHER]):
            hw_where.append("h.teacher_id=?")
            hw_params.append(user["id"])
        elif has_role(user, [ROLE_STUDENT]):
            hw_where.append("hs.student_id=?")
            hw_params.append(user["id"])
        elif has_role(user, [ROLE_PARENT]):
            hw_where.append("s.guardian_name=?")
            hw_params.append(user["name"])
        where_sql = ("WHERE " + " AND ".join(hw_where)) if hw_where else ""
        homework_rows = []
        if load_homework:
            homework_rows = conn.execute(
                f"""SELECT DISTINCT h.id, h.class_id, h.teacher_id, h.title, h.description, h.due_date, h.status, h.created_at,
                c.name AS class_name, u.name AS teacher_name,
                (SELECT COUNT(*) FROM students st WHERE st.current_class_id=h.class_id) AS target_count,
                (SELECT COUNT(*) FROM homework_submissions hs2 WHERE hs2.homework_id=h.id AND hs2.submitted=1) AS submitted_count
                FROM homework h
                LEFT JOIN classes c ON c.id=h.class_id
                LEFT JOIN users u ON u.id=h.teacher_id
                LEFT JOIN homework_submissions hs ON hs.homework_id=h.id
                LEFT JOIN students s ON s.user_id=hs.student_id
                {where_sql}
                ORDER BY h.id DESC LIMIT 200""",
                tuple(hw_params),
            ).fetchall()

        selected_homework = None
        if selected_homework_id.isdigit():
            selected_homework = conn.execute(
                """SELECT h.*, c.name AS class_name, u.name AS teacher_name
                FROM homework h
                LEFT JOIN classes c ON c.id=h.class_id
                LEFT JOIN users u ON u.id=h.teacher_id
                WHERE h.id=?""",
                (selected_homework_id,),
            ).fetchone()
            if selected_homework and has_role(user, [ROLE_TEACHER]) and str(selected_homework["teacher_id"] or "") != str(user["id"]):
                selected_homework = None

        submission_rows = []
        if selected_homework:
            class_students = conn.execute("SELECT user_id, student_no, name_ko FROM students WHERE current_class_id=? ORDER BY name_ko, id", (selected_homework["class_id"],)).fetchall()
            sub_map = {
                str(r["student_id"]): r
                for r in conn.execute("SELECT * FROM homework_submissions WHERE homework_id=?", (selected_homework["id"],)).fetchall()
            }
            for st in class_students:
                submission_rows.append((st, sub_map.get(str(st["user_id"]))))

        class_picker = render_picker_block(
            t("picker.class"), "class_q", query.get("class_q", ""), "selected_class_id", selected_class_id,
            (selected_class["name"] if selected_class else ""), class_candidates, "/homework", CURRENT_LANG,
            {"selected_homework_id": selected_homework_id, "load": "1"},
            query_enabled=class_query_enabled,
        )

        hw_list_html = ""
        for hrow in homework_rows:
            progress = f"{hrow['submitted_count'] or 0}/{hrow['target_count'] or 0}"
            hw_list_html += f"<tr><td><a href='/homework?lang={CURRENT_LANG}&selected_class_id={hrow['class_id']}&selected_homework_id={hrow['id']}'>{hrow['title']}</a></td><td>{hrow['class_name'] or '-'}</td><td>{hrow['due_date'] or '-'}</td><td>{hrow['teacher_name'] or '-'}</td><td>{progress}</td></tr>"

        sub_table_rows = ""
        for st, sub in submission_rows:
            submitted_checked = "checked" if sub and str(sub["submitted"]) == "1" else ""
            submitted_at = (sub["submitted_at"] if sub and sub["submitted_at"] else "-")
            feedback_val = h(sub["feedback"] if sub and sub["feedback"] else "")
            sub_table_rows += f"""
            <tr>
              <td>{h(st['student_no'] or '-')}</td><td>{h(st['name_ko'] or '-')}</td>
              <td><input type='checkbox' name='submitted_{st['user_id']}' value='1' {submitted_checked}></td>
              <td>{h(submitted_at)}</td>
              <td><input name='feedback_{st['user_id']}' value='{feedback_val}' style='min-width:240px'></td>
            </tr>
            """

        html = render_html(t("homework.title"), f"""
        {class_picker}
        <div class='card'>
          <h4>{t('homework.class_context')}</h4>
          <div>{t('academics.class_name')}: <strong>{selected_class['name'] if selected_class else '-'}</strong></div>
          <div>{t('homework.teacher')}: <strong>{selected_class['teacher_name'] if selected_class and selected_class['teacher_name'] else '-'}</strong></div>
          <div>{t('academics.course_level')}: <strong>{selected_class['course_name'] if selected_class and selected_class['course_name'] else '-'} / {selected_class['level_name'] if selected_class and selected_class['level_name'] else '-'}</strong></div>
        </div>
        <div class='card'>
          <h4>{t('homework.add')}</h4>
          <form method='post' class='form-row'>
            <input type='hidden' name='type' value='homework_create'>
            <input type='hidden' name='class_id' value='{selected_class_id}'>
            <label>{t('academics.class_name')} <input value='{selected_class['name'] if selected_class else '-'}' readonly></label>
            <label>{t('field.title')} <input name='title' required></label>
            <label>{t('homework.description')} <input name='description' style='min-width:220px'></label>
            <label>{t('field.due_date')} <input name='due_date' placeholder='2026-03-15'></label>
            <label>{t('homework.status')} <select name='status'><option value='active'>{t('homework.status.active')}</option><option value='closed'>{t('homework.status.closed')}</option></select></label>
            <button>{t('common.save')}</button>
          </form>
        </div>
        <div class='card'>
          <h4>{t('homework.list')}</h4>
          {'' if load_homework else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}
          <table>
            <tr><th>{t('field.title')}</th><th>{t('academics.class_name')}</th><th>{t('field.due_date')}</th><th>{t('field.writer')}</th><th>{t('homework.progress')}</th></tr>
            {hw_list_html if load_homework else ''}
            {(f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_homework and not hw_list_html) else ''}
          </table>
        </div>
        <div class='card'>
          <h4>{t('homework.submission_panel')}</h4>
          <div>{t('homework.select_homework')}: <strong>{selected_homework['title'] if selected_homework else '-'}</strong></div>
          <form method='post'>
            <input type='hidden' name='type' value='submission_bulk'>
            <input type='hidden' name='homework_id' value='{selected_homework['id'] if selected_homework else ''}'>
            <table>
              <tr><th>{t('students.field.student_no')}</th><th>{t('field.name')}</th><th>{t('students.field.submitted')}</th><th>{t('students.field.submitted_at')}</th><th>{t('students.field.feedback')}</th></tr>
              {sub_table_rows or f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>"}
            </table>
            <div style='margin-top:10px'><button>{t('homework.save_students')}</button></div>
          </form>
        </div>
        """, user, current_menu="homework", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/exams":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER, ROLE_PARENT, ROLE_STUDENT]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        flash_msg = ""
        flash_type = "success"
        q_exam_class_id = (query.get("selected_class_id", "") or "").strip()
        q_exam_student_id = (query.get("selected_student_id", "") or "").strip()
        q_exam_class_q = (query.get("class_q", "") or "").strip()
        q_exam_student_q = (query.get("student_q", "") or "").strip()
        exam_picker_query_enabled = query.get("do_search", "") == "1" or bool(q_exam_class_q) or bool(q_exam_student_q) or bool(q_exam_class_id) or bool(q_exam_student_id)
        class_candidates = fetch_class_candidates(conn, q_exam_class_q, limit=10) if exam_picker_query_enabled else []
        student_candidates = fetch_student_candidates(conn, q_exam_student_q, limit=10) if exam_picker_query_enabled else []
        selected_class = conn.execute("SELECT id, name FROM classes WHERE id=?", (q_exam_class_id,)).fetchone() if q_exam_class_id.isdigit() else None
        selected_student = conn.execute("SELECT id, user_id, student_no, name_ko FROM students WHERE id=?", (q_exam_student_id,)).fetchone() if q_exam_student_id.isdigit() else None
        selected_student_user_id = str(selected_student["user_id"]) if selected_student else q_exam_student_id
        load_exams = query.get("load", "") == "1" or bool(q_exam_class_id) or bool(selected_student_user_id)
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            typ = d.get("type")
            errs = []
            class_input_id = d.get("class_id") or q_exam_class_id
            score_exam_id = d.get("exam_id")
            if has_role(user, [ROLE_TEACHER]):
                if typ == "exam":
                    class_ok = conn.execute("SELECT id FROM classes WHERE id=? AND teacher_id=?", (class_input_id, user["id"])).fetchone()
                    if not class_ok:
                        conn.close()
                        status, headers, body = forbidden_html(user, t('forbidden.teacher_class_only'))
                        start_response(status, headers)
                        return [body]
                elif typ == "score":
                    exam_ok = conn.execute("""SELECT e.id FROM exams e JOIN classes c ON c.id=e.class_id WHERE e.id=? AND c.teacher_id=?""", (score_exam_id, user["id"])).fetchone()
                    if not exam_ok:
                        conn.close()
                        status, headers, body = forbidden_html(user, t('forbidden.teacher_exam_only'))
                        start_response(status, headers)
                        return [body]
            if typ == "exam":
                if not ensure_exists(conn, "classes", class_input_id):
                    add_error(errs, "class_id", "존재하지 않는 반입니다")
                if not (d.get("name") or "").strip():
                    add_error(errs, "name", "필수값입니다")
                if d.get("exam_date") and not is_valid_date(d.get("exam_date")):
                    add_error(errs, "exam_date", "YYYY-MM-DD 형식이어야 합니다")
                if d.get("linked_book_id") and not ensure_exists(conn, "books", d.get("linked_book_id")):
                    add_error(errs, "linked_book_id", "존재하지 않는 도서입니다")
                if not errs:
                    conn.execute("INSERT INTO exams(class_id, name, exam_date, report, linked_book_id, created_at) VALUES(?,?,?,?,?,?)", (class_input_id, d.get("name"), d.get("exam_date"), d.get("report"), d.get("linked_book_id") or None, now()))
            elif typ == "score":
                score_v = as_float(d.get("score"))
                if not ensure_exists(conn, "exams", d.get("exam_id")):
                    add_error(errs, "exam_id", "존재하지 않는 시험입니다")
                student_input_id = d.get("student_id") or selected_student_user_id
                if not ensure_exists(conn, "users", student_input_id, extra_where="role='student'"):
                    add_error(errs, "student_id", "존재하지 않는 학생입니다")
                if score_v is None or score_v < 0 or score_v > 100:
                    add_error(errs, "score", "0~100 범위의 숫자여야 합니다")
                if not errs:
                    conn.execute("INSERT INTO exam_scores(exam_id, student_id, score, created_at) VALUES(?,?,?,?)", (d.get("exam_id"), student_input_id, score_v, now()))
            if errs:
                flash_msg = format_errors(errs)
                flash_type = "error"
                log_event(conn, "ERROR", path, "시험/성적 저장 검증 실패", "\n".join(errs), user["id"])
            else:
                conn.commit()
                flash_msg = "저장되었습니다"
            load_exams = True

        exams = []
        scores = []
        if load_exams:
            if has_role(user, [ROLE_TEACHER]):
                exams = conn.execute("SELECT e.* FROM exams e JOIN classes c ON c.id=e.class_id WHERE c.teacher_id=? ORDER BY e.id DESC", (user["id"],)).fetchall()
                scores = conn.execute("SELECT es.* FROM exam_scores es JOIN exams e ON e.id=es.exam_id JOIN classes c ON c.id=e.class_id WHERE c.teacher_id=? ORDER BY es.id DESC", (user["id"],)).fetchall()
            elif has_role(user, [ROLE_STUDENT]):
                exams = conn.execute("SELECT e.* FROM exams e JOIN exam_scores es ON es.exam_id=e.id WHERE es.student_id=? ORDER BY e.id DESC", (user["id"],)).fetchall()
                scores = conn.execute("SELECT * FROM exam_scores WHERE student_id=? ORDER BY id DESC", (user["id"],)).fetchall()
            elif has_role(user, [ROLE_PARENT]):
                exams = conn.execute("SELECT DISTINCT e.* FROM exams e JOIN exam_scores es ON es.exam_id=e.id JOIN students s ON s.user_id=es.student_id WHERE s.guardian_name=? ORDER BY e.id DESC", (user["name"],)).fetchall()
                scores = conn.execute("SELECT es.* FROM exam_scores es JOIN students s ON s.user_id=es.student_id WHERE s.guardian_name=? ORDER BY es.id DESC", (user["name"],)).fetchall()
            else:
                exams = conn.execute("SELECT * FROM exams ORDER BY id DESC").fetchall()
                scores = conn.execute("SELECT * FROM exam_scores ORDER BY id DESC").fetchall()

            if q_exam_class_id.isdigit():
                exams = [r for r in exams if str(r["class_id"] or "") == q_exam_class_id]
                allowed_exam_ids = {str(r["id"]) for r in exams}
                scores = [r for r in scores if str(r["exam_id"] or "") in allowed_exam_ids]
            if str(selected_student_user_id).isdigit():
                scores = [r for r in scores if str(r["student_id"] or "") == str(selected_student_user_id)]
                allowed_exam_ids = {str(r["exam_id"]) for r in scores}
                exams = [r for r in exams if str(r["id"] or "") in allowed_exam_ids]

        class_picker = render_picker_block(t("picker.class"), "class_q", q_exam_class_q, "selected_class_id", q_exam_class_id,
                                         (selected_class["name"] if selected_class else ""), class_candidates, "/exams", CURRENT_LANG,
                                         {"selected_student_id": q_exam_student_id, "student_q": q_exam_student_q, "load": "1"}, query_enabled=exam_picker_query_enabled)
        student_picker = render_picker_block(t("picker.student"), "student_q", q_exam_student_q, "selected_student_id", q_exam_student_id,
                                           (f"{selected_student['name_ko']} ({selected_student['student_no'] or '-'} / U:{selected_student['user_id']})" if selected_student else ""), student_candidates, "/exams", CURRENT_LANG,
                                           {"selected_class_id": q_exam_class_id, "class_q": q_exam_class_q, "load": "1"}, query_enabled=exam_picker_query_enabled)
        exam_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['class_id']}</td><td>{r['name']}</td><td>{r['exam_date'] or '-'}</td><td>{r['report'] or '-'}</td></tr>" for r in exams])
        score_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['exam_id']}</td><td>{r['student_id']}</td><td>{r['score']}</td></tr>" for r in scores])
        html = render_html(t("exams.title"), f"""
        {class_picker if has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]) else ''}
        {student_picker if has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]) else ''}
        <div class='card'><h4>{t("exams.add")}</h4><form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'><input type='hidden' name='type' value='exam'>{t("field.class_id")}<input name='class_id' value='{h(q_exam_class_id)}' placeholder='ID'> {t("field.exam_name")}<input name='name'> {t("field.exam_date")}<input name='exam_date'> {t("field.report")}<input name='report'> {t("field.book_id")}<input name='linked_book_id'><button>{t('common.save')}</button></form></div>
        <div class='card'><h4>{t("exams.score_input")}</h4><form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'><input type='hidden' name='type' value='score'>{t("field.exam_id")}<input name='exam_id'> {t("field.student_id")}<input name='student_id' value='{h(selected_student_user_id)}' placeholder='User ID'> {t("field.score")}<input name='score'><button>{t('common.save')}</button></form></div>
        <div class='card'>
          <h4>{t('common.search')}</h4>
          <form method='get' class='mobile-stack query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='load' value='1'>
            <div class='filter-grid'>
              <label>{t('field.class_id')} <input name='selected_class_id' value='{h(q_exam_class_id)}'></label>
              <label>{t('field.student_id')} <input name='selected_student_id' value='{h(q_exam_student_id)}'></label>
            </div>
            <div class='btn-row'>
              <button>{t('common.search')}</button>
              <a class='btn secondary' href='/exams?lang={CURRENT_LANG}'>{t('common.reset')}</a>
            </div>
          </form>
        </div>
        <div class='card'><h4>{t("exams.list")}</h4>{'' if load_exams else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}<table><tr><th>{t("field.id")}</th><th>{t("field.class_id")}</th><th>{t("field.exam_name")}</th><th>{t("field.exam_date")}</th><th>{t("field.report")}</th></tr>{exam_rows if load_exams else ''}{(f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_exams and not exam_rows) else ''}</table></div>
        <div class='card'><h4>{t("exams.scores")}</h4>{'' if load_exams else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}<table><tr><th>{t("field.id")}</th><th>{t("field.exam_id")}</th><th>{t("field.student_id")}</th><th>{t("field.score")}</th></tr>{score_rows if load_exams else ''}{(f"<tr><td colspan='4' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_exams and not score_rows) else ''}</table></div>
        """, user, current_menu="exams", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/counseling":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        flash_msg = ""
        flash_type = "success"
        q_c_student_id = (query.get("student_id", "") or "").strip()
        q_c_parent_id = (query.get("parent_id", "") or "").strip()
        q_c_special = (query.get("is_special_note", "") or "").strip()
        q_c_selected_student_id = (query.get("selected_student_id", "") or "").strip()
        q_c_student_q = (query.get("student_q", "") or "").strip()
        q_c_parent_q = (query.get("parent_q", "") or "").strip()
        counseling_picker_query_enabled = query.get("do_search", "") == "1" or bool(q_c_student_q) or bool(q_c_parent_q) or bool(q_c_selected_student_id)
        student_candidates = fetch_student_candidates(conn, q_c_student_q, limit=10) if counseling_picker_query_enabled else []
        selected_student = conn.execute("SELECT id, user_id, student_no, name_ko FROM students WHERE id=?", (q_c_selected_student_id,)).fetchone() if q_c_selected_student_id.isdigit() else None
        q_c_student_user_id = str(selected_student["user_id"]) if selected_student else q_c_student_id
        parent_rows = conn.execute("SELECT id, name, username FROM users WHERE role='parent' AND (?='' OR name LIKE ? OR username LIKE ?) ORDER BY id DESC LIMIT 10", (q_c_parent_q, f"%{q_c_parent_q}%", f"%{q_c_parent_q}%")).fetchall() if counseling_picker_query_enabled else []
        parent_candidates = [{"id": r["id"], "label": f"{r['name']} ({r['username']})"} for r in parent_rows]
        selected_parent = conn.execute("SELECT id, name, username FROM users WHERE role='parent' AND id=?", (q_c_parent_id,)).fetchone() if q_c_parent_id.isdigit() else None
        load_counseling = query.get("load", "") == "1" or bool(q_c_student_user_id) or bool(q_c_parent_id) or bool(q_c_special)
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            errs = []
            student_input_id = d.get("student_id") or q_c_student_user_id
            if not ensure_exists(conn, "users", student_input_id, extra_where="role='student'"):
                add_error(errs, "student_id", "존재하지 않는 학생입니다")
            parent_input_id = d.get("parent_id") or q_c_parent_id
            if parent_input_id and not ensure_exists(conn, "users", parent_input_id, extra_where="role='parent'"):
                add_error(errs, "parent_id", "존재하지 않는 학부모입니다")
            if not (d.get("memo") or "").strip():
                add_error(errs, "memo", "필수값입니다")
            if errs:
                flash_msg = format_errors(errs)
                flash_type = "error"
                log_event(conn, "ERROR", path, "상담 저장 검증 실패", "\n".join(errs), user["id"])
            else:
                conn.execute("INSERT INTO counseling(student_id, parent_id, memo, is_special_note, created_by, created_at) VALUES(?,?,?,?,?,?)", (student_input_id, parent_input_id or None, d.get("memo"), 1 if d.get("is_special_note") else 0, user["id"], now()))
                conn.commit()
                flash_msg = "저장되었습니다"
            load_counseling = True

        rows = []
        if load_counseling:
            where = []
            params = []
            if str(q_c_student_user_id).isdigit():
                where.append("student_id=?")
                params.append(str(q_c_student_user_id))
            if q_c_parent_id.isdigit():
                where.append("parent_id=?")
                params.append(q_c_parent_id)
            if q_c_special in ("1", "0"):
                where.append("is_special_note=?")
                params.append(q_c_special)
            where_sql = (" WHERE " + " AND ".join(where)) if where else ""
            rows = conn.execute(f"SELECT * FROM counseling{where_sql} ORDER BY id DESC", tuple(params)).fetchall()

        student_picker = render_picker_block(t("picker.student"), "student_q", q_c_student_q, "selected_student_id", q_c_selected_student_id,
                                            (f"{selected_student['name_ko']} ({selected_student['student_no'] or '-'} / U:{selected_student['user_id']})" if selected_student else ""),
                                            student_candidates, "/counseling", CURRENT_LANG,
                                            {"parent_id": q_c_parent_id, "parent_q": q_c_parent_q, "load": "1", "is_special_note": q_c_special}, query_enabled=counseling_picker_query_enabled)
        parent_picker = render_picker_block(t("picker.parent"), "parent_q", q_c_parent_q, "parent_id", q_c_parent_id,
                                           (f"{selected_parent['name']} ({selected_parent['username']})" if selected_parent else ""),
                                           parent_candidates, "/counseling", CURRENT_LANG,
                                           {"selected_student_id": q_c_selected_student_id, "student_q": q_c_student_q, "load": "1", "is_special_note": q_c_special}, query_enabled=counseling_picker_query_enabled)
        table_rows = "".join([
            f"<tr><td>{r['id']}</td><td>{r['student_id']}</td><td>{r['parent_id'] or '-'}</td><td>{r['memo'] or '-'}</td><td>{'Y' if r['is_special_note'] else '-'}</td><td>{r['created_at'] or '-'}</td></tr>"
            for r in rows
        ])
        html = render_html(t("counseling.title"), f"""
        {student_picker}
        {parent_picker}
        <div class='card'>
          <h4>{t('common.search')}</h4>
          <form method='get' class='mobile-stack query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='load' value='1'>
            <div class='filter-grid'>
              <label>{t('counseling.student_id')} <input name='student_id' value='{h(q_c_student_user_id)}'></label>
              <label>{t('counseling.parent_id')} <input name='parent_id' value='{h(q_c_parent_id)}'></label>
              <label>{t('counseling.special')}
                <select name='is_special_note'>
                  <option value=''>{t('academics.day_all')}</option>
                  <option value='1' {'selected' if q_c_special=='1' else ''}>{t('common.yes')}</option>
                  <option value='0' {'selected' if q_c_special=='0' else ''}>{t('common.no')}</option>
                </select>
              </label>
            </div>
            <div class='btn-row'>
              <button>{t('common.search')}</button>
              <a class='btn secondary' href='/counseling?lang={CURRENT_LANG}'>{t('common.reset')}</a>
            </div>
          </form>
        </div>
        <div class='card'>
          <form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'>
            <label>{t('counseling.student_id')} <input name='student_id' value='{h(q_c_student_user_id)}' placeholder='User ID'></label>
            <label>{t('counseling.parent_id')} <input name='parent_id' value='{h(q_c_parent_id)}'></label>
            <label>{t('counseling.memo')} <input name='memo'></label>
            <label>{t('counseling.special')} <input type='checkbox' name='is_special_note' value='1'></label>
            <button>{t("common.save")}</button>
          </form>
        </div>
        <div class='card'>
          <h4>{t('counseling.list')}</h4>
          {'' if load_counseling else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}
          <table>
            <tr><th>ID</th><th>{t('counseling.student_id')}</th><th>{t('counseling.parent_id')}</th><th>{t('counseling.memo')}</th><th>{t('counseling.special')}</th><th>created_at</th></tr>
            {table_rows if load_counseling else ''}
            {(f"<tr><td colspan='6' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_counseling and not table_rows) else ''}
          </table>
        </div>
        """, user, current_menu="counseling", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/payments":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_PARENT, ROLE_STUDENT]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        selected_student_id = (query.get("selected_student_id", "") or "").strip()
        student_query_enabled = query.get("do_search", "") == "1" or bool((query.get("student_q", "") or "").strip()) or bool(selected_student_id)
        load_payments = query.get("load", "") == "1" or bool(selected_student_id)
        student_candidates = fetch_student_candidates(conn, query.get("student_q", ""), limit=10) if student_query_enabled else []
        selected_student = conn.execute("SELECT id, name_ko, student_no FROM students WHERE id=?", (selected_student_id,)).fetchone() if selected_student_id else None
        flash_msg = ""
        flash_type = "success"
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER]):
            d = parse_body(environ)
            student_input_id = d.get("student_id") or selected_student_id
            student_row = conn.execute("SELECT user_id FROM students WHERE id=?", (student_input_id,)).fetchone() if str(student_input_id).isdigit() else None
            student_user_id = student_row["user_id"] if student_row else student_input_id
            errs = []
            amount_v = as_float(d.get("amount"))
            hours_v = as_float(d.get("package_hours") or 0)
            remain_v = as_int(d.get("remaining_classes") or 0)
            if not ensure_exists(conn, "users", student_user_id, extra_where="role='student'"):
                add_error(errs, "student_id", "존재하지 않는 학생입니다")
            if not is_valid_date(d.get("paid_date") or ""):
                add_error(errs, "paid_date", "YYYY-MM-DD 형식이어야 합니다")
            if amount_v is None or amount_v < 0:
                add_error(errs, "amount", "0 이상의 숫자여야 합니다")
            if hours_v is None or hours_v < 0:
                add_error(errs, "package_hours", "0 이상의 숫자여야 합니다")
            if remain_v is None or remain_v < 0:
                add_error(errs, "remaining_classes", "0 이상의 정수여야 합니다")
            if errs:
                flash_msg = format_errors(errs)
                flash_type = "error"
                log_event(conn, "ERROR", path, "수납 저장 검증 실패", "\n".join(errs), user["id"])
            else:
                conn.execute("INSERT INTO payments(student_id, paid_date, amount, package_hours, remaining_classes, created_at) VALUES(?,?,?,?,?,?)", (student_user_id, d.get("paid_date"), amount_v, hours_v, remain_v, now()))
                conn.commit()
                flash_msg = "저장되었습니다"
            load_payments = True

        rows = []
        if load_payments:
            if has_role(user, [ROLE_STUDENT]):
                rows = conn.execute("SELECT * FROM payments WHERE student_id=? ORDER BY id DESC", (user["id"],)).fetchall()
            elif has_role(user, [ROLE_PARENT]):
                rows = conn.execute("SELECT p.* FROM payments p JOIN students s ON s.user_id=p.student_id WHERE s.guardian_name=? ORDER BY p.id DESC", (user["name"],)).fetchall()
            else:
                if selected_student_id and selected_student_id.isdigit():
                    selected_user = conn.execute("SELECT user_id FROM students WHERE id=?", (selected_student_id,)).fetchone()
                    if selected_user:
                        rows = conn.execute("SELECT * FROM payments WHERE student_id=? ORDER BY id DESC", (selected_user["user_id"],)).fetchall()
                    else:
                        rows = []
                else:
                    rows = conn.execute("SELECT * FROM payments ORDER BY id DESC").fetchall()

        student_picker = render_picker_block(t("picker.student"), "student_q", query.get("student_q", ""), "selected_student_id", selected_student_id,
                                            (f"{selected_student['name_ko']} ({selected_student['student_no'] or '-'})" if selected_student else ""),
                                            student_candidates, "/payments", CURRENT_LANG, {"load": "1"}, query_enabled=student_query_enabled)
        row_html = "".join([f"<tr><td>{r['id']}</td><td>{r['student_id']}</td><td>{r['paid_date']}</td><td>{r['amount']}</td><td>{r['package_hours']}</td><td>{r['remaining_classes']}</td></tr>" for r in rows])
        html = render_html(t("payments.title"), f"""
        {student_picker if has_role(user, [ROLE_OWNER, ROLE_MANAGER]) else ''}
        <div class='card'>
          <form method='get' class='btn-row query-form'>
            <input type='hidden' name='lang' value='{CURRENT_LANG}'>
            <input type='hidden' name='load' value='1'>
            <input type='hidden' name='selected_student_id' value='{selected_student_id}'>
            <button>{t('common.search')}</button>
            <a class='btn secondary admin-action-link' data-preserve-scroll='1' href='/payments?lang={CURRENT_LANG}'>{t('common.reset')}</a>
          </form>
        </div>
        <div class='card'>
          <h4>{t("payments.input")}</h4>
          <form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'>
            <input type='hidden' name='student_id' value='{selected_student_id}'>
            <label>{t('common.selected')} <input value='{(selected_student['name_ko'] if selected_student else '-')}' readonly></label>
            <label>{t('field.student_id')} <input value='{selected_student_id}' readonly></label>
            <label>{t("field.paid_date")} <input name='paid_date' placeholder='2026-03-06'></label>
            <label>{t("field.amount")} <input name='amount'></label>
            <label>{t("field.package_hours")} <input name='package_hours'></label>
            <label>{t("field.remaining_classes")} <input name='remaining_classes'></label>
            <button>{t('common.save')}</button>
          </form>
        </div>
        <div class='card'><h4>{t("payments.list")}</h4>{'' if load_payments else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}<table><tr><th>{t("field.id")}</th><th>{t("field.student_id")}</th><th>{t("field.paid_date")}</th><th>{t("field.amount")}</th><th>{t("field.package_hours")}</th><th>{t("field.remaining_classes")}</th></tr>{row_html if load_payments else ''}{(f"<tr><td colspan='6' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_payments and not row_html) else ''}</table></div>
        """, user, current_menu="payments", flash_msg=flash_msg, flash_type=flash_type)
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if False and path == "/announcements":  # moved to readingtown.routes.notifications
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            conn.execute("INSERT INTO announcements(title, content, created_by, created_at) VALUES(?,?,?,?)", (d.get("title"), d.get("content"), user["id"], now()))
            conn.commit()
        rows = conn.execute("SELECT * FROM announcements ORDER BY id DESC").fetchall()
        noti = conn.execute("SELECT * FROM notifications ORDER BY id DESC LIMIT 50").fetchall()
        ann_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['title']}</td><td>{r['content']}</td><td>{r['created_by']}</td><td>{r['created_at']}</td></tr>" for r in rows])
        noti_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['type']}</td><td>{r['target_user_id'] or '-'}</td><td>{r['payload']}</td><td>{r['created_at']}</td></tr>" for r in noti])
        html = render_html(t("ann.title"), f"""
        <div class='card'><h4>{t("ann.write")}</h4><form method='post' class='form-row'>{t("field.title")}<input name='title'> {t('field.content')}<input name='content'><button>{t('common.save')}</button></form></div>
        <div class='card'><h4>{t("ann.list")}</h4><table><tr><th>{t("field.id")}</th><th>{t("field.title")}</th><th>{t("field.content")}</th><th>{t("field.writer")}</th><th>{t("field.created_at")}</th></tr>{ann_rows or f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>"}</table></div>
        <div class='card'><h4>{t("ann.noti")}</h4><table><tr><th>{t("field.id")}</th><th>{t("field.type")}</th><th>{t("field.target")}</th><th>{t("field.data")}</th><th>{t("field.created_at")}</th></tr>{noti_rows or f"<tr><td colspan='5' class='empty-msg'>{t('common.no_data')}</td></tr>"}</table></div>
        """, user, current_menu="announcements")
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if path == "/library":
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        selected_student_id = (query.get("selected_student_id", "") or "").strip()
        selected_teacher_id = (query.get("selected_teacher_id", "") or "").strip()
        picker_query_enabled = query.get("do_search", "") == "1" or bool((query.get("student_q", "") or "").strip()) or bool((query.get("teacher_q", "") or "").strip()) or bool(selected_student_id) or bool(selected_teacher_id)
        load_library = query.get("load", "") == "1" or bool(selected_student_id) or bool(selected_teacher_id)
        student_candidates = fetch_student_candidates(conn, query.get("student_q", ""), limit=10) if picker_query_enabled else []
        teacher_candidates = fetch_teacher_candidates(conn, query.get("teacher_q", ""), limit=10) if picker_query_enabled else []
        selected_student = conn.execute("SELECT id, name_ko, student_no FROM students WHERE id=?", (selected_student_id,)).fetchone() if selected_student_id else None
        selected_teacher = fetch_teacher_by_id(conn, selected_teacher_id) if selected_teacher_id else None
        if method == "POST" and has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            d = parse_body(environ)
            typ = d.get("type")
            if typ == "book":
                conn.execute("INSERT INTO books(code, title, status, created_at) VALUES(?,?,?,?)", (d.get("code"), d.get("title"), "available", now()))
            elif typ == "loan":
                book = conn.execute("SELECT * FROM books WHERE code=?", (d.get("code"),)).fetchone()
                if book:
                    student_input_id = d.get("student_id") or selected_student_id
                    student_row = conn.execute("SELECT user_id FROM students WHERE id=?", (student_input_id,)).fetchone() if str(student_input_id).isdigit() else None
                    if student_row:
                        conn.execute("UPDATE books SET status='borrowed' WHERE id=?", (book["id"],))
                        conn.execute("INSERT INTO book_loans(book_id, student_id, loaned_at, handled_by, created_at) VALUES(?,?,?,?,?)", (book["id"], student_row["user_id"], now(), d.get("teacher_id") or selected_teacher_id or user["id"], now()))
                    else:
                        log_event(conn, "ERROR", path, "도서 대출 실패", f"invalid student_id={student_input_id}", user["id"])
            elif typ == "return":
                book = conn.execute("SELECT * FROM books WHERE code=?", (d.get("code"),)).fetchone()
                if book:
                    conn.execute("UPDATE books SET status='available' WHERE id=?", (book["id"],))
                    conn.execute("UPDATE book_loans SET returned_at=? WHERE book_id=? AND returned_at IS NULL", (now(), book["id"]))
            conn.commit()
            load_library = True
        books = conn.execute("SELECT * FROM books ORDER BY id DESC").fetchall() if load_library else []
        loans = conn.execute("SELECT * FROM book_loans ORDER BY id DESC").fetchall() if load_library else []
        student_picker = render_picker_block(t("picker.student"), "student_q", query.get("student_q", ""), "selected_student_id", selected_student_id,
                                            (f"{selected_student['name_ko']} ({selected_student['student_no'] or '-'})" if selected_student else ""),
                                            student_candidates, "/library", CURRENT_LANG,
                                            {"selected_teacher_id": selected_teacher_id, "teacher_q": query.get("teacher_q", ""), "load": "1"}, query_enabled=picker_query_enabled)
        teacher_picker = render_picker_block(t("picker.teacher"), "teacher_q", query.get("teacher_q", ""), "selected_teacher_id", selected_teacher_id,
                                            (f"{selected_teacher['name']} ({selected_teacher['username']})" if selected_teacher else ""),
                                            teacher_candidates, "/library", CURRENT_LANG,
                                            {"selected_student_id": selected_student_id, "student_q": query.get("student_q", ""), "load": "1"}, query_enabled=picker_query_enabled)
        book_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['code']}</td><td>{r['title']}</td><td>{r['status']}</td></tr>" for r in books])
        loan_rows = "".join([f"<tr><td>{r['id']}</td><td>{r['book_id']}</td><td>{r['student_id']}</td><td>{r['loaned_at']}</td><td>{r['returned_at'] or '-'}</td><td>{r['handled_by']}</td></tr>" for r in loans])
        html = render_html(t("library.title"), f"""
        {student_picker}
        {teacher_picker}
        <div class='card'><h4>{t("library.book_add")}</h4><form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'><input type='hidden' name='type' value='book'>{t("field.code")}<input name='code'> {t("field.title")}<input name='title'><button>{t('common.save')}</button></form></div>
        <div class='card'><h4>{t("library.loan")}</h4><form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'><input type='hidden' name='type' value='loan'><input type='hidden' name='student_id' value='{selected_student_id}'><input type='hidden' name='teacher_id' value='{selected_teacher_id}'>{t("field.code")}<input name='code'> {t("common.selected")}<input value='{(selected_student["name_ko"] if selected_student else "-")}' readonly> {t("field.student_id")}<input value='{selected_student_id}' readonly> {t("field.teacher_id")}<input value='{selected_teacher_id}' readonly><button>{t('common.save')}</button></form></div>
        <div class='card'><h4>{t("library.return")}</h4><form method='post' class='form-row preserve-scroll-form' data-preserve-scroll='1'><input type='hidden' name='type' value='return'>{t("field.code")}<input name='code'><button>{t('common.save')}</button></form></div>
        <div class='card'><h4>{t("library.books")}</h4>{'' if load_library else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}<table><tr><th>{t("field.id")}</th><th>{t("field.code")}</th><th>{t("field.title")}</th><th>{t("field.status")}</th></tr>{book_rows if load_library else ''}{(f"<tr><td colspan='4' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_library and not book_rows) else ''}</table></div>
        <div class='card'><h4>{t("library.history")}</h4>{'' if load_library else ("<div class='empty-msg'>" + t('common.query_to_load') + "</div>")}<table><tr><th>{t("field.id")}</th><th>{t("field.book_id")}</th><th>{t("field.student_id")}</th><th>{t("field.loaned_at")}</th><th>{t("field.returned_at")}</th><th>{t("field.handler")}</th></tr>{loan_rows if load_library else ''}{(f"<tr><td colspan='6' class='empty-msg'>{t('common.no_data')}</td></tr>") if (load_library and not loan_rows) else ''}</table></div>
        """, user, current_menu="library")
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if False and path == "/logs":  # moved to readingtown.routes.logs
        if not has_role(user, [ROLE_OWNER]):
            conn.close()
            status, headers, body = forbidden_html(user)
            start_response(status, headers)
            return [body]
        ensure_logs_table(conn)
        ensure_logs_columns(conn)
        rows = conn.execute("SELECT id, level, route, user_id, message, detail, created_at FROM app_logs ORDER BY id DESC LIMIT 300").fetchall()
        row_html = "".join([
            f"<tr><td>{r['id']}</td><td>{r['level']}</td><td>{r['route'] or '-'}</td><td>{r['user_id'] or '-'}</td><td>{r['message']}</td><td>{(r['detail'] or '-')}</td><td>{r['created_at']}</td></tr>"
            for r in rows
        ])
        html = render_html(menu_t("logs"), f"""
        <div class='card'>
          <h4>{menu_t('logs')}</h4>
          <table>
            <tr><th>ID</th><th>level</th><th>route</th><th>user_id</th><th>message</th><th>detail</th><th>created_at</th></tr>
            {row_html or f"<tr><td colspan='7' class='empty-msg'>{t('common.no_data')}</td></tr>"}
          </table>
        </div>
        """, user, current_menu="logs")
        status, headers, body = text_resp(html)
        conn.close()
        start_response(status, headers)
        return [body]
    if False and path == "/api/announcements" and method == "GET":  # moved to readingtown.routes.api
        rows = conn.execute("SELECT * FROM announcements ORDER BY id DESC").fetchall()
        conn.close()
        status, headers, body = json_resp([dict(r) for r in rows])
        start_response(status, headers)
        return [body]
    if path == "/api/books/loan-by-code" and method == "POST":
        d = parse_body(environ)
        if not has_role(user, [ROLE_OWNER, ROLE_MANAGER, ROLE_TEACHER]):
            conn.close()
            status, headers, body = forbidden_json()
            start_response(status, headers)
            return [body]
        book = conn.execute("SELECT * FROM books WHERE code=?", (d.get("code"),)).fetchone()
        if not book:
            conn.close()
            status, headers, body = json_resp({"error": t('library.not_found')}, "404 Not Found")
            start_response(status, headers)
            return [body]
        student_input_id = d.get("student_id")
        student_row = conn.execute("SELECT user_id FROM students WHERE id=?", (student_input_id,)).fetchone() if str(student_input_id).isdigit() else None
        if not student_row:
            conn.close()
            status, headers, body = json_resp({"error": "student_id: 존재하지 않는 학생입니다"}, "400 Bad Request")
            start_response(status, headers)
            return [body]
        conn.execute("UPDATE books SET status='borrowed' WHERE id=?", (book["id"],))
        conn.execute("INSERT INTO book_loans(book_id, student_id, loaned_at, handled_by, created_at) VALUES(?,?,?,?,?)", (book["id"], student_row["user_id"], now(), user["id"], now()))
        conn.commit()
        conn.close()
        status, headers, body = json_resp({"message": t('library.loan_done')})
        start_response(status, headers)
        return [body]
    conn.close()
    start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
    return ["Not Found".encode("utf-8")]
if __name__ == "__main__":
    init_db()
    with make_server("0.0.0.0", 8000, app) as httpd:
        print(t('server.start') + ': http://127.0.0.1:8000')
        httpd.serve_forever()









